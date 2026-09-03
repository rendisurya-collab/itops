"""
Query Executor module untuk execute query SELECT dan process hasil (teks/Excel).
"""
import io
import logging
from typing import Dict, Tuple
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logger = logging.getLogger(__name__)


class QueryExecutor:
    """
    Execute query SELECT dan process hasil data.
    """

    def __init__(self, db_connection=None):
        """
        Initialize QueryExecutor.

        Args:
            db_connection: Database connection object (dari servicedesk_client atau jira_client)
        """
        self.db_connection = db_connection

    def execute_select(self, query: str, db_name: str | None = None) -> Tuple[bool, list, str]:
        """
        Execute query SELECT.

        Args:
            query: SQL query string (harus SELECT)
            db_name: Database name (opsional, untuk DatabaseConnector)

        Return:
            (success: bool, rows: list of tuples, error: str)
        """
        if not query:
            return False, [], "Query kosong"

        query_lower = query.strip().lower()
        if not query_lower.startswith("select"):
            return False, [], "Query harus berupa SELECT statement"

        try:
            if not self.db_connection:
                return False, [], "Database connection belum dikonfigurasi"

            # Check apakah db_connection adalah DatabaseConnector atau legacy connection
            from db_connector import DatabaseConnector
            
            if isinstance(self.db_connection, DatabaseConnector):
                # Use DatabaseConnector (support multiple databases)
                if not db_name:
                    # Fallback: ambil database pertama
                    databases = self.db_connection.config.get('databases', [])
                    if not databases:
                        return False, [], "Tidak ada database di config"
                    db_name = databases[0].get('name')
                
                success, rows, error = self.db_connection.execute_query(db_name, query)
                return success, rows, error
            else:
                # Legacy: assume connection punya method execute()
                rows = self.db_connection.execute(query)

            if not rows:
                return True, [], None  # Success tapi empty result

            logger.info(f"Query executed: {len(rows)} rows returned")
            return True, rows, None

        except Exception as e:
            error_msg = f"Query execution error: {str(e)}"
            logger.error(error_msg)
            return False, [], error_msg

    def format_rows_as_text(self, rows: list, max_chars: int = 3500) -> Tuple[str, bool]:
        """
        Format hasil query sebagai teks plain untuk Telegram.

        Args:
            rows: List of tuples (hasil query)
            max_chars: Max character limit untuk teks

        Return:
            (formatted_text, exceeded_limit: bool)
        """
        if not rows:
            return "Tidak ada data", False

        try:
            lines = []
            for i, row in enumerate(rows, 1):
                # Convert tuple to readable string
                if isinstance(row, dict):
                    # Jika hasil berbentuk dict
                    row_str = " | ".join(f"{k}: {v}" for k, v in row.items())
                else:
                    # Jika hasil berbentuk tuple
                    row_str = " | ".join(str(v) for v in row)

                lines.append(f"{i}. {row_str}")

            text = "\n".join(lines)
            exceeded = len(text) > max_chars

            if exceeded:
                # Truncate dengan indikasi
                text = text[:max_chars] + f"\n\n... (Teks terlalu panjang, total {len(text)} chars. Lihat file Excel untuk data lengkap)"

            return text, exceeded

        except Exception as e:
            logger.error(f"Error formatting rows: {e}")
            return f"Error formatting data: {str(e)}", True

    def export_rows_to_excel(self, rows: list, query_name: str = "Query Result") -> Tuple[bool, io.BytesIO, str]:
        """
        Export hasil query ke file Excel (.xlsx) dalam BytesIO buffer.

        Args:
            rows: List of tuples (hasil query)
            query_name: Nama sheet / header

        Return:
            (success: bool, buffer: io.BytesIO, error: str)
        """
        if not HAS_OPENPYXL:
            return False, None, "openpyxl library not installed"

        if not rows:
            return False, None, "Tidak ada data untuk di-export"

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = query_name[:31]  # Excel sheet name max 31 chars

            # Header styling
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Add header row (ambil dari first row jika ada)
            if rows:
                first_row = rows[0]
                num_cols = len(first_row) if isinstance(first_row, (tuple, list)) else 1

                for col_idx in range(1, num_cols + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = f"Column {col_idx}"
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Add data rows
            for row_idx, row in enumerate(rows, 2):
                if isinstance(row, (tuple, list)):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                elif isinstance(row, dict):
                    for col_idx, (key, value) in enumerate(row.items(), 1):
                        if row_idx == 2:  # Set header untuk dict rows
                            header_cell = ws.cell(row=1, column=col_idx)
                            header_cell.value = key
                            header_cell.fill = header_fill
                            header_cell.font = header_font
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = value

            # Auto-fit columns
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Write to BytesIO
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            logger.info(f"Excel exported: {len(rows)} rows, size={buffer.getbuffer().nbytes} bytes")
            return True, buffer, None

        except Exception as e:
            error_msg = f"Excel export error: {str(e)}"
            logger.error(error_msg)
            return False, None, error_msg

    def process_query_result(
        self,
        rows: list,
        query_name: str = "Query Result",
        max_text_chars: int = 3500
    ) -> Tuple[str, bytes | None, str]:
        """
        Process hasil query dan tentukan format output (teks atau Excel).

        Workflow:
        1. Jika rows kosong -> return pesan "no data", tanpa export
        2. Jika rows > 0 dan teks <= max_text_chars -> return teks saja
        3. Jika rows > 0 dan teks > max_text_chars -> return teks + export Excel

        Args:
            rows: List of tuples
            query_name: Nama query/sheet
            max_text_chars: Max character untuk output teks

        Return:
            (text_message, excel_bytes, error_message)
        """
        # Check if empty
        if not rows:
            return "Tidak ada data untuk ditampilkan", None, None

        # Format as text
        text, exceeded = self.format_rows_as_text(rows, max_text_chars)

        # If text exceeded, export to Excel
        if exceeded:
            success, buffer, error = self.export_rows_to_excel(rows, query_name)
            if success:
                excel_bytes = buffer.getvalue()
                return text, excel_bytes, None
            else:
                # Fall back to text saja jika Excel export gagal
                logger.warning(f"Excel export failed: {error}, fallback to text")
                return text, None, error

        return text, None, None
