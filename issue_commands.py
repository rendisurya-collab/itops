"""
Issue Commands Handler - /noteissue dan /listissue commands untuk Telegram bot.
Terintegrasi dengan IssueLogger untuk Google Sheets operations.
"""

import logging
from telegram import Update
from telegram.ext import ContextTypes
from telegram.constants import ParseMode
from issue_logger import IssueLogger
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Global IssueLogger instance
issue_logger = None


def init_issue_logger():
    """Initialize global IssueLogger instance."""
    global issue_logger
    issue_logger = IssueLogger()
    return issue_logger.worksheet is not None


async def handle_noteissue(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handler untuk /noteissue command.
    Pencatatan issue baru ke Google Sheets.
    
    Format input:
    /noteissue
    Source: Web App / Mobile / POS
    Kendala: Detail issue di sini
    Action: Action resolved di sini
    """
    try:
        # Get user input (everything after /noteissue)
        if not context.args:
            # User hanya kirim /noteissue tanpa text
            # Minta user mengirim kembali dengan format
            help_text = r"""
❌ **Format /noteissue salah!**

Silakan kirim dengan format:

```
/noteissue
Source: Web App / Mobile / POS
Ticket Number: TIK-123456 (opsional)
Kendala: Detail issue yang terjadi
Action: Action untuk resolve issue
```

**Contoh:**
```
/noteissue
Source: Web App
Ticket Number: TIK-2026-001
Kendala: POS Digital tidak bisa payment setelah update
Action: Clear cache browser & restart service
```
"""
            await update.message.reply_text(help_text, parse_mode=ParseMode.MARKDOWN)
            return
        
        # Get full message text (exclude /noteissue command)
        message_text = update.message.text
        if message_text.startswith('/noteissue'):
            # Remove /noteissue from text
            input_text = message_text[len('/noteissue'):].strip()
        else:
            input_text = ' '.join(context.args)
        
        # Parse input
        success, parsed = issue_logger.parse_noteissue_input(input_text)
        if not success:
            error_msg = parsed.get('error', 'Error parsing input')
            await update.message.reply_text(
                f"❌ **Error:** {error_msg}\n\nSilakan cek format pesan Anda.",
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Extract data
        source = parsed['source']
        ticket_number = parsed.get('ticket_number', '')
        detail_issue = parsed['detail_issue']
        action_resolved = parsed['action_resolved']
        chat_id = str(update.effective_chat.id)
        
        # Get username from Telegram user object
        username = update.effective_user.username or update.effective_user.first_name or "Unknown"
        
        # Append to Google Sheets (now with username and ticket_number)
        append_success, timestamp = issue_logger.append_issue(
            chat_id=chat_id,
            username=username,
            ticket_number=ticket_number,
            source=source,
            detail_issue=detail_issue,
            action_resolved=action_resolved
        )
        
        if not append_success:
            await update.message.reply_text(
                f"❌ **Gagal mencatat issue!**\n\n{timestamp}",
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Success response
        response = f"""✅ **Issue Berhasil Dicatat!**

• **Tanggal:** {timestamp}
• **Username:** {username}
• **Ticket Number:** {ticket_number if ticket_number else '-'}
• **Source:** {source}
• **Detail Issue:** {detail_issue}
• **Action Resolved:** {action_resolved}
"""
        await update.message.reply_text(response, parse_mode=ParseMode.MARKDOWN)
        logger.info(f"Issue recorded: chat_id={chat_id}, username={username}, ticket={ticket_number}, source={source}")
        
    except Exception as e:
        logger.error(f"Error in handle_noteissue: {e}")
        await update.message.reply_text(
            f"❌ **Error:** {str(e)}",
            parse_mode=ParseMode.MARKDOWN
        )


async def handle_listissue(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    Handler untuk /listissue command.
    Menampilkan daftar semua issue dari Google Sheets.
    
    Dual output:
    - Jika text <= 3500 chars: kirim sebagai text
    - Jika text > 3500 chars: kirim sebagai Excel file
    """
    try:
        # Show loading message
        loading_msg = await update.message.reply_text("⏳ Mengambil data issue...")
        
        # Fetch issues from Google Sheets
        fetch_success, rows, error_msg = issue_logger.fetch_all_issues()
        
        if not fetch_success:
            await loading_msg.delete()
            await update.message.reply_text(
                f"❌ **Error:** {error_msg}",
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        if not rows:
            await loading_msg.delete()
            await update.message.reply_text(
                "📭 Tidak ada issue yang tercatat.",
                parse_mode=ParseMode.MARKDOWN
            )
            return
        
        # Format as text
        text_output = format_issues_as_text(rows)
        
        # Check length threshold
        if len(text_output) <= 3500:
            # Send as text
            await loading_msg.delete()
            await update.message.reply_text(
                text_output,
                parse_mode=ParseMode.MARKDOWN
            )
            logger.info(f"Sent {len(rows)} issues as text ({len(text_output)} chars)")
            
        else:
            # Convert to Excel and send as file
            excel_buffer = create_excel_from_issues(rows)
            
            await loading_msg.delete()
            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=excel_buffer,
                filename="issue_logs.xlsx",
                caption="📄 Daftar Issue Log terlampir dalam file Excel karena jumlah data melebihi batas karakter."
            )
            logger.info(f"Sent {len(rows)} issues as Excel file")
            
    except Exception as e:
        logger.error(f"Error in handle_listissue: {e}")
        await update.message.reply_text(
            f"❌ **Error:** {str(e)}",
            parse_mode=ParseMode.MARKDOWN
        )


def format_issues_as_text(rows: list) -> str:
    """
    Format issues sebagai readable text untuk Telegram.
    
    Args:
        rows: List of lists dari Google Sheets
        
    Return:
        str: Formatted text
    """
    lines = []
    lines.append("DAFTAR ISSUE LOG\n")
    
    for idx, row in enumerate(rows, 1):
        if len(row) >= 7:
            timestamp = row[0]
            chat_id = row[1]
            username = row[2]
            ticket_number = row[3]
            source = row[4]
            detail = row[5]
            action = row[6]
            
            # Truncate long values
            detail = detail[:80] + "..." if len(detail) > 80 else detail
            action = action[:80] + "..." if len(action) > 80 else action
            
            lines.append(f"{idx}. {source} ({timestamp})")
            lines.append(f"   reporter: {username}")
            if ticket_number and ticket_number != "-":
                lines.append(f"   ticketnumber: {ticket_number}")
            lines.append(f"   detail issue: {detail}")
            lines.append(f"   action resolved: {action}")
            lines.append("")
    
    # Wrap in code block for monospace
    text = "\n".join(lines)
    return f"```\n{text}\n```"


def create_excel_from_issues(rows: list) -> BytesIO:
    """
    Create Excel file from issues data.
    
    Args:
        rows: List of lists dari Google Sheets
        
    Return:
        BytesIO: Excel file buffer
    """
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IssueLogs"
    
    # Add headers (now with Reporter and Ticket Number)
    headers = ['Tanggal Issue', 'Chat ID', 'Reporter', 'Ticket Number', 'Source', 'Detail Issue', 'Action Resolved']
    ws.append(headers)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add data rows
    for row in rows:
        ws.append(row)
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Tanggal
    ws.column_dimensions['B'].width = 12  # Chat ID
    ws.column_dimensions['C'].width = 15  # Username
    ws.column_dimensions['D'].width = 15  # Ticket Number
    ws.column_dimensions['E'].width = 15  # Source
    ws.column_dimensions['F'].width = 40  # Detail Issue
    ws.column_dimensions['G'].width = 40  # Action Resolved
    
    # Wrap text for columns F dan G
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
