import asyncio
import datetime as dt
import json
import logging
import os
import re
import subprocess
import tempfile
from zoneinfo import ZoneInfo

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ParseMode
from telegram.error import NetworkError, TimedOut
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

import config
import guidance_store
import user_accounts
from export_excel import build_export_excel
from jira_client import JiraClient, JiraError
from servicedesk_client import SDPClient, SDPError

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

jira = JiraClient()  # dipakai untuk fitur yang tidak butuh atribusi personal (/tasks, /export)
sdp = SDPClient() if config.sdp_configured() else None

_JIRA_CLIENT_CACHE = {}  # telegram_user_id (str) -> JiraClient


# ==============================================================================
# KONFIGURASI FILE & LOGIKA SHIFT / ROUND-ROBIN / AUDIT LOG SERVICEDESK PLUS
# ==============================================================================
TECH_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "technicians.json")
SDP_ASSIGN_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sdp_assign_state.json")
TICKET_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ticket_activity.log")

SHIFT_MORNING_START = getattr(config, "SHIFT_MORNING_START", "07:00")
SHIFT_NIGHT_START = getattr(config, "SHIFT_NIGHT_START", "19:00")


def write_ticket_log(ticket_id: str, tech_name: str, tech_email: str, status: str, shift: str, is_success: bool, error_msg: str = ""):
    """Menyimpan setiap riwayat aksi update tiket ke dalam file ticket_activity.log"""
    timestamp = dt.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
    status_tag = "SUCCESS" if is_success else "FAILED"
    
    log_line = (
        f"[{timestamp}] [{status_tag}] Ticket ID: #{ticket_id} | "
        f"Assigned To: {tech_name} ({tech_email}) | Status: {status} | Shift: {shift}"
    )
    if not is_success and error_msg:
        log_line += f" | Error: {error_msg}"
        
    try:
        with open(TICKET_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
    except Exception as e:
        logger.error(f"Gagal menulis ke file log tiket: {e}")


def load_technicians() -> list:
    """Membaca daftar 4 teknisi dari file JSON"""
    if not os.path.exists(TECH_FILE):
        logger.warning(f"File {TECH_FILE} tidak ditemukan. Otomasi pembagian tiket tidak dapat berjalan.")
        return []
    try:
        with open(TECH_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Gagal membaca file {TECH_FILE}: {e}")
        return []


def get_current_shift() -> str:
    """Menentukan shift aktif dari 4 jadwal shift"""
    now = dt.datetime.now(TZ).time()
    
# GANTI BAGIAN LAMA DI ATAS DENGAN 1 BARIS INI:

import openpyxl
import datetime as dt
import pytz

def get_current_shift_from_excel(excel_path: str = "jadwal_shift.xlsx") -> dict:
    now_dt = dt.datetime.now()
    today_str = now_dt.strftime("%Y-%m-%d")
    current_time = now_dt.time()
    TZ = pytz.timezone("Asia/Jakarta")

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue

            tgl_val, jam_mulai_val, jam_selesai_val, nama_shift, teknisi = row[0], row[1], row[2], row[3], row[4]

            if isinstance(tgl_val, (dt.date, dt.datetime)):
                tgl_str = tgl_val.strftime("%Y-%m-%d")
            else:
                tgl_str = str(tgl_val).strip()

            if tgl_str == today_str:
                t_start = jam_mulai_val if isinstance(jam_mulai_val, dt.time) else dt.datetime.strptime(str(jam_mulai_val).strip(), "%H:%M").time()
                t_end = jam_selesai_val if isinstance(jam_selesai_val, dt.time) else dt.datetime.strptime(str(jam_selesai_val).strip(), "%H:%M").time()

                if t_start <= current_time < t_end:
                    return {
                        "shift": str(nama_shift).strip(),
                        "technician": str(teknisi).strip(),
                    }

    except Exception as e:
        logger.error(f"Gagal membaca file Excel shift: {e}")

    return {"shift": "OFF_SHIFT", "technician": None}


# 2. BARU PANGGIL FUNGSI TERSEBUT DI BAWAHNYA (misal di dalam auto_assign_sdp_tickets)
async def auto_assign_sdp_tickets(context):
    shift_info = get_current_shift_from_excel("jadwal_shift.xlsx")
    # ...
    
shift_info = get_current_shift_from_excel("jadwal_shift.xlsx")

shift_name = shift_info["shift"]         # Contoh hasil: "SHIFT_1"
tech_on_duty = shift_info["technician"] # Contoh hasil: "Ahmad" atau "Ahmad, Budi"

if tech_on_duty:
    # Lanjutkan logika Round Robin & Auto Assign ke SDP
    pass

def load_assign_state() -> dict:
    """Membaca indeks giliran teknisi terakhir dari file JSON"""
    try:
        with open(SDP_ASSIGN_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_assigned": {"SHIFT_1": -1, "SHIFT_2": -1, "SHIFT_3": -1, "SHIFT_4": -1}}

def save_assign_state(state: dict):
    """Menyimpan indeks giliran teknisi terbaru"""
    with open(SDP_ASSIGN_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)


def select_technician_round_robin(shift: str) -> dict:
    """Mengambil teknisi berikutnya secara bergantian dan adil di shift yang sama"""
    all_techs = load_technicians()
    on_duty_techs = [t for t in all_techs if t.get("shift") == shift]

    if not on_duty_techs:
        raise ValueError(f"Tidak ada teknisi yang terdaftar untuk shift {shift}")

    state = load_assign_state()
    last_index = state.get("last_assigned", {}).get(shift, -1)

    next_index = (last_index + 1) % len(on_duty_techs)

    if "last_assigned" not in state:
        state["last_assigned"] = {}
    state["last_assigned"][shift] = next_index
    save_assign_state(state)

    selected_tech = on_duty_techs[next_index]
    logger.info(f"🔄 [ROUND-ROBIN] Ticket dialokasikan ke: {selected_tech['name']} (Shift {shift})")
    return selected_tech

def get_current_shift_from_excel(excel_path: str = "jadwal_shift.xlsx") -> dict:
    now_dt = dt.datetime.now(TZ)
    today_str = now_dt.strftime("%Y-%m-%d")
    current_time = now_dt.time()

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:  # Memastikan ada 5 kolom
                continue

            tgl_val, jam_mulai_val, jam_selesai_val, nama_shift, teknisi = (
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
            )

            # Format Tanggal
            if isinstance(tgl_val, (dt.date, dt.datetime)):
                tgl_str = tgl_val.strftime("%Y-%m-%d")
            else:
                tgl_str = str(tgl_val).strip()

            # Pengecekan Tanggal Hari Ini
            if tgl_str == today_str:
                # Format Jam Mulai & Selesai
                t_start = (
                    jam_mulai_val
                    if isinstance(jam_mulai_val, dt.time)
                    else dt.datetime.strptime(
                        str(jam_mulai_val).strip(), "%H:%M"
                    ).time()
                )
                t_end = (
                    jam_selesai_val
                    if isinstance(jam_selesai_val, dt.time)
                    else dt.datetime.strptime(
                        str(jam_selesai_val).strip(), "%H:%M"
                    ).time()
                )

                # Cek apakah jam sekarang masuk rentang shift
                if t_start <= current_time < t_end:
                    return {
                        "shift": str(nama_shift).strip(),
                        "technician": str(teknisi).strip(),
                    }

    except Exception as e:
        logger.error(f"Gagal membaca file Excel shift: {e}")

    return {"shift": "OFF_SHIFT", "technician": None}

async def auto_assign_sdp_tickets(context: ContextTypes.DEFAULT_TYPE):
    """
    Job Background:
    1. GET ticket dengan Status = Open HANYA untuk grup divisi di SDP_NOTIFY_GROUPS
    2. Cek shift & tentukan teknisi via Round-Robin
    3. Update Technician + Status = In Progress Investigation
    4. VERIFICATION & Simpan ke File Log
    """
    
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    try:
        # 1. GET tickets yang di-filter KHUSUS untuk grup divisi kamu
        open_tickets = await asyncio.to_thread(
            sdp.list_requests, 15, "Open", config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal mengambil tiket Open untuk otomasi assign: {e}")
        return

    if not open_tickets:
        return

    current_shift = get_current_shift()

    for req in open_tickets:
        ticket_id = req.get("id")
        if not ticket_id:
            continue

        ticket_group = (req.get("group") or {}).get("name", "")

        try:
            # 2. Tentukan teknisi
            assigned_tech = select_technician_round_robin(current_shift)

            # 3. Update Technician + Status = In Progress Investigation
            update_payload = {
                "request": {
                    "status": {"name": "In Progress Investigation"},
                    "technician": {
                        "email_id": assigned_tech["email"]
                    }
                }
            }

            def _do_update():
                url = f"{sdp.base_url}/api/v3/requests/{ticket_id}"
                data = {"input_data": json.dumps(update_payload)}
                res = sdp.session.put(url, data=data)
                return res.status_code == 200, res.json()

            success, res_json = await asyncio.to_thread(_do_update)

            if success:
                # 4. VERIFICATION
                verified_req = await asyncio.to_thread(sdp.get_request, ticket_id)
                v_status = (verified_req.get("status") or {}).get("name", "")
                v_tech = (verified_req.get("technician") or {}).get("name", "")

                # 📝 CATAT KE FILE LOG
                write_ticket_log(
                    ticket_id=str(ticket_id),
                    tech_name=v_tech,
                    tech_email=assigned_tech['email'],
                    status=v_status,
                    shift=current_shift,
                    is_success=True
                )

                # Ubah baris 219 dari:
                f"{reply_note}"

                # Menjadi string langsung:
                "Tiket telah di-assign secara otomatis."

                # Notifikasi ke Telegram
                msg_text = (
                    f"✅ *[Otomasi Assign Tiket]*\n"
                    f"• Ticket ID: `#{ticket_id}`\n"
                    f"• Group: *{ticket_group}*\n"
                    f"• Subject: {req.get('subject', '')}\n"
                    f"• Assigned To: *{v_tech}* ({assigned_tech['email']})\n"
                    f"• Status Terbaru: *{v_status}*\n"
                    f"• Shift: {current_shift}"
                    f"{reply_note}"
                )
                await _broadcast_notify(context, msg_text)
            else:
                # 📝 CATAT KETIKA GAGAL UPDATE
                write_ticket_log(
                    ticket_id=str(ticket_id),
                    tech_name=assigned_tech['name'],
                    tech_email=assigned_tech['email'],
                    status="In Progress Investigation",
                    shift=current_shift,
                    is_success=False,
                    error_msg=str(res_json)
                )
                logger.error(f"Gagal update tiket #{ticket_id}: {res_json}")

        except Exception as e:
            # 📝 CATAT KETIKA EXCEPTION ERROR
            write_ticket_log(
                ticket_id=str(ticket_id),
                tech_name="N/A",
                tech_email="N/A",
                status="N/A",
                shift=current_shift,
                is_success=False,
                error_msg=str(e)
            )
            logger.exception(f"Error pada otomasi tiket #{ticket_id}: {e}")


# ==============================================================================
# LOGIKA SCRIPT UTAMA & TELEGRAM BOT HANDLERS
# ==============================================================================

def get_jira_client_for_user(user_id) -> JiraClient:
    """Ambil JiraClient pakai kredensial pribadi user ini (kalau sudah daftar
    lewat /myjira). Return None kalau belum daftar."""
    uid = str(user_id)
    if uid in _JIRA_CLIENT_CACHE:
        return _JIRA_CLIENT_CACHE[uid]
    account = user_accounts.get_account(uid)
    if not account:
        return None
    client = JiraClient(email=account["email"], api_token=account["api_token"])
    _JIRA_CLIENT_CACHE[uid] = client
    return client


def invalidate_jira_cache(user_id):
    _JIRA_CLIENT_CACHE.pop(str(user_id), None)


async def ensure_jira_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ambil JiraClient pribadi si pengirim pesan. Kalau belum daftar akun,
    kasih tahu untuk /myjira dulu, dan return None."""
    client = get_jira_client_for_user(update.effective_user.id)
    if client is None:
        await update.effective_message.reply_text(
            "Kamu belum daftarkan akun Jira pribadi di bot ini.\n"
            "Ketik /myjira untuk daftar dulu (email Jira + API token kamu sendiri), "
            "supaya logwork tercatat atas nama kamu sendiri di Jira."
        )
    return client

TZ = ZoneInfo(config.TIMEZONE)

TIME_PATTERN = re.compile(r"^(\d+w\s*)?(\d+d\s*)?(\d+h\s*)?(\d+m\s*)?$", re.IGNORECASE)

ISSUE_KEY_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9]+-\d+)\b")
JAM_RE = re.compile(r"\bjam\s+(\d{1,2})(?:[:.](\d{2}))?\s*(pagi|siang|sore|malam)\b", re.IGNORECASE)
DURASI_RE = re.compile(r"(?:selama|durasi)\s+([^,\.]+)", re.IGNORECASE)
DESKRIPSI_RE = re.compile(r"deskripsi(?:nya)?(?:\s+pekerjaan(?:nya|ya))?\s*[:\-]?\s*(.+)", re.IGNORECASE)


def _parse_jam_indo(hour: int, minute: int, period: str) -> dt.time:
    period = (period or "").lower()
    if period == "pagi":
        if hour == 12:
            hour = 0
    elif period == "siang":
        if hour < 12:
            hour += 12
    elif period in ("sore", "malam"):
        if hour == 12 and period == "malam":
            hour = 0
        elif hour < 12:
            hour += 12
    return dt.time(hour=hour % 24, minute=minute)


def parse_natural_log(text: str) -> dict:
    """Coba ekstrak issue key, jam, durasi, dan deskripsi dari satu kalimat bebas."""
    result = {}

    m = ISSUE_KEY_RE.search(text)
    if m:
        result["issue_key"] = m.group(1).upper()

    m = JAM_RE.search(text)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        result["start_time"] = _parse_jam_indo(hour, minute, m.group(3))

    m = DURASI_RE.search(text)
    if m:
        raw = m.group(1)
        raw = re.sub(r"\bjam\b", "h", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\b(menit|mnt)\b", "m", raw, flags=re.IGNORECASE)
        tokens = re.findall(r"\d+\s*[hm]", raw, flags=re.IGNORECASE)
        if tokens:
            result["duration"] = " ".join(t.replace(" ", "") for t in tokens)

    m = DESKRIPSI_RE.search(text)
    if m:
        result["description"] = m.group(1).strip().rstrip(".")

    if re.search(r"\bkemarin\b", text, re.IGNORECASE):
        result["date_offset"] = -1

    return result


def _extract_project_prefix(text: str):
    """Kenali 'TDBU-' atau 'TIC' sebagai kode project."""
    text = text.strip()
    if re.fullmatch(r"[A-Za-z]{2,10}-", text):
        return text[:-1].upper()
    if re.fullmatch(r"[A-Z]{2,10}", text):
        return text
    return None

# Conversation states
(
    LOG_ISSUE,
    LOG_TIME,
    LOG_DESC,
    LOG_DATE,
    LOG_CONFIRM,
    PICK_ISSUE_FOR_EDIT,
    PICK_WORKLOG_EDIT,
    EDIT_TIME,
    EDIT_DESC,
    PICK_ISSUE_FOR_DELETE,
    PICK_WORKLOG_DELETE,
    CONFIRM_DELETE,
    ADD_GUIDE_TITLE,
    ADD_GUIDE_KEYWORDS,
    ADD_GUIDE_CONTENT,
    ADD_GUIDE_CONFIRM,
    DEL_GUIDE_PICK,
    DEL_GUIDE_CONFIRM,
    EDIT_GUIDE_PICK,
    EDIT_GUIDE_MENU,
    EDIT_GUIDE_TITLE,
    EDIT_GUIDE_KEYWORDS,
    EDIT_GUIDE_CONTENT,
    ADD_GUIDE_ACTION_ASK,
    ADD_GUIDE_ACTION_SCRIPT,
    ADD_GUIDE_ACTION_FLAG,
    ADD_GUIDE_ACTION_TYPE,
    EDIT_GUIDE_ACTION_SCRIPT,
    EDIT_GUIDE_ACTION_FLAG,
    EDIT_GUIDE_ACTION_TYPE,
    RUN_ACTION_PARAM,
    RUN_ACTION_CONFIRM,
    ADD_GUIDE_ACTION_MODE,
    EDIT_GUIDE_ACTION_MODE,
    MYJIRA_EMAIL,
    MYJIRA_TOKEN,
    MYJIRA_CONFIRM,
) = range(37)


def _thread_id_from_update(update: Update):
    msg = update.effective_message
    if msg is not None and getattr(msg, "is_topic_message", False):
        return msg.message_thread_id
    return None


def _is_allowed_chat(chat_id) -> bool:
    return str(chat_id) in config.allowed_chat_ids()


def restricted(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        if not _is_allowed_chat(update.effective_chat.id):
            await update.effective_message.reply_text(
                "Maaf, saya tidak kenal dengan anda"
            )
            return ConversationHandler.END
        return await func(update, context, *args, **kwargs)

    return wrapper


def seconds_to_human(total_seconds: int) -> str:
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    return " ".join(parts) if parts else "0m"


# ---------------- basic commands ----------------

@restricted
async def chatid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    msg = update.message
    text = f"Chat ID di sini: `{chat.id}`\nTipe: {chat.type}"
    if msg.is_topic_message and msg.message_thread_id:
        text += f"\nTopic ID: `{msg.message_thread_id}`"
    await msg.reply_text(text, parse_mode=ParseMode.MARKDOWN)


# ---------------- /myjira: daftarkan akun Jira pribadi ----------------

@restricted
async def myjira_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""
    user_id = update.effective_user.id

    if arg == "remove":
        ok = user_accounts.remove_account(user_id)
        invalidate_jira_cache(user_id)
        await update.message.reply_text(
            "Akun Jira kamu sudah dihapus dari bot." if ok else "Kamu belum daftarkan akun Jira."
        )
        return ConversationHandler.END

    if arg == "status":
        account = user_accounts.get_account(user_id)
        if account:
            await update.message.reply_text(
                f"Akun Jira kamu terdaftar: {account['email']}\n\n"
                "Ganti akun: /myjira\nHapus: /myjira remove"
            )
        else:
            await update.message.reply_text(
                "Kamu belum daftarkan akun Jira. Ketik /myjira untuk daftar."
            )
        return ConversationHandler.END

    return await myjira_start(update, context)


async def myjira_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Daftarkan akun Jira pribadi kamu, supaya logwork/edit/delete tercatat atas nama kamu sendiri.\n\n"
        "Masukkan **email akun Jira** kamu:",
        parse_mode=ParseMode.MARKDOWN,
    )
    return MYJIRA_EMAIL


async def myjira_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email = update.message.text.strip()
    if "@" not in email:
        await update.message.reply_text("Sepertinya itu bukan email yang valid, coba lagi:")
        return MYJIRA_EMAIL
    context.user_data["myjira_email"] = email
    await update.message.reply_text(
        "Sekarang masukkan **API token** Jira kamu.\n\n"
        "Belum punya? Buat dulu di: https://id.atlassian.com/manage-profile/security/api-tokens\n"
        "(klik \"Create API token\", copy tokennya, lalu paste di sini)",
        parse_mode=ParseMode.MARKDOWN,
    )
    return MYJIRA_TOKEN


async def myjira_token(update: Update, context: ContextTypes.DEFAULT_TYPE):
    token = update.message.text.strip()
    if not token:
        await update.message.reply_text("Tokennya kosong, coba lagi:")
        return MYJIRA_TOKEN
    context.user_data["myjira_token"] = token
    email = context.user_data["myjira_email"]

    await update.message.reply_text("Mengecek akun ke Jira, mohon tunggu...")
    try:
        test_client = JiraClient(email=email, api_token=token)
        resp = await asyncio.to_thread(
            test_client._request, "GET", test_client._url("/rest/api/3/myself")
        )
        me = test_client._safe_json(resp)
        display_name = me.get("displayName", "")
    except Exception as e:
        await update.message.reply_text(
            f"Gagal verifikasi ke Jira:\n{e}\n\nCek lagi email/token-nya, atau /cancel untuk batal."
        )
        return MYJIRA_TOKEN

    context.user_data["myjira_display_name"] = display_name
    keyboard = [
        [
            InlineKeyboardButton("✅ Ya, simpan", callback_data="myjira_save_yes"),
            InlineKeyboardButton("Batal", callback_data="myjira_save_no"),
        ]
    ]
    await update.message.reply_text(
        f"Berhasil verifikasi! Akun Jira: {display_name} ({email})\n\nSimpan sebagai akun kamu?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return MYJIRA_CONFIRM


async def myjira_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "myjira_save_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    user_id = update.effective_user.id
    user_accounts.set_account(
        user_id, d["myjira_email"], d["myjira_token"], d.get("myjira_display_name", "")
    )
    invalidate_jira_cache(user_id)
    await query.edit_message_text(
        f"Akun Jira kamu ({d['myjira_email']}) berhasil disimpan. ✅\n"
        "Sekarang /log, /edit, /delete, /today, /week bakal pakai akun ini."
    )
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Halo! Saya itops_bot bantu logwork ke Jira dan Otomasi ServiceDesk.\n\n"
        "Perintah yang tersedia:\n"
        "/myjira - daftarkan akun Jira pribadi kamu (wajib sebelum /log dkk)\n"
        "/log - isi logwork baru\n"
        "/today - rekap logwork hari ini\n"
        "/week - rekap logwork minggu ini\n"
        "/tasks - lihat daftar task di suatu project (contoh: /tasks TDBU)\n"
        "/export - download report Excel task suatu project (contoh: /export TIC)\n"
        "/sdtickets [filter] - lihat tiket ServiceDesk Plus\n"
        "/sdticket <id> - lihat detail 1 tiket ServiceDesk Plus\n"
        "/sdreminder <menit>|off|status - atur reminder berkala utk tiket Open\n"
        "/edit - edit logwork yang sudah ada\n"
        "/delete - hapus logwork\n"
        "/guide <kata kunci> - munculkan guidance/panduan tersimpan\n"
        "/run <kata kunci> - langsung eksekusi action/script guidance\n"
        "/addguide - tambah guidance baru\n"
        "/listguide - lihat semua guidance tersimpan\n"
        "/cancel - batalkan proses yang sedang berjalan"
    )


@restricted
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Oke, dibatalkan.")
    return ConversationHandler.END


# ==============================================================================
# HELPER & HANDLER BULK LOGWORK (BARU)
# ==============================================================================

MONTH_MAP = {
    "januari": 1, "jan": 1, "februari": 2, "feb": 2, "maret": 3, "mar": 3,
    "april": 4, "apr": 4, "mei": 5, "juni": 6, "jun": 6, "juli": 7, "jul": 7,
    "agustus": 8, "agt": 8, "agu": 8, "september": 9, "sep": 9, "oktober": 10, "okt": 10,
    "november": 11, "nov": 11, "desember": 12, "des": 12
}

def parse_bulk_log_text(text: str) -> list:
    items = []
    lines = text.splitlines()
    current_year = dt.datetime.now(TZ).year

    for line in lines:
        line_str = line.strip()
        if not line_str:
            continue

        match_key = ISSUE_KEY_RE.search(line_str)
        if not match_key:
            continue
        issue_key = match_key.group(1).upper()

        dur_match = re.search(r'(?:selama|durasi)?\s*(\d+\s*(?:jam|h|menit|mnt|m))+', line_str, re.IGNORECASE)
        time_spent = "1h"
        if dur_match:
            raw_dur = dur_match.group(0)
            raw_dur = re.sub(r'\bjam\b', 'h', raw_dur, flags=re.IGNORECASE)
            raw_dur = re.sub(r'\b(menit|mnt)\b', 'm', raw_dur, flags=re.IGNORECASE)
            tokens = re.findall(r'\d+\s*[hm]', raw_dur, flags=re.IGNORECASE)
            if tokens:
                time_spent = " ".join(t.replace(" ", "") for t in tokens)

        date_obj = dt.datetime.now(TZ).date()
        date_match = re.search(r'(\d{1,2})\s+([a-zA-Z]+)(?:\s+(\d{4}))?', line_str, re.IGNORECASE)
        if date_match:
            day = int(date_match.group(1))
            month_str = date_match.group(2).lower()
            year = int(date_match.group(3)) if date_match.group(3) else current_year
            month = MONTH_MAP.get(month_str, date_obj.month)
            try:
                date_obj = dt.date(year, month, day)
            except ValueError:
                pass

        time_match = re.search(r'jam\s+(\d{1,2})[:.](\d{2})', line_str, re.IGNORECASE)
        start_time = None
        if time_match:
            start_time = dt.time(int(time_match.group(1)), int(time_match.group(2)))

        quotes = re.findall(r'"([^"]*)"', line_str)
        if quotes:
            comment = quotes[0]
        else:
            clean_text = re.sub(r'task\s+[A-Za-z0-9-]+', '', line_str, flags=re.IGNORECASE)
            clean_text = re.sub(r'pada tanggal.*', '', clean_text, flags=re.IGNORECASE)
            clean_text = re.sub(r'selama.*', '', clean_text, flags=re.IGNORECASE)
            comment = clean_text.strip(" .a-b-c-1-2-3-")

        items.append({
            "issue_key": issue_key,
            "comment": comment or "Worklog update",
            "time_spent": time_spent,
            "date": date_obj,
            "start_time": start_time
        })

    return items


async def show_bulk_log_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items = context.user_data.get("bulk_items", [])
    if not items:
        await update.message.reply_text("Tidak ada data logwork yang berhasil dibaca. Coba lagi.")
        return ConversationHandler.END

    lines = [f"📋 <b>Daftar {len(items)} Logwork yang akan dikirim:</b>\n"]
    for idx, item in enumerate(items, 1):
        jam_str = f" jam {item['start_time'].strftime('%H:%M')}" if item.get('start_time') else ""
        lines.append(
            f"<b>{idx}. {item['issue_key']}</b> - {item['time_spent']}\n"
            f"   • Deskripsi: <i>{item['comment']}</i>\n"
            f"   • Tanggal: {item['date'].strftime('%d %b %Y')}{jam_str}\n"
        )

    lines.append("Kirim semua logwork ini ke Jira?")

    keyboard = [
        [
            InlineKeyboardButton("✅ Submit Semua", callback_data="bulk_confirm_yes"),
            InlineKeyboardButton("❌ Batal", callback_data="bulk_confirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("\n".join(lines), reply_markup=markup, parse_mode=ParseMode.HTML)
    return LOG_BULK_CONFIRM


async def bulk_log_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "bulk_confirm_no":
        await query.edit_message_text("❌ Proses logwork dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    items = context.user_data.get("bulk_items", [])
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text(f"⏳ Sedang mengirim {len(items)} logwork ke Jira, mohon tunggu...")

    success_count = 0
    failed_items = []

    for item in items:
        try:
            await asyncio.to_thread(
                jira_client.add_worklog,
                item["issue_key"],
                item["time_spent"],
                item["comment"],
                item["date"],
                item.get("start_time"),
            )
            success_count += 1
        except JiraError as e:
            failed_items.append(f"{item['issue_key']}: {e}")

    result_text = f"✅ <b>Berhasil mengirim {success_count} dari {len(items)} logwork ke Jira!</b>"
    if failed_items:
        result_text += "\n\n⚠️ <b>Gagal pada task berikut:</b>\n" + "\n".join(failed_items)

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=result_text,
        parse_mode=ParseMode.HTML,
        message_thread_id=_thread_id_from_update(update)
    )

    context.user_data.clear()
    return ConversationHandler.END

LOG_BULK_CONFIRM = 37

# ---------------- /log flow ----------------

@restricted
async def log_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    text = update.message.text or ""
    after = re.sub(r"^/log(@\w+)?\s*", "", text, flags=re.IGNORECASE).strip()

    # Cek apakah input mengandung multi-baris / banyak task
    if "\n" in after or len(re.findall(r'[A-Za-z]+-\d+', after)) > 1:
        parsed_items = parse_bulk_log_text(after)
        if parsed_items:
            context.user_data["bulk_items"] = parsed_items
            return await show_bulk_log_confirm(update, context)

    if after:
        context.user_data.update(parse_natural_log(after))
    return await advance_log_flow(update, context)


async def advance_log_flow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    msg = update.effective_message

    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END

    if not d.get("issue_key"):
        await msg.reply_text("Isi ke issue Jira mana? (contoh: PROJ-123)")
        return LOG_ISSUE

    if "issue_title" not in d:
        try:
            d["issue_title"] = await asyncio.to_thread(jira_client.get_issue_summary_title, d["issue_key"])
        except JiraError:
            await msg.reply_text(
                f"Issue {d['issue_key']} tidak ditemukan atau tidak bisa diakses. "
                "Coba ketik issue key yang benar (contoh: PROJ-123), atau /cancel."
            )
            d.pop("issue_key", None)
            return LOG_ISSUE

    if not d.get("time_spent"):
        if d.get("duration"):
            d["time_spent"] = d["duration"]
        else:
            await msg.reply_text(
                f'Issue: {d["issue_key"]} - "{d["issue_title"]}"\n\n'
                "Berapa lama waktu yang dihabiskan? (contoh: 2h, 1h 30m, 45m)"
            )
            return LOG_TIME

    if not d.get("comment"):
        if d.get("description"):
            d["comment"] = d["description"]
        else:
            await msg.reply_text("Deskripsi pekerjaannya apa? (bisa singkat saja)")
            return LOG_DESC

    if not d.get("date"):
        if d.get("date_offset"):
            today = dt.datetime.now(TZ).date()
            d["date"] = today + dt.timedelta(days=d["date_offset"])
        else:
            keyboard = [
                [
                    InlineKeyboardButton("Hari ini", callback_data="date_today"),
                    InlineKeyboardButton("Kemarin", callback_data="date_yesterday"),
                ],
                [InlineKeyboardButton("Ketik tanggal manual", callback_data="date_manual")],
            ]
            await msg.reply_text(
                "Tanggal logwork-nya kapan?", reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return LOG_DATE

    return await log_show_confirm(update, context)


async def log_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["issue_key"] = update.message.text.strip().upper()
    context.user_data.pop("issue_title", None)
    return await advance_log_flow(update, context)


async def log_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    time_text = update.message.text.strip()
    if not TIME_PATTERN.match(time_text.replace(" ", "")) and not TIME_PATTERN.match(time_text):
        await update.message.reply_text(
            "Format waktu kurang tepat. Gunakan contoh seperti: 2h, 1h 30m, 45m, 1d. Coba lagi:"
        )
        return LOG_TIME
    context.user_data["time_spent"] = time_text
    return await advance_log_flow(update, context)


async def log_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["comment"] = update.message.text.strip()
    return await advance_log_flow(update, context)


async def log_date_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = dt.datetime.now(TZ).date()
    if query.data == "date_today":
        context.user_data["date"] = today
        return await log_show_confirm(update, context)
    elif query.data == "date_yesterday":
        context.user_data["date"] = today - dt.timedelta(days=1)
        return await log_show_confirm(update, context)
    else:
        await query.edit_message_text("Ketik tanggalnya, format: YYYY-MM-DD (contoh: 2026-08-01)")
        return LOG_DATE


async def log_date_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        date_obj = dt.datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Format tanggal salah. Contoh yang benar: 2026-08-01")
        return LOG_DATE
    context.user_data["date"] = date_obj
    return await log_show_confirm(update, context)


async def log_show_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    tanggal_line = d["date"].isoformat()
    if d.get("start_time"):
        tanggal_line += f" (jam {d['start_time'].strftime('%H:%M')})"
    text = (
        "Konfirmasi logwork berikut:\n\n"
        f"Issue: {d['issue_key']} - \"{d['issue_title']}\"\n"
        f"Waktu: {d['time_spent']}\n"
        f"Deskripsi: {d['comment']}\n"
        f"Tanggal: {tanggal_line}\n\n"
        "Kirim ke Jira sekarang?"
    )
    keyboard = [
        [
            InlineKeyboardButton("Ya, kirim", callback_data="confirm_yes"),
            InlineKeyboardButton("Batal", callback_data="confirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    return LOG_CONFIRM


async def log_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(
            jira_client.add_worklog,
            d["issue_key"],
            d["time_spent"],
            d["comment"],
            d["date"],
            d.get("start_time"),
        )
        await query.edit_message_text(
            f"Berhasil dicatat ke {d['issue_key']} ({d['time_spent']}). 👍"
        )
    except JiraError as e:
        await query.edit_message_text(f"Gagal mengirim ke Jira:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


# ---------------- /today & /week ----------------

def format_summary(entries: list, heading: str) -> str:
    if not entries:
        return f"{heading}\n\nBelum ada logwork tercatat."
    lines = [heading, ""]
    grand_total = 0
    for item in entries:
        total = seconds_to_human(item["total_seconds"])
        grand_total += item["total_seconds"]
        lines.append(f"• {item['issue_key']} - \"{item['summary']}\" — {total}")
        for e in item["entries"]:
            comment = f" ({e['comment']})" if e["comment"] else ""
            lines.append(f"    - {e['date']} · {e['time_spent']}{comment}")
    lines.append("")
    lines.append(f"Total: {seconds_to_human(grand_total)}")
    return "\n".join(lines)


@restricted
async def today_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return
    today = dt.datetime.now(TZ).date()
    try:
        entries = await asyncio.to_thread(jira_client.get_summary, today, today)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data dari Jira:\n{e}")
        return
    await update.message.reply_text(
        _safe_telegram_text(format_summary(entries, f"Rekap hari ini ({today.isoformat()}):"))
    )


@restricted
async def week_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return
    today = dt.datetime.now(TZ).date()
    start = today - dt.timedelta(days=today.weekday())
    end = today
    try:
        entries = await asyncio.to_thread(jira_client.get_summary, start, end)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data dari Jira:\n{e}")
        return
    await update.message.reply_text(
        _safe_telegram_text(
            format_summary(entries, f"Rekap minggu ini ({start.isoformat()} s/d {end.isoformat()}):")
        )
    )


# ---------------- /tasks ----------------

@restricted
async def list_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text.lower().startswith("/tasks"):
        parts = text.split(maxsplit=1)
        if len(parts) < 2 or not parts[1].strip():
            await update.message.reply_text(
                "Ketik kode project-nya, contoh: /tasks TDBU\n"
                "Atau langsung ketik kode project-nya saja, contoh: TDBU- atau TIC"
            )
            return
        project_key = re.sub(r"-$", "", parts[1].strip()).upper()
    else:
        project_key = _extract_project_prefix(text)
        if not project_key:
            return

    try:
        issues = await asyncio.to_thread(jira.search_project_issues, project_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil daftar task project {project_key}:\n{e}")
        return

    if not issues:
        await update.message.reply_text(
            f"Tidak ada task ditemukan untuk project {project_key}.\n"
            "Cek lagi kode project-nya, atau pastikan kamu punya akses ke project tersebut."
        )
        return

    lines = [f"📋 Task terbaru di project {project_key} (maks 50, urut yang terakhir diupdate):", ""]
    for issue in issues:
        key = issue["key"]
        summary = issue["fields"].get("summary", "")
        status = (issue["fields"].get("status") or {}).get("name", "")
        lines.append(f"• {key} - {summary} [{status}]")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines)))


# ---------------- ServiceDesk Plus ----------------
import html
import re

import html
import re

def _safe_telegram_text(text: str, limit: int = 3800) -> str:
    """Memotong teks jika melebihi batas karakter Telegram"""
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n... (dipotong, terlalu banyak data -- coba filter lebih spesifik)"

def _strip_html(text: str) -> str:
    if not text:
        return ""
    
    # 1. Decode HTML entities (&nbsp;, &amp;, dll)
    decoded_text = html.unescape(text)
    
    # 2. Ganti tag pemisah baris HTML (<br>, <p>, <div>, <li>) dengan newline (\n)
    decoded_text = re.sub(r"(?i)<(br|p|div|li)\b[^>]*>", "\n", decoded_text)
    
    # 3. Hapus sisa tag HTML lainnya
    clean_text = re.sub(r"<[^>]+>", "", decoded_text)
    
    # 4. Rapikan spasi horizontal tanpa menghapus baris baru
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in clean_text.splitlines()]
    
    # 5. Gabungkan kembali baris yang tidak kosong
    return "\n".join(line for line in lines if line)

def _format_sdp_request_line(req: dict) -> str:
    """Merapikan format teks tampilan tiket untuk notifikasi & daftar tiket"""
    req_id = req.get("id", "")
    subject = req.get("subject", "")
    status = (req.get("status") or {}).get("name", "")
    requester = (req.get("requester") or {}).get("name", "")
    
    # Ambil full text deskripsi yang sudah dibersihkan dari HTML/entity
    description = _strip_html(req.get("description", ""))
    
    desc_line = f"\nDescription:\n{description}" if description else ""
    
    return (
        f"Ticket #{req_id}\n"
        f"Subject: {subject}\n"
        f"Requester: {requester}\n"
        f"Status: {status}"
        f"{desc_line}"
    )
    
SDP_FILTER_RE = re.compile(r"(status|group)\s*:\s*(.+?)(?=\s+(?:status|group)\s*:|$)", re.IGNORECASE)


def _parse_sdp_filters(text: str) -> dict:
    filters = {}
    for m in SDP_FILTER_RE.finditer(text):
        filters[m.group(1).lower()] = m.group(2).strip()
    return filters


@restricted
async def sdtickets_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not sdp:
        await update.message.reply_text(
            "ServiceDesk Plus belum dikonfigurasi. Isi dulu SDP_BASE_URL dan SDP_API_KEY di file .env, "
            "lalu restart bot."
        )
        return
    parts = update.message.text.split(maxsplit=1)
    raw_filter = parts[1].strip() if len(parts) > 1 else ""
    filters = _parse_sdp_filters(raw_filter)
    status_filter = filters.get("status")
    group_filter = filters.get("group")

    used_default_groups = False
    if group_filter is None and config.SDP_NOTIFY_GROUPS:
        group_filter = config.SDP_NOTIFY_GROUPS
        used_default_groups = True
    elif group_filter and group_filter.strip().lower() == "all":
        group_filter = None

    try:
        requests_list = await asyncio.to_thread(
            sdp.list_requests, 15, status_filter, group_filter
        )
    except SDPError as e:
        await update.message.reply_text(f"Gagal mengambil data dari ServiceDesk Plus:\n{e}")
        return

    if not requests_list:
        await update.message.reply_text("Tidak ada tiket ditemukan dengan filter tersebut.")
        return

    filter_bits = []
    if status_filter:
        filter_bits.append(f"status: {status_filter}")
    if used_default_groups:
        filter_bits.append(f"group kamu: {', '.join(config.SDP_NOTIFY_GROUPS)}")
    elif group_filter:
        filter_bits.append(f"group: {group_filter}")
    filter_note = f" ({', '.join(filter_bits)})" if filter_bits else ""

    lines = [f"🎫 15 tiket ServiceDesk Plus terbaru{filter_note}:", ""]
    for req in requests_list:
        lines.append(_format_sdp_request_line(req))
        lines.append("")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines).rstrip()))

@restricted
async def sdticket_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not sdp:
        await update.message.reply_text(
            "ServiceDesk Plus belum dikonfigurasi. Isi dulu SDP_BASE_URL dan SDP_API_KEY di file .env, "
            "lalu restart bot."
        )
        return
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await update.message.reply_text("Ketik: /sdticket <ID tiket>\ncontoh: /sdticket 12345")
        return
    request_id = parts[1].strip()

    try:
        req = await asyncio.to_thread(sdp.get_request, request_id)
    except SDPError as e:
        await update.message.reply_text(f"Gagal mengambil tiket {request_id}:\n{e}")
        return

    subject = req.get("subject", "")
    status = (req.get("status") or {}).get("name", "")
    requester = (req.get("requester") or {}).get("name", "")
    priority = (req.get("priority") or {}).get("name", "")
    created = (req.get("created_time") or {}).get("display_value", "")
    description = req.get("description", "") or ""
    description = re.sub(r"<[^>]+>", "", description).strip()

    # Tampilkan FULL TEXT deskripsi (tanpa dipotong [:1500])
    text = (
        f"🎫 Tiket #{request_id}\n\n"
        f"Subjek: {subject}\n"
        f"Status: {status}\n"
        f"Requester: {requester}\n"
        f"Priority: {priority}\n"
        f"Dibuat: {created}\n\n"
        f"Deskripsi:\n{description or '(kosong)'}"
    )

    # Gunakan _safe_telegram_text agar tidak error jika pesan terlalu panjang untuk limit Telegram
    await update.message.reply_text(_safe_telegram_text(text))


# ---------------- /export ----------------

@restricted
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await update.message.reply_text("Ketik: /export <KODE_PROJECT>\ncontoh: /export TIC")
        return
    project_key = re.sub(r"-$", "", parts[1].strip()).upper()

    await update.message.reply_text(
        f"Sedang mengambil data project {project_key} dari Jira (termasuk riwayat status "
        "Dev/QA Done), mohon tunggu sebentar..."
    )
    try:
        rows = await asyncio.to_thread(jira.get_project_export, project_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data:\n{e}")
        return

    if not rows:
        await update.message.reply_text(
            f"Tidak ada task ditemukan di project {project_key}. "
            "Cek lagi kode project-nya, atau pastikan kamu punya akses."
        )
        return

    filepath = build_export_excel(project_key, rows)
    try:
        with open(filepath, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(filepath),
                caption=f"📊 Report project {project_key} ({len(rows)} task)",
            )
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


# ---------------- /edit flow ----------------

@restricted
async def edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END
    await update.message.reply_text("Issue Jira mana yang mau diedit worklog-nya? (contoh: PROJ-123)")
    return PICK_ISSUE_FOR_EDIT


async def edit_pick_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    issue_key = update.message.text.strip().upper()
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        return ConversationHandler.END
    try:
        worklogs = await asyncio.to_thread(jira_client.get_my_recent_worklogs, issue_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil worklog:\n{e}")
        return PICK_ISSUE_FOR_EDIT
    if not worklogs:
        await update.message.reply_text(
            f"Tidak ada worklog kamu di {issue_key} dalam 14 hari terakhir. Coba issue lain atau /cancel."
        )
        return PICK_ISSUE_FOR_EDIT
    context.user_data["issue_key"] = issue_key
    buttons = []
    for w in worklogs:
        label = f"{w['started'][:10]} · {w.get('timeSpent', '')}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"wl_{w['id']}")])
    await update.message.reply_text(
        "Pilih worklog yang mau diedit:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return PICK_WORKLOG_EDIT


async def edit_pick_worklog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    worklog_id = query.data.replace("wl_", "")
    context.user_data["worklog_id"] = worklog_id
    await query.edit_message_text(
        "Masukkan waktu baru (contoh: 2h 30m), atau ketik - untuk tidak mengubah waktu:"
    )
    return EDIT_TIME


async def edit_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["new_time"] = None if text == "-" else text
    await update.message.reply_text(
        "Masukkan deskripsi baru, atau ketik - untuk tidak mengubah deskripsi:"
    )
    return EDIT_DESC


async def edit_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    new_comment = None if text == "-" else text
    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(
            jira_client.update_worklog, d["issue_key"], d["worklog_id"], d.get("new_time"), new_comment
        )
        await update.message.reply_text("Worklog berhasil diperbarui. ✅")
    except JiraError as e:
        await update.message.reply_text(f"Gagal update worklog:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


# ---------------- /delete flow ----------------

@restricted
async def delete_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END
    await update.message.reply_text("Issue Jira mana yang worklog-nya mau dihapus? (contoh: PROJ-123)")
    return PICK_ISSUE_FOR_DELETE


async def delete_pick_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    issue_key = update.message.text.strip().upper()
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        return ConversationHandler.END
    try:
        worklogs = await asyncio.to_thread(jira_client.get_my_recent_worklogs, issue_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil worklog:\n{e}")
        return PICK_ISSUE_FOR_DELETE
    if not worklogs:
        await update.message.reply_text(
            f"Tidak ada worklog kamu di {issue_key} dalam 14 hari terakhir. Coba issue lain atau /cancel."
        )
        return PICK_ISSUE_FOR_DELETE
    context.user_data["issue_key"] = issue_key
    buttons = []
    for w in worklogs:
        label = f"{w['started'][:10]} · {w.get('timeSpent', '')}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"wl_{w['id']}")])
    await update.message.reply_text(
        "Pilih worklog yang mau dihapus:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return PICK_WORKLOG_DELETE


async def delete_pick_worklog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    worklog_id = query.data.replace("wl_", "")
    context.user_data["worklog_id"] = worklog_id
    keyboard = [
        [
            InlineKeyboardButton("Ya, hapus", callback_data="del_yes"),
            InlineKeyboardButton("Batal", callback_data="del_no"),
        ]
    ]
    await query.edit_message_text(
        "Yakin mau hapus worklog ini?", reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return CONFIRM_DELETE


async def delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "del_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(jira_client.delete_worklog, d["issue_key"], d["worklog_id"])
        await query.edit_message_text("Worklog berhasil dihapus. 🗑️")
    except JiraError as e:
        await update.message.reply_text(f"Gagal menghapus worklog:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


# ---------------- guidance / knowledge base ----------------

async def send_guidance_entry(update: Update, context: ContextTypes.DEFAULT_TYPE, entry: dict):
    caption = f"📘 {entry['title']}\n\n{entry['content']}".strip()
    attachment = entry.get("attachment")
    action = entry.get("action")
    chat_id = update.effective_chat.id

    action_markup = None
    if action:
        action_markup = InlineKeyboardMarkup(
            [[InlineKeyboardButton("⚙️ Jalankan Script", callback_data=f"runaction_{entry['id']}")]]
        )

    if attachment and attachment.get("file_id"):
        if update.callback_query:
            try:
                await update.callback_query.edit_message_reply_markup(reply_markup=None)
            except Exception:
                pass
        thread_id = _thread_id_from_update(update)
        await context.bot.send_document(
            chat_id=chat_id,
            document=attachment["file_id"],
            filename=attachment.get("file_name"),
            caption=caption[:1024] if caption else None,
            message_thread_id=thread_id,
        )
        if len(caption) > 1024:
            await context.bot.send_message(chat_id=chat_id, text=caption, message_thread_id=thread_id)
        if action_markup:
            await context.bot.send_message(
                chat_id=chat_id,
                text="Guidance ini bisa langsung dieksekusi:",
                reply_markup=action_markup,
                message_thread_id=thread_id,
            )
        return

    if update.callback_query:
        await update.callback_query.edit_message_text(caption, reply_markup=action_markup)
    else:
        await update.message.reply_text(caption, reply_markup=action_markup)


async def send_guidance_matches(update: Update, context: ContextTypes.DEFAULT_TYPE, query: str, matches: list):
    if len(matches) == 1:
        await send_guidance_entry(update, context, matches[0])
        return
    buttons = [
        [InlineKeyboardButton(m["title"], callback_data=f"gid_{m['id']}")] for m in matches[:8]
    ]
    await update.message.reply_text(
        f"Ada beberapa guidance yang cocok dengan '{query}', pilih salah satu:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


@restricted
async def guide_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    query = parts[1].strip() if len(parts) > 1 else ""
    if not query:
        await update.message.reply_text(
            "Ketik: /guide <kata kunci>\ncontoh: /guide koneksi jaringan error"
        )
        return
    matches = guidance_store.find_matches(query)
    if not matches:
        await update.message.reply_text(
            f"Guidance untuk '{query}' tidak ditemukan.\n"
            "Ketik /listguide untuk lihat semua topik yang tersedia, "
            "atau /addguide untuk menambahkan yang baru."
        )
        return
    await send_guidance_matches(update, context, query, matches)


async def guide_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("gid_", "")
    entry = guidance_store.get_guidance(gid)
    if not entry:
        await query.edit_message_text("Guidance tidak ditemukan (mungkin sudah dihapus).")
        return
    await send_guidance_entry(update, context, entry)


# ---------------- eksekusi action / script dari guidance ----------------

def _action_param_list(action: dict) -> list:
    if action.get("params"):
        return action["params"]
    flag = action.get("param_flag")
    if flag:
        return [{"flag": flag, "type": action.get("param_type", "text")}]
    return []


async def _ask_next_action_param(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    flags = d["run_flags"]
    idx = d["run_flag_index"]

    if idx >= len(flags):
        return await _show_run_confirm(update, context)

    current = flags[idx]
    flag_name = current.get("flag", "")
    ptype = current.get("type", "text")
    if ptype == "file":
        text = f"Upload file untuk parameter {flag_name}:"
    else:
        text = f"Masukkan nilai untuk parameter {flag_name}\n(kalau lebih dari 1 nilai, pisahkan dengan koma)"

    if update.callback_query:
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text=text, message_thread_id=_thread_id_from_update(update)
        )
    else:
        await update.effective_message.reply_text(text)
    return RUN_ACTION_PARAM


async def _show_run_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    lines = ["Konfirmasi:", "", f"Script: {d['run_script_path']}"]
    for flag_name, info in d["run_values"].items():
        lines.append(f"{flag_name}: {info['display']}")
    lines.append("")
    lines.append("Jalankan sekarang?")

    keyboard = [
        [
            InlineKeyboardButton("✅ Ya, jalankan", callback_data="runconfirm_yes"),
            InlineKeyboardButton("Batal", callback_data="runconfirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="\n".join(lines),
            reply_markup=markup,
            message_thread_id=_thread_id_from_update(update),
        )
    else:
        await update.effective_message.reply_text("\n".join(lines), reply_markup=markup)
    return RUN_ACTION_CONFIRM


async def _start_run_action(update: Update, context: ContextTypes.DEFAULT_TYPE, entry: dict):
    action = entry["action"]
    flags = _action_param_list(action)
    context.user_data.clear()
    context.user_data["run_guide_id"] = entry["id"]
    context.user_data["run_script_path"] = action.get("script_path", "")
    context.user_data["run_flags"] = flags

    paired = action.get("input_mode") == "paired" and len(flags) > 1
    context.user_data["run_paired_mode"] = paired

    if paired:
        flag_names = ", ".join(f["flag"].lstrip("-") for f in flags)
        example = "\n".join(
            "8000044321, SS20" if i == 0 else "8000044322, SS21" for i in range(2)
        )
        text = (
            f"Masukkan data, satu baris per kombinasi ({flag_names}):\n\n"
            f"Contoh:\n{example}"
        )
        msg = update.effective_message
        if update.callback_query:
            await context.bot.send_message(
                chat_id=update.effective_chat.id, text=text, message_thread_id=_thread_id_from_update(update)
            )
        else:
            await msg.reply_text(text)
        return RUN_ACTION_PARAM

    context.user_data["run_flag_index"] = 0
    context.user_data["run_values"] = {}
    return await _ask_next_action_param(update, context)


async def run_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    query = parts[1].strip() if len(parts) > 1 else ""
    if not query:
        await update.message.reply_text("Ketik: /run <kata kunci guidance>\ncontoh: /run delete reservation")
        return ConversationHandler.END

    matches = guidance_store.find_matches(query)
    if not matches:
        await update.message.reply_text(
            f"Guidance untuk '{query}' tidak ditemukan.\nKetik /listguide untuk lihat semua topik."
        )
        return ConversationHandler.END

    matches_with_action = [m for m in matches if m.get("action")]
    if not matches_with_action:
        await update.message.reply_text(
            f"Guidance '{matches[0]['title']}' ditemukan, tapi belum punya action/script.\n"
            "Tambahkan lewat /editguide dulu (tombol '⚙️ Tambah Action'), "
            "atau pakai /guide untuk lihat isinya saja."
        )
        return ConversationHandler.END

    if len(matches_with_action) > 1:
        buttons = [
            [InlineKeyboardButton(f"▶️ {m['title']}", callback_data=f"runaction_{m['id']}")]
            for m in matches_with_action[:8]
        ]
        await update.message.reply_text(
            f"Ada beberapa guidance dengan action yang cocok dengan '{query}', pilih salah satu:",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return ConversationHandler.END

    entry = matches_with_action[0]
    await update.message.reply_text(f"Jalankan: {entry['title']}")
    return await _start_run_action(update, context, entry)


async def run_action_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_allowed_chat(update.effective_chat.id):
        return ConversationHandler.END
    gid = query.data.replace("runaction_", "")
    entry = guidance_store.get_guidance(gid)
    thread_id = _thread_id_from_update(update)
    if not entry or not entry.get("action"):
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text="Action tidak ditemukan.", message_thread_id=thread_id
        )
        return ConversationHandler.END

    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=f"Jalankan: {entry['title']}", message_thread_id=thread_id
    )
    return await _start_run_action(update, context, entry)


async def run_action_param(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    flags = d.get("run_flags")
    if not flags:
        await update.message.reply_text("Sesi eksekusi tidak ditemukan lagi, dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    if d.get("run_paired_mode"):
        text = (update.message.text or "").strip()
        if not text:
            await update.message.reply_text("Datanya kosong, coba lagi (kirim ulang semua baris):")
            return RUN_ACTION_PARAM
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        values = {f["flag"]: [] for f in flags}
        for line in lines:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != len(flags):
                flag_names = ", ".join(f["flag"].lstrip("-") for f in flags)
                await update.message.reply_text(
                    f"Baris '{line}' jumlah nilainya ({len(parts)}) tidak sesuai jumlah "
                    f"parameter ({len(flags)}: {flag_names}). Coba lagi, kirim ulang semua baris:"
                )
                return RUN_ACTION_PARAM
            for f, val in zip(flags, parts):
                values[f["flag"]].append(val)
        d["run_values"] = {
            flag_name: {"type": "text", "value": vals, "display": ", ".join(vals)}
            for flag_name, vals in values.items()
        }
        return await _show_run_confirm(update, context)

    idx = d["run_flag_index"]
    current = flags[idx]
    flag_name = current.get("flag", "")
    ptype = current.get("type", "text")

    if ptype == "file":
        if not update.message.document:
            await update.message.reply_text("Mohon upload file yang dibutuhkan, atau ketik /cancel.")
            return RUN_ACTION_PARAM
        doc = update.message.document
        tg_file = await doc.get_file()
        local_path = os.path.join(tempfile.gettempdir(), doc.file_name)
        await tg_file.download_to_drive(local_path)
        d["run_values"][flag_name] = {"type": "file", "value": local_path, "display": doc.file_name}
    else:
        text = (update.message.text or "").strip()
        if not text:
            await update.message.reply_text("Nilainya kosong, coba lagi:")
            return RUN_ACTION_PARAM
        values = [v.strip() for v in text.split(",") if v.strip()]
        d["run_values"][flag_name] = {"type": "text", "value": values, "display": text}

    d["run_flag_index"] += 1
    return await _ask_next_action_param(update, context)


def _filter_item_detail_lines(text: str) -> str:
    item_line_re = re.compile(r"^\s*\[\d+\]")
    kept_lines = [line for line in text.splitlines() if not item_line_re.match(line)]
    result_lines = []
    prev_blank = False
    for line in kept_lines:
        is_blank = not line.strip()
        if is_blank and prev_blank:
            continue
        result_lines.append(line)
        prev_blank = is_blank
    return "\n".join(result_lines).strip()


async def run_action_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "runconfirm_no":
        await query.edit_message_text("Dibatalkan, script tidak dijalankan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    script_path = d.get("run_script_path", "")
    run_values = d.get("run_values", {})
    if not script_path:
        await query.edit_message_text("Sesi eksekusi tidak ditemukan lagi, dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text(f"⏳ Menjalankan {os.path.basename(script_path)}, mohon tunggu...")

    cmd = ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script_path]
    for flag_name, info in run_values.items():
        cmd.append(flag_name)
        if info["type"] == "file":
            cmd.append(info["value"])
        else:
            cmd.extend(info["value"])

    chat_id = update.effective_chat.id
    thread_id = _thread_id_from_update(update)
    script_dir = os.path.dirname(script_path) or None
    try:
        result = await asyncio.to_thread(
            subprocess.run, cmd, capture_output=True, text=True, timeout=300, cwd=script_dir
        )
        output = (result.stdout or "").strip()
        err = (result.stderr or "").strip()
        output = _filter_item_detail_lines(output)
        combined = output + (f"\n\n[stderr]\n{err}" if err else "")
        combined = combined or "(tidak ada output dari script)"
        if len(combined) > 3500:
            combined = combined[:3500] + "\n... (output dipotong, terlalu panjang)"
        icon = "✅" if result.returncode == 0 else "⚠️"
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"{icon} Selesai (exit code {result.returncode}):\n\n{combined}",
            message_thread_id=thread_id,
        )
    except subprocess.TimeoutExpired:
        await context.bot.send_message(
            chat_id=chat_id,
            text="⚠️ Script berjalan lebih dari 5 menit, dihentikan paksa (timeout).",
            message_thread_id=thread_id,
        )
    except FileNotFoundError:
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"⚠️ Gagal menjalankan: file script tidak ditemukan di path:\n{script_path}",
            message_thread_id=thread_id,
        )
    except Exception as e:
        await context.bot.send_message(
            chat_id=chat_id, text=f"⚠️ Gagal menjalankan script:\n{e}", message_thread_id=thread_id
        )

    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def list_guide(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text(
            "Belum ada guidance tersimpan. Tambahkan lewat /addguide, "
            "atau isi langsung file guidance.json."
        )
        return
    lines = ["📚 Daftar guidance yang tersimpan:", ""]
    for item in data:
        kw = ", ".join(item.get("keywords", [])[:3])
        lines.append(f"• [{item['id']}] {item['title']}")
        if kw:
            lines.append(f"    kata kunci: {kw}")
        lines.append("")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines).rstrip()))


async def guide_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _is_allowed_chat(update.effective_chat.id):
        return
    text = update.message.text
    matches = guidance_store.find_matches(text)
    if not matches:
        return
    await send_guidance_matches(update, context, text, matches)


@restricted
async def addguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Judul guidance-nya apa?\ncontoh: Cara mengatasi koneksi jaringan error"
    )
    return ADD_GUIDE_TITLE


async def addguide_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["guide_title"] = update.message.text.strip()
    await update.message.reply_text(
        "Kata kunci pemicunya apa saja? Pisahkan dengan koma.\n"
        "Ini dipakai supaya bot bisa mengenali kapan harus munculkan guidance ini "
        "(lewat /guide maupun kalau kamu ketik bebas).\n\n"
        "Contoh: koneksi jaringan error, jaringan error, network issue"
    )
    return ADD_GUIDE_KEYWORDS


async def addguide_keywords(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keywords = [k.strip() for k in update.message.text.split(",") if k.strip()]
    if not keywords:
        await update.message.reply_text("Kata kuncinya kosong, coba lagi (pisahkan dengan koma):")
        return ADD_GUIDE_KEYWORDS
    context.user_data["guide_keywords"] = keywords
    await update.message.reply_text(
        "Sekarang kirim isi/detail guidance-nya (boleh panjang & multi-baris).\n"
        "Atau kalau ada file pendukung (script, dokumen, dll), langsung upload file-nya "
        "di sini juga boleh — caption di file-nya akan dipakai sebagai deskripsi."
    )
    return ADD_GUIDE_CONTENT


async def addguide_content(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document:
        doc = update.message.document
        context.user_data["guide_content"] = (update.message.caption or "").strip()
        context.user_data["guide_attachment"] = {
            "file_id": doc.file_id,
            "file_name": doc.file_name,
        }
    else:
        context.user_data["guide_content"] = update.message.text.strip()
        context.user_data["guide_attachment"] = None

    context.user_data["guide_action"] = None
    keyboard = [
        [
            InlineKeyboardButton("Ya", callback_data="ga_yes"),
            InlineKeyboardButton("Tidak", callback_data="ga_no"),
        ]
    ]
    await update.message.reply_text(
        "Apakah guidance ini bisa dieksekusi otomatis oleh bot (menjalankan script "
        "PowerShell di laptop kamu)?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return ADD_GUIDE_ACTION_ASK


async def addguide_action_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "ga_no":
        return await addguide_show_preview(update, context)
    await query.edit_message_text(
        "Masukkan path lengkap file script-nya di laptop kamu.\n"
        r"Contoh: D:\promovoucher\cancel_reservation.ps1"
    )
    return ADD_GUIDE_ACTION_SCRIPT


async def addguide_action_script(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["action_script_path"] = update.message.text.strip()
    await update.message.reply_text(
        "Nama parameter/flag-nya apa? (sesuai yang dipakai script)\n"
        "Kalau lebih dari 1 parameter, pisahkan dengan koma.\n\n"
        "Contoh 1 parameter: -CsvFile\n"
        "Contoh banyak parameter: -ArticleId, -SourceId"
    )
    return ADD_GUIDE_ACTION_FLAG


async def addguide_action_flag(update: Update, context: ContextTypes.DEFAULT_TYPE):
    flags = []
    for part in update.message.text.split(","):
        f = part.strip()
        if not f:
            continue
        if not f.startswith("-"):
            f = "-" + f
        flags.append(f)
    if not flags:
        await update.message.reply_text("Nama parameternya kosong, coba lagi:")
        return ADD_GUIDE_ACTION_FLAG

    context.user_data["action_flags"] = flags

    if len(flags) == 1:
        keyboard = [
            [
                InlineKeyboardButton("Teks bebas", callback_data="atype_text"),
                InlineKeyboardButton("Upload file", callback_data="atype_file"),
            ]
        ]
        await update.message.reply_text(
            "Nilai parameter ini nanti diminta dalam bentuk apa?",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return ADD_GUIDE_ACTION_TYPE

    keyboard = [
        [InlineKeyboardButton("Satu-satu per parameter", callback_data="amode_separate")],
        [InlineKeyboardButton("Baris per baris (berpasangan)", callback_data="amode_paired")],
    ]
    await update.message.reply_text(
        "Cara input datanya nanti gimana?\n\n"
        "• Satu-satu per parameter: bot tanya tiap parameter terpisah "
        f"({', '.join(flags)}), boleh multi-nilai dipisah koma tiap parameter.\n"
        "• Baris per baris (berpasangan): kamu isi sekali jalan, satu baris = "
        f"satu kombinasi ({', '.join(f.lstrip('-') for f in flags)}), contoh:\n"
        "8000044321, SS20\n8000044322, SS21",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return ADD_GUIDE_ACTION_MODE


async def addguide_action_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    flags = context.user_data.get("action_flags", [])
    input_mode = "paired" if query.data == "amode_paired" else "separate"
    action = {
        "script_path": context.user_data.get("action_script_path", ""),
        "params": [{"flag": f, "type": "text"} for f in flags],
    }
    if input_mode == "paired":
        action["input_mode"] = "paired"
    context.user_data["guide_action"] = action
    return await addguide_show_preview(update, context)


async def addguide_action_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    param_type = "file" if query.data == "atype_file" else "text"
    flags = context.user_data.get("action_flags", [])
    context.user_data["guide_action"] = {
        "script_path": context.user_data.get("action_script_path", ""),
        "params": [{"flag": flags[0], "type": param_type}] if flags else [],
    }
    return await addguide_show_preview(update, context)


async def addguide_show_preview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    file_line = f"\nFile: {d['guide_attachment']['file_name']}" if d.get("guide_attachment") else ""
    action = d.get("guide_action")
    action_line = ""
    if action:
        param_bits = [
            f"{p['flag']} ({'file' if p.get('type') == 'file' else 'teks'})"
            for p in action.get("params", [])
        ]
        action_line = f"\nAction: {action['script_path']} | " + ", ".join(param_bits)
        if action.get("input_mode") == "paired":
            action_line += " | mode: baris berpasangan"
    preview = (
        "Konfirmasi guidance berikut:\n\n"
        f"Judul: {d['guide_title']}\n"
        f"Kata kunci: {', '.join(d['guide_keywords'])}\n"
        f"{file_line}"
        f"{action_line}\n"
        f"Isi:\n{d['guide_content'] or '(tidak ada teks deskripsi)'}\n\n"
        "Simpan guidance ini?"
    )
    keyboard = [
        [
            InlineKeyboardButton("Ya, simpan", callback_data="guide_save_yes"),
            InlineKeyboardButton("Batal", callback_data="guide_save_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(preview, reply_markup=markup)
    else:
        await update.message.reply_text(preview, reply_markup=markup)
    return ADD_GUIDE_CONFIRM


async def addguide_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "guide_save_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    d = context.user_data
    entry = guidance_store.add_guidance(
        d["guide_title"],
        d["guide_keywords"],
        d["guide_content"],
        d.get("guide_attachment"),
        d.get("guide_action"),
    )
    await query.edit_message_text(f"Guidance '{entry['title']}' berhasil disimpan. ✅ (ID: {entry['id']})")
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def delguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text("Belum ada guidance tersimpan.")
        return ConversationHandler.END
    buttons = [
        [InlineKeyboardButton(f"[{d['id']}] {d['title']}", callback_data=f"delg_{d['id']}")]
        for d in data
    ]
    await update.message.reply_text(
        "Pilih guidance yang mau dihapus:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return DEL_GUIDE_PICK


async def delguide_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("delg_", "")
    context.user_data["del_guide_id"] = gid
    entry = guidance_store.get_guidance(gid)
    title = entry["title"] if entry else gid
    keyboard = [
        [
            InlineKeyboardButton("Ya, hapus", callback_data="delg_yes"),
            InlineKeyboardButton("Batal", callback_data="delg_no"),
        ]
    ]
    await query.edit_message_text(
        f"Yakin mau hapus guidance '{title}'?", reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return DEL_GUIDE_CONFIRM


async def delguide_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "delg_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    gid = context.user_data.get("del_guide_id")
    ok = guidance_store.delete_guidance(gid)
    await query.edit_message_text(
        "Guidance berhasil dihapus. 🗑️" if ok else "Gagal menghapus (mungkin sudah dihapus)."
    )
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def editguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text("Belum ada guidance tersimpan.")
        return ConversationHandler.END
    buttons = [
        [InlineKeyboardButton(f"[{d['id']}] {d['title']}", callback_data=f"editg_{d['id']}")]
        for d in data
    ]
    await update.message.reply_text(
        "Pilih guidance yang mau diedit:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return EDIT_GUIDE_PICK


async def editguide_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("editg_", "")
    entry = guidance_store.get_guidance(gid)
    if not entry:
        await query.edit_message_text("Guidance tidak ditemukan (mungkin sudah dihapus).")
        return ConversationHandler.END
    context.user_data["edit_guide_id"] = gid
    context.user_data["edit_title"] = entry["title"]
    context.user_data["edit_keywords"] = entry.get("keywords", [])
    context.user_data["edit_content"] = entry.get("content", "")
    context.user_data["edit_attachment"] = entry.get("attachment")
    context.user_data["edit_action"] = entry.get("action")
    return await show_edit_menu(update, context)


async def show_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    file_line = ""
    if d.get("edit_attachment"):
        file_line = f"\nFile terlampir: {d['edit_attachment']['file_name']}"
    content_preview = d["edit_content"][:200] or "(kosong)"
    action = d.get("edit_action")
    action_line = ""
    if action:
        param_bits = [
            f"{p['flag']} ({'file' if p.get('type') == 'file' else 'teks'})"
            for p in _action_param_list(action)
        ]
        action_line = f"\nAction: {action.get('script_path','')} | " + ", ".join(param_bits)
        if action.get("input_mode") == "paired":
            action_line += " | mode: baris berpasangan"
    text = (
        "Mengedit guidance:\n\n"
        f"Judul: {d['edit_title']}\n"
        f"Kata kunci: {', '.join(d['edit_keywords'])}"
        f"{file_line}"
        f"{action_line}\n"
        f"Isi: {content_preview}\n\n"
        "Mau ubah bagian mana?"
    )
    keyboard = [
        [
            InlineKeyboardButton("✏️ Judul", callback_data="ef_title"),
            InlineKeyboardButton("✏️ Kata kunci", callback_data="ef_keywords"),
        ],
        [InlineKeyboardButton("✏️ Isi / File", callback_data="ef_content")],
        [
            InlineKeyboardButton(
                "⚙️ Hapus Action" if action else "⚙️ Tambah Action", callback_data="ef_action"
            )
        ],
        [
            InlineKeyboardButton("✅ Selesai & Simpan", callback_data="ef_done"),
            InlineKeyboardButton("Batal", callback_data="ef_cancel"),
        ],
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    return EDIT_GUIDE_MENU


async def editguide_menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "ef_cancel":
        await query.edit_message_text("Dibatalkan, tidak ada perubahan yang disimpan.")
        context.user_data.clear()
        return ConversationHandler.END

    if choice == "ef_done":
        d = context.user_data
        keywords = d["edit_keywords"]
        if isinstance(keywords, str):
            keywords = [k.strip() for k in keywords.split(",") if k.strip()]
        updates = {
            "title": d["edit_title"],
            "keywords": [k.strip().lower() for k in keywords],
            "content": d["edit_content"],
            "attachment": d.get("edit_attachment"),
            "action": d.get("edit_action"),
        }
        entry = guidance_store.update_guidance(d["edit_guide_id"], updates)
        if entry:
            await query.edit_message_text(f"Guidance '{entry['title']}' berhasil diperbarui. ✅")
        else:
            await query.edit_message_text("Gagal menyimpan perubahan (guidance mungkin sudah dihapus).")
        context.user_data.clear()
        return ConversationHandler.END

    if choice == "ef_action":
        if context.user_data.get("edit_action"):
            context.user_data["edit_action"] = None
            return await show_edit_menu(update, context)
        await query.edit_message_text(
            "Masukkan path lengkap file script-nya di laptop kamu.\n"
            r"Contoh: D:\promovoucher\cancel_reservation.ps1"
        )
        return EDIT_GUIDE_ACTION_SCRIPT

    if choice == "ef_title":
        await query.edit_message_text("Ketik judul baru:")
        return EDIT_GUIDE_TITLE
    if choice == "ef_keywords":
        await query.edit_message_text("Ketik kata kunci baru, pisahkan dengan koma:")
        return EDIT_GUIDE_KEYWORDS
    if choice == "ef_content":
        await query.edit_message_text(
            "Kirim isi/detail baru (teks biasa, atau upload file kalau ada)."
        )
        return EDIT_GUIDE_CONTENT


async def editguide_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["edit_title"] = update.message.text.strip()
    return await show_edit_menu(update, context)


async def editguide_keywords(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keywords = [k.strip() for k in update.message.text.split(",") if k.strip()]
    if not keywords:
        await update.message.reply_text("Kata kuncinya kosong, coba lagi (pisahkan dengan koma):")
        return EDIT_GUIDE_KEYWORDS
    context.user_data["edit_keywords"] = keywords
    return await show_edit_menu(update, context)


async def editguide_content(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document:
        doc = update.message.document
        context.user_data["edit_content"] = (update.message.caption or "").strip()
        context.user_data["edit_attachment"] = {"file_id": doc.file_id, "file_name": doc.file_name}
    else:
        context.user_data["edit_content"] = update.message.text.strip()
        context.user_data["edit_attachment"] = None
    return await show_edit_menu(update, context)


async def editguide_action_script(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["_edit_action_script_path"] = update.message.text.strip()
    await update.message.reply_text(
        "Nama parameter/flag-nya apa? (sesuai yang dipakai script)\n"
        "Kalau lebih dari 1 parameter, pisahkan dengan koma.\n\n"
        "Contoh 1 parameter: -CsvFile\n"
        "Contoh banyak parameter: -ArticleId, -SourceId"
    )
    return EDIT_GUIDE_ACTION_FLAG


async def editguide_action_flag(update: Update, context: ContextTypes.DEFAULT_TYPE):
    flags = []
    for part in update.message.text.split(","):
        f = part.strip()
        if not f:
            continue
        if not f.startswith("-"):
            f = "-" + f
        flags.append(f)
    if not flags:
        await update.message.reply_text("Nama parameternya kosong, coba lagi:")
        return EDIT_GUIDE_ACTION_FLAG

    context.user_data["_edit_action_flags"] = flags

    if len(flags) == 1:
        keyboard = [
            [
                InlineKeyboardButton("Teks bebas", callback_data="etype_text"),
                InlineKeyboardButton("Upload file", callback_data="etype_file"),
            ]
        ]
        await update.message.reply_text(
            "Nilai parameter ini nanti diminta dalam bentuk apa?",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return EDIT_GUIDE_ACTION_TYPE

    keyboard = [
        [InlineKeyboardButton("Satu-satu per parameter", callback_data="emode_separate")],
        [InlineKeyboardButton("Baris per baris (berpasangan)", callback_data="emode_paired")],
    ]
    await update.message.reply_text(
        "Cara input datanya nanti gimana?\n\n"
        "• Satu-satu per parameter: bot tanya tiap parameter terpisah "
        f"({', '.join(flags)}), boleh multi-nilai dipisah koma tiap parameter.\n"
        "• Baris per baris (berpasangan): kamu isi sekali jalan, satu baris = "
        f"satu kombinasi ({', '.join(f.lstrip('-') for f in flags)}), contoh:\n"
        "8000044321, SS20\n8000044322, SS21",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return EDIT_GUIDE_ACTION_MODE


async def editguide_action_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    flags = context.user_data.get("_edit_action_flags", [])
    input_mode = "paired" if query.data == "emode_paired" else "separate"
    action = {
        "script_path": context.user_data.pop("_edit_action_script_path", ""),
        "params": [{"flag": f, "type": "text"} for f in flags],
    }
    if input_mode == "paired":
        action["input_mode"] = "paired"
    context.user_data["edit_action"] = action
    return await show_edit_menu(update, context)


async def editguide_action_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    param_type = "file" if query.data == "etype_file" else "text"
    flags = context.user_data.get("_edit_action_flags", [])
    context.user_data["edit_action"] = {
        "script_path": context.user_data.pop("_edit_action_script_path", ""),
        "params": [{"flag": flags[0], "type": param_type}] if flags else [],
    }
    return await show_edit_menu(update, context)


# ---------------- reminder ----------------

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Terjadi error tak terduga:", exc_info=context.error)

    if isinstance(context.error, (TimedOut, NetworkError)):
        return

    target_chat_id = config.TELEGRAM_USER_ID
    target_thread_id = None
    if isinstance(update, Update) and update.effective_chat:
        target_chat_id = update.effective_chat.id
        target_thread_id = _thread_id_from_update(update)

    try:
        await context.bot.send_message(
            chat_id=target_chat_id,
            message_thread_id=target_thread_id,
            text=(
                "⚠️ Terjadi error saat memproses perintah terakhir kamu:\n"
                f"{context.error}\n\n"
                "Coba ulangi lagi, atau ketik /cancel lalu mulai ulang."
            ),
        )
    except Exception:
        logger.exception("Gagal mengirim notifikasi error ke user")


SDP_NOTIFY_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sdp_notify_state.json")


def _load_sdp_notify_state() -> dict:
    try:
        with open(SDP_NOTIFY_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_seen_id": 0}


def _save_sdp_notify_state(state: dict):
    with open(SDP_NOTIFY_STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f)


async def check_new_sdp_tickets(context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    state = _load_sdp_notify_state()
    last_seen_id = state.get("last_seen_id", 0)

    try:
        requests_list = await asyncio.to_thread(
            sdp.list_requests, 30, None, config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal cek tiket baru ServiceDesk Plus: {e}")
        return

    max_id_seen = last_seen_id
    new_items = []
    for req in requests_list:
        try:
            rid = int(req.get("id", 0))
        except (TypeError, ValueError):
            continue
        max_id_seen = max(max_id_seen, rid)
        if rid > last_seen_id:
            new_items.append((rid, req))

    if last_seen_id == 0:
        state["last_seen_id"] = max_id_seen
        _save_sdp_notify_state(state)
        return

    new_items.sort(key=lambda pair: pair[0])
    for _, req in new_items:
        group_name = (req.get("group") or {}).get("name", "")
        text = f"*New Ticket ({group_name}):\n\n{_format_sdp_request_line(req)}"
        await _broadcast_notify(context, text)

    if new_items:
        state["last_seen_id"] = max_id_seen
        _save_sdp_notify_state(state)


@restricted
async def sdreminder_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        await update.message.reply_text(
            "Fitur ini butuh SDP_NOTIFY_GROUPS diisi dulu di .env (nama-nama group yang mau dipantau)."
        )
        return

    parts = update.message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""
    state = _load_sdp_notify_state()

    if not arg or arg == "status":
        interval = state.get("open_reminder_interval_minutes", config.SDP_OPEN_REMINDER_DEFAULT_MINUTES)
        if interval and interval > 0:
            await update.message.reply_text(
                f"Reminder tiket Open sedang AKTIF, tiap {interval} menit.\n\n"
                "Ubah interval: /sdreminder <menit>\n"
                "Matikan: /sdreminder off"
            )
        else:
            await update.message.reply_text(
                "Reminder tiket Open sedang NONAKTIF.\n\n"
                "Aktifkan: /sdreminder <menit>, contoh: /sdreminder 30"
            )
        return

    if arg == "off":
        state["open_reminder_interval_minutes"] = 0
        _save_sdp_notify_state(state)
        await update.message.reply_text("Reminder tiket Open dimatikan.")
        return

    if not arg.isdigit() or int(arg) <= 0:
        await update.message.reply_text(
            "Ketik salah satu:\n"
            "/sdreminder <menit> - aktifkan/ubah interval, contoh: /sdreminder 30\n"
            "/sdreminder off - matikan\n"
            "/sdreminder status - cek status sekarang"
        )
        return

    minutes = int(arg)
    state["open_reminder_interval_minutes"] = minutes
    state["last_open_reminder_at"] = None
    _save_sdp_notify_state(state)
    await update.message.reply_text(
        f"Oke, reminder tiket Open sekarang aktif tiap {minutes} menit untuk group:\n"
        f"{', '.join(config.SDP_NOTIFY_GROUPS)}"
    )


async def check_open_ticket_reminders(context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    state = _load_sdp_notify_state()
    interval = state.get("open_reminder_interval_minutes", config.SDP_OPEN_REMINDER_DEFAULT_MINUTES)
    if not interval or interval <= 0:
        return

    now = dt.datetime.now(TZ)
    last_at_str = state.get("last_open_reminder_at")
    if last_at_str:
        try:
            last_at = dt.datetime.fromisoformat(last_at_str)
        except ValueError:
            last_at = None
    else:
        last_at = None

    if last_at and (now - last_at) < dt.timedelta(minutes=interval):
        return

    try:
        open_tickets = await asyncio.to_thread(
            sdp.list_requests, 50, "Open", config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal cek tiket Open untuk reminder: {e}")
        return

    state["last_open_reminder_at"] = now.isoformat()
    _save_sdp_notify_state(state)

    if not open_tickets:
        return

    lines = [f"⏰ Reminder: masih ada {len(open_tickets)} tiket berstatus Open:", ""]
    for req in open_tickets[:15]:
        lines.append(_format_sdp_request_line(req))
        lines.append("")
    if len(open_tickets) > 15:
        lines.append(f"...dan {len(open_tickets) - 15} tiket Open lainnya.")
    await _broadcast_notify(context, _safe_telegram_text("\n".join(lines).rstrip()))


async def _broadcast_notify(context: ContextTypes.DEFAULT_TYPE, text: str):
    for chat_id, thread_id in config.notify_targets():
        try:
            # Gunakan parse_mode=ParseMode.HTML atau hapus parse_mode agar aman dari error parsing Markdown
            await context.bot.send_message(
                chat_id=chat_id, 
                text=text, 
                message_thread_id=thread_id, 
                parse_mode=ParseMode.HTML
            )
        except Exception:
            logger.exception(f"Gagal kirim notifikasi ke chat_id={chat_id} thread_id={thread_id}")


async def daily_reminder(context: ContextTypes.DEFAULT_TYPE):
    await _broadcast_notify(context, "⏰ Jangan lupa isi logwork hari ini ya! Ketik /log untuk mulai.")


async def interval_reminder(context: ContextTypes.DEFAULT_TYPE):
    now = dt.datetime.now(TZ)
    if not (config.REMINDER_START_HOUR <= now.hour < config.REMINDER_END_HOUR):
        return
    await _broadcast_notify(context, "⏰ Reminder: jangan lupa isi logwork ya! Ketik /log untuk mulai.")


# ---------------- main ----------------

def main():
    config.validate()

    import asyncio

    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    app = (
        Application.builder()
        .token(config.TELEGRAM_BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .pool_timeout(30)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("chatid", chatid_command))

    myjira_conv = ConversationHandler(
        entry_points=[CommandHandler("myjira", myjira_command)],
        states={
            MYJIRA_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, myjira_email)],
            MYJIRA_TOKEN: [MessageHandler(filters.TEXT & ~filters.COMMAND, myjira_token)],
            MYJIRA_CONFIRM: [CallbackQueryHandler(myjira_confirm, pattern="^myjira_save_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(myjira_conv)
    app.add_handler(CommandHandler("today", today_summary))
    app.add_handler(CommandHandler("week", week_summary))

    log_conv = ConversationHandler(
        entry_points=[CommandHandler("log", log_start)],
        states={
            LOG_ISSUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_issue)],
            LOG_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_time)],
            LOG_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_desc)],
            LOG_DATE: [
                CallbackQueryHandler(log_date_choice, pattern="^date_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, log_date_manual),
            ],
            LOG_CONFIRM: [CallbackQueryHandler(log_confirm, pattern="^confirm_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    log_conv = ConversationHandler(
        entry_points=[CommandHandler("log", log_start)],
        states={
            LOG_ISSUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_issue)],
            LOG_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_time)],
            LOG_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_desc)],
            LOG_DATE: [
                CallbackQueryHandler(log_date_choice, pattern="^date_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, log_date_manual),
            ],
            LOG_CONFIRM: [CallbackQueryHandler(log_confirm, pattern="^confirm_")],
            # Tambahkan baris ini untuk menangkap tombol konfirmasi bulk logwork:
            LOG_BULK_CONFIRM: [CallbackQueryHandler(bulk_log_confirm_callback, pattern="^bulk_confirm_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    
    app.add_handler(log_conv)

    edit_conv = ConversationHandler(
        entry_points=[CommandHandler("edit", edit_start)],
        states={
            PICK_ISSUE_FOR_EDIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_pick_issue)],
            PICK_WORKLOG_EDIT: [CallbackQueryHandler(edit_pick_worklog, pattern="^wl_")],
            EDIT_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_time)],
            EDIT_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_desc)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(edit_conv)

    delete_conv = ConversationHandler(
        entry_points=[CommandHandler("delete", delete_start)],
        states={
            PICK_ISSUE_FOR_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_pick_issue)],
            PICK_WORKLOG_DELETE: [CallbackQueryHandler(delete_pick_worklog, pattern="^wl_")],
            CONFIRM_DELETE: [CallbackQueryHandler(delete_confirm, pattern="^del_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(delete_conv)
    app.add_handler(CommandHandler("tasks", list_tasks))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("sdtickets", sdtickets_command))
    app.add_handler(CommandHandler("sdticket", sdticket_command))
    app.add_handler(CommandHandler("sdreminder", sdreminder_command))
    app.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND & filters.Regex(r"^([A-Za-z]{2,10}-|[A-Z]{2,10})$"),
            list_tasks,
        )
    )

    # Guidance / knowledge base
    app.add_handler(CommandHandler("guide", guide_command))
    app.add_handler(CommandHandler("listguide", list_guide))
    app.add_handler(CallbackQueryHandler(guide_pick_callback, pattern="^gid_"))

    addguide_conv = ConversationHandler(
        entry_points=[CommandHandler("addguide", addguide_start)],
        states={
            ADD_GUIDE_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_title)],
            ADD_GUIDE_KEYWORDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_keywords)],
            ADD_GUIDE_CONTENT: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, addguide_content
                )
            ],
            ADD_GUIDE_ACTION_ASK: [CallbackQueryHandler(addguide_action_ask, pattern="^ga_")],
            ADD_GUIDE_ACTION_SCRIPT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_action_script)
            ],
            ADD_GUIDE_ACTION_FLAG: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_action_flag)
            ],
            ADD_GUIDE_ACTION_TYPE: [CallbackQueryHandler(addguide_action_type, pattern="^atype_")],
            ADD_GUIDE_ACTION_MODE: [CallbackQueryHandler(addguide_action_mode, pattern="^amode_")],
            ADD_GUIDE_CONFIRM: [CallbackQueryHandler(addguide_confirm, pattern="^guide_save_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(addguide_conv)

    delguide_conv = ConversationHandler(
        entry_points=[CommandHandler("delguide", delguide_start)],
        states={
            DEL_GUIDE_PICK: [CallbackQueryHandler(delguide_pick, pattern="^delg_")],
            DEL_GUIDE_CONFIRM: [CallbackQueryHandler(delguide_confirm, pattern="^delg_(yes|no)$")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(delguide_conv)

    editguide_conv = ConversationHandler(
        entry_points=[CommandHandler("editguide", editguide_start)],
        states={
            EDIT_GUIDE_PICK: [CallbackQueryHandler(editguide_pick, pattern="^editg_")],
            EDIT_GUIDE_MENU: [CallbackQueryHandler(editguide_menu_choice, pattern="^ef_")],
            EDIT_GUIDE_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_title)],
            EDIT_GUIDE_KEYWORDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_keywords)],
            EDIT_GUIDE_CONTENT: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, editguide_content
                )
            ],
            EDIT_GUIDE_ACTION_SCRIPT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_action_script)
            ],
            EDIT_GUIDE_ACTION_FLAG: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_action_flag)
            ],
            EDIT_GUIDE_ACTION_TYPE: [
                CallbackQueryHandler(editguide_action_type, pattern="^etype_")
            ],
            EDIT_GUIDE_ACTION_MODE: [
                CallbackQueryHandler(editguide_action_mode, pattern="^emode_")
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(editguide_conv)

    run_action_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(run_action_start, pattern="^runaction_"),
            CommandHandler("run", run_command),
        ],
        states={
            RUN_ACTION_PARAM: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, run_action_param
                )
            ],
            RUN_ACTION_CONFIRM: [
                CallbackQueryHandler(run_action_confirm, pattern="^runconfirm_")
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    app.add_handler(run_action_conv)

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, guide_free_text))
    app.add_error_handler(error_handler)

    # --------------------------------------------------------------------------
    # BACKGROUND JOBS SERVICEDESK PLUS & REMINDERS
    # --------------------------------------------------------------------------
    if sdp and config.SDP_NOTIFY_GROUPS:
        #Job Otomasi Assign Tiket ServiceDesk Plus (Berjalan tiap 5 menit)
        app.job_queue.run_repeating(
            auto_assign_sdp_tickets,
            interval=dt.timedelta(minutes=5),
            first=10,
        )
        logger.info(
            f"Otomasi Auto-Assign Tiket SDP aktif (tiap 5 menit) untuk grup: {', '.join(config.SDP_NOTIFY_GROUPS)}"
        )

        # Job Cek Tiket Baru
        app.job_queue.run_repeating(
            check_new_sdp_tickets,
            interval=dt.timedelta(minutes=config.SDP_NOTIFY_INTERVAL_MINUTES),
            first=15,
        )
        logger.info(
            f"Notifikasi tiket baru SDP aktif tiap {config.SDP_NOTIFY_INTERVAL_MINUTES} menit "
            f"untuk group: {', '.join(config.SDP_NOTIFY_GROUPS)}"
        )

        # Job Reminder Tiket Open
        app.job_queue.run_repeating(
            check_open_ticket_reminders,
            interval=dt.timedelta(minutes=1),
            first=30,
        )

    if config.REMINDER_INTERVAL_MINUTES > 0:
        app.job_queue.run_repeating(
            interval_reminder,
            interval=dt.timedelta(minutes=config.REMINDER_INTERVAL_MINUTES),
            first=10,
        )
        logger.info(
            f"Reminder aktif tiap {config.REMINDER_INTERVAL_MINUTES} menit, "
            f"jam {config.REMINDER_START_HOUR}:00-{config.REMINDER_END_HOUR}:00"
        )
    else:
        app.job_queue.run_daily(
            daily_reminder,
            time=dt.time(hour=config.REMINDER_HOUR, minute=config.REMINDER_MINUTE, tzinfo=TZ),
        )
        logger.info(f"Reminder aktif tiap hari jam {config.REMINDER_HOUR}:{config.REMINDER_MINUTE:02d}")

    logger.info("Bot mulai berjalan...")
    app.run_polling()


if __name__ == "__main__":
    main()