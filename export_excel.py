import os
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "No",
    "Task SD",
    "Task Dev",
    "Summary",
    "Created",
    "Status SD",
    "Status Dev",
    "Tanggal Dev Done",
    "Status QA",
    "Tanggal QA Done",
]
COLUMN_WIDTHS = [5, 12, 14, 12, 55, 14, 14, 18, 18, 18]


def build_export_excel(project_key: str, rows: list) -> str:
    """Bikin file .xlsx dari data report project (list of dict dengan key:
    key, task_dev, summary, status_dev, status_qa, created). Return path file."""
    wb = Workbook()
    ws = wb.active
    ws.title = project_key[:31] or "Report"  # nama sheet max 31 karakter

    ws.append(HEADERS)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial")
    for i, row in enumerate(rows, start=1):
        ws.append(
            [
                i,
                row.get("key", ""),
                row.get("task_dev", ""),
                row.get("summary", ""),
                row.get("created", ""),
                row.get("status_sd", ""),
                row.get("status_dev", ""),
                row.get("dev_done_date", ""),
                row.get("status_qa", ""),
                row.get("qa_done_date", ""),
            ]
        )
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=i + 1, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_cells in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(HEADERS)):
        for cell in row_cells:
            cell.border = border

    ws.freeze_panes = "A2"

    filepath = os.path.join(tempfile.gettempdir(), f"report_{project_key}.xlsx")
    wb.save(filepath)
    return filepath
