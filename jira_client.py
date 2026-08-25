import os
import tempfile

from openpyxl import Workbook

import datetime as dt
from zoneinfo import ZoneInfo

import requests

import config


class JiraError(Exception):
    pass


class JiraClient:
    def __init__(self, email: str = None, api_token: str = None):
        self.base_url = config.JIRA_BASE_URL
        self.email = email or config.JIRA_EMAIL
        self.auth = (self.email, api_token or config.JIRA_API_TOKEN)
        self.session = requests.Session()
        self.session.auth = self.auth
        self.session.headers.update({"Content-Type": "application/json"})
        self.tz = ZoneInfo(config.TIMEZONE)

    # ---------- helpers ----------

    def _url(self, path):
        return f"{self.base_url}{path}"

    @staticmethod
    def _format_started(date_obj: dt.date, tz: ZoneInfo, time_obj: dt.time = None) -> str:
        """Format tanggal (+ jam custom kalau ada) sesuai format yang diminta Jira."""
        if time_obj is None:
            time_obj = dt.datetime.now(tz).time()
        started = dt.datetime.combine(date_obj, time_obj, tzinfo=tz)
        s = started.strftime("%Y-%m-%dT%H:%M:%S.%f%z")
        date_part, rest = s.split(".")
        micro, tzpart = rest[:6], rest[6:]
        millis = micro[:3]
        return f"{date_part}.{millis}{tzpart}"

    @staticmethod
    def _to_adf(text: str) -> dict:
        return {
            "type": "doc",
            "version": 1,
            "content": [
                {"type": "paragraph", "content": [{"type": "text", "text": text}]}
            ],
        }

    @staticmethod
    def _adf_to_text(adf) -> str:
        if not adf or not isinstance(adf, dict):
            return ""
        parts = []
        for block in adf.get("content", []):
            for node in block.get("content", []):
                if node.get("type") == "text":
                    parts.append(node.get("text", ""))
        return " ".join(parts).strip()

    def _request(self, method: str, url: str, **kwargs) -> requests.Response:
        try:
            resp = self.session.request(method, url, timeout=20, **kwargs)
        except requests.exceptions.RequestException as e:
            raise JiraError(f"Gagal terhubung ke Jira ({e.__class__.__name__}): {e}")
        self._check(resp)
        return resp

    def _safe_json(self, resp: requests.Response):
        if not resp.content:
            return {}
        try:
            return resp.json()
        except ValueError:
            raise JiraError(
                f"Respons dari Jira tidak sesuai dugaan (bukan JSON): {resp.text[:200]!r}"
            )

    def _check(self, resp: requests.Response):
        if not resp.ok:
            try:
                detail = resp.json()
            except Exception:
                detail = resp.text
            raise JiraError(f"Jira error {resp.status_code}: {detail}")

    # ---------- worklog CRUD ----------

    def add_worklog(
        self,
        issue_key: str,
        time_spent: str,
        comment: str,
        date_obj: dt.date,
        start_time: dt.time = None,
    ):
        payload = {
            "timeSpent": time_spent,
            "comment": self._to_adf(comment) if comment else self._to_adf(""),
            "started": self._format_started(date_obj, self.tz, start_time),
        }
        resp = self._request(
            "POST", self._url(f"/rest/api/3/issue/{issue_key}/worklog"), json=payload
        )
        return self._safe_json(resp)

    def get_worklogs(self, issue_key: str) -> list:
        resp = self._request("GET", self._url(f"/rest/api/3/issue/{issue_key}/worklog"))
        return self._safe_json(resp).get("worklogs", [])

    def get_my_recent_worklogs(self, issue_key: str, days: int = 14) -> list:
        """Ambil worklog milik user (email akun ini sendiri) pada issue tertentu, N hari terakhir."""
        worklogs = self.get_worklogs(issue_key)
        cutoff = dt.datetime.now(self.tz) - dt.timedelta(days=days)
        mine = []
        for w in worklogs:
            author_email = (w.get("author") or {}).get("emailAddress", "")
            if author_email.lower() != self.email.lower():
                continue
            started = dt.datetime.strptime(w["started"][:19], "%Y-%m-%dT%H:%M:%S")
            started = started.replace(tzinfo=self.tz)
            if started >= cutoff:
                mine.append(w)
        mine.sort(key=lambda w: w["started"], reverse=True)
        return mine

    def update_worklog(self, issue_key: str, worklog_id: str, time_spent: str = None, comment: str = None):
        payload = {}
        if time_spent:
            payload["timeSpent"] = time_spent
        if comment is not None:
            payload["comment"] = self._to_adf(comment)
        resp = self._request(
            "PUT",
            self._url(f"/rest/api/3/issue/{issue_key}/worklog/{worklog_id}"),
            json=payload,
        )
        return self._safe_json(resp)

    def delete_worklog(self, issue_key: str, worklog_id: str):
        self._request(
            "DELETE", self._url(f"/rest/api/3/issue/{issue_key}/worklog/{worklog_id}")
        )

    # ---------- daftar task per project ----------

    def search_project_issues(self, project_key: str, max_results: int = 50) -> list:
        jql = f'project = "{project_key}" ORDER BY updated DESC'
        resp = self._request(
            "GET",
            self._url("/rest/api/3/search/jql"),
            params={"jql": jql, "fields": "summary,status", "maxResults": max_results},
        )
        return self._safe_json(resp).get("issues", [])

    def get_project_export(self, project_key: str, max_results: int = 100) -> list:
        """Ambil data report per task di suatu project, termasuk linked issue
        'relates to' untuk kolom Task Dev / Status Dev / Status QA.

        Aturan: di antara linked issue bertipe 'relates to', yang judulnya
        diawali kata 'Testing' dianggap task QA (dipakai Status QA-nya),
        sisanya (non-Testing pertama yang ditemukan) dianggap Task Dev."""
        jql = f'project = "{project_key}" ORDER BY created DESC'
        resp = self._request(
            "GET",
            self._url("/rest/api/3/search/jql"),
            params={
                "jql": jql,
                "fields": "summary,created,status,issuelinks",
                "maxResults": max_results,
            },
        )
        issues = self._safe_json(resp).get("issues", [])

        rows = []
        for issue in issues:
            fields = issue.get("fields", {})
            status_sd = (fields.get("status") or {}).get("name", "")
            task_dev, status_dev, status_qa, qa_key = "", "", "", ""

            for link in fields.get("issuelinks", []):
                type_name = (link.get("type", {}).get("name") or "").lower()
                if "relate" not in type_name:
                    continue
                linked = link.get("outwardIssue") or link.get("inwardIssue")
                if not linked:
                    continue
                linked_summary = (linked.get("fields", {}).get("summary") or "").strip()
                linked_status = (linked.get("fields", {}).get("status") or {}).get("name", "")
                if linked_summary.lower().startswith("testing"):
                    status_qa = linked_status
                    qa_key = linked.get("key", "")
                elif not task_dev:
                    task_dev = linked.get("key", "")
                    status_dev = linked_status

            dev_done_date = self.get_status_done_date(task_dev) if task_dev else ""
            qa_done_date = self.get_status_done_date(qa_key) if qa_key else ""

            rows.append(
                {
                    "key": issue.get("key", ""),
                    "status_sd": status_sd,
                    "task_dev": task_dev,
                    "summary": fields.get("summary", ""),
                    "status_dev": status_dev,
                    "status_qa": status_qa,
                    "created": self._format_created_display(fields.get("created", "")),
                    "dev_done_date": dev_done_date,
                    "qa_done_date": qa_done_date,
                }
            )
        return rows

    def get_status_done_date(self, issue_key: str) -> str:
        """Cari tanggal terakhir status issue ini berubah jadi 'Done' lewat
        history/changelog. Return string sudah terformat, atau '' kalau
        belum pernah/tidak ketemu."""
        try:
            resp = self._request(
                "GET",
                self._url(f"/rest/api/3/issue/{issue_key}"),
                params={"expand": "changelog", "fields": "status"},
            )
        except JiraError:
            return ""
        data = self._safe_json(resp)
        histories = data.get("changelog", {}).get("histories", [])
        done_raw = ""
        for h in histories:
            for item in h.get("items", []):
                if item.get("field") == "status" and (item.get("toString") or "").strip().lower() == "done":
                    done_raw = h.get("created", "")  # ambil yang paling akhir kalau lebih dari 1
        return self._format_created_display(done_raw) if done_raw else ""

    @staticmethod
    def _format_created_display(raw: str) -> str:
        """Format '2026-06-15T14:50:00.000+0700' jadi '2026-06-15 14:50:00'."""
        if not raw:
            return ""
        try:
            parsed = dt.datetime.strptime(raw[:19], "%Y-%m-%dT%H:%M:%S")
        except ValueError:
            return raw
        return parsed.strftime("%Y-%m-%d %H:%M:%S")

    # ---------- rekap ----------

    def _search_issues_with_worklog_in_range(self, start_date: dt.date, end_date: dt.date) -> list:
        jql = (
            f'worklogAuthor = currentUser() AND worklogDate >= "{start_date.isoformat()}" '
            f'AND worklogDate <= "{end_date.isoformat()}"'
        )
        resp = self._request(
            "GET",
            self._url("/rest/api/3/search/jql"),
            params={"jql": jql, "fields": "summary", "maxResults": 100},
        )
        return self._safe_json(resp).get("issues", [])

    def get_summary(self, start_date: dt.date, end_date: dt.date) -> list:
        """Return list of dict: {issue_key, summary, total_seconds, entries: [...]}"""
        issues = self._search_issues_with_worklog_in_range(start_date, end_date)
        result = []
        for issue in issues:
            key = issue["key"]
            title = issue["fields"]["summary"]
            worklogs = self.get_worklogs(key)
            entries = []
            total_seconds = 0
            for w in worklogs:
                author_email = (w.get("author") or {}).get("emailAddress", "")
                if author_email.lower() != self.email.lower():
                    continue
                started_date = dt.datetime.strptime(w["started"][:10], "%Y-%m-%d").date()
                if not (start_date <= started_date <= end_date):
                    continue
                total_seconds += w.get("timeSpentSeconds", 0)
                entries.append(
                    {
                        "id": w["id"],
                        "date": started_date.isoformat(),
                        "time_spent": w.get("timeSpent", ""),
                        "comment": self._adf_to_text(w.get("comment")),
                    }
                )
            if entries:
                result.append(
                    {
                        "issue_key": key,
                        "summary": title,
                        "total_seconds": total_seconds,
                        "entries": entries,
                    }
                )
        return result

    def get_issue_summary_title(self, issue_key: str) -> str:
        resp = self._request(
            "GET", self._url(f"/rest/api/3/issue/{issue_key}"), params={"fields": "summary"}
        )
        return self._safe_json(resp)["fields"]["summary"]
        
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = [
    "No",
    "Task SD",
    "Status SD",
    "Task Dev",
    "Summary",
    "Status Dev",
    "Status QA",
    "Created",
    "Tanggal Dev Done",
    "Tanggal QA Done",
]
COLUMN_WIDTHS = [5, 12, 14, 12, 55, 14, 14, 18, 18, 18]


def build_export_excel(project_key: str, rows: list) -> str:
    """Bikin file .xlsx dari data report project (list of dict dengan key:
    key, task_dev, summary, status_dev, status_qa, created). Return path file."""
    wb = Workbook()
    ws = wb.active
    ws.title = project_key[:31] or "Report"  # nama sheet max 31 karakter

    ws.append(HEADERS)

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial")
    for i, row in enumerate(rows, start=1):
        ws.append(
            [
                i,
                row.get("key", ""),
                row.get("status_sd", ""),
                row.get("task_dev", ""),
                row.get("summary", ""),
                row.get("status_dev", ""),
                row.get("status_qa", ""),
                row.get("created", ""),
                row.get("dev_done_date", ""),
                row.get("qa_done_date", ""),
            ]
        )
        for col_idx in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=i + 1, column=col_idx)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))

    for idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_cells in ws.iter_rows(min_row=1, max_row=len(rows) + 1, min_col=1, max_col=len(HEADERS)):
        for cell in row_cells:
            cell.border = border

    ws.freeze_panes = "A2"

    filepath = os.path.join(tempfile.gettempdir(), f"report_{project_key}.xlsx")
    wb.save(filepath)
    return filepath
