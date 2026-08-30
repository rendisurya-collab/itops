import asyncio
import datetime as dt
import html
import json
import logging
import os
import re
import subprocess
import tempfile
import warnings
from zoneinfo import ZoneInfo

import openpyxl
import pytz
import requests
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ParseMode
from telegram.error import BadRequest, Conflict, NetworkError, TimedOut
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)
from telegram.warnings import PTBUserWarning

import config
import guidance_store
import kb_store
import user_accounts
from export_excel import build_export_excel
from grab_client import GrabClient, GrabError, GrabNotFound
from jira_client import JiraClient, JiraError
from servicedesk_client import SDPClient, SDPError

# Sembunyikan peringatan PTBUserWarning agar terminal bersih
warnings.filterwarnings("ignore", category=PTBUserWarning)

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)

jira = JiraClient()  # dipakai untuk fitur yang tidak butuh atribusi personal (/tasks, /export)
sdp = SDPClient() if config.sdp_configured() else None
grab = GrabClient() if config.grab_configured() else None

_JIRA_CLIENT_CACHE = {}  # telegram_user_id (str) -> JiraClient

TZ = ZoneInfo(config.TIMEZONE)

# ==============================================================================
# KONFIGURASI FILE & LOGIKA SHIFT / ROUND-ROBIN / AUDIT LOG SERVICEDESK PLUS
# ==============================================================================
TECH_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "technicians.json")
SDP_ASSIGN_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sdp_assign_state.json")
TICKET_LOG_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ticket_activity.log")


def write_ticket_log(ticket_id: str, tech_name: str, tech_email: str, status: str, shift: str, is_success: bool, error_msg: str = ""):
    """Menyimpan setiap riwayat aksi update tiket ke file log DAN Google Sheets"""
    timestamp = dt.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
    status_tag = "SUCCESS" if is_success else "FAILED"
    
    log_line = (
        f"[{timestamp}] [{status_tag}] Ticket ID: #{ticket_id} | "
        f"Assigned To: {tech_name} ({tech_email}) | Status: {status} | Shift: {shift}"
    )
    if not is_success and error_msg:
        log_line += f" | Error: {error_msg}"
        
    # Tulis ke file lokal
    try:
        with open(TICKET_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(log_line + "\n")
    except Exception as e:
        logger.error(f"Gagal menulis ke file log tiket: {e}")

    # Tulis ke Google Sheets (async-safe, jalan di background)
    try:
        _append_to_google_sheets([
            timestamp, str(ticket_id), tech_name, tech_email, status, shift, status_tag, error_msg
        ])
    except Exception as e:
        logger.error(f"Gagal menulis ke Google Sheets: {type(e).__name__}: {e}")


def _append_to_google_sheets(row_data: list):
    """Append satu baris ke Google Sheets."""
    creds_json = config.GOOGLE_SHEETS_CREDENTIALS
    sheet_id = config.GOOGLE_SHEETS_SPREADSHEET_ID

    if not creds_json:
        logger.warning("GOOGLE_SHEETS_CREDENTIALS env kosong, skip logging ke Sheets.")
        return
    if not sheet_id:
        logger.warning("GOOGLE_SHEETS_SPREADSHEET_ID env kosong, skip logging ke Sheets.")
        return

    import gspread
    from google.oauth2.service_account import Credentials

    try:
        creds_dict = json.loads(creds_json)
    except json.JSONDecodeError as e:
        logger.error(f"GOOGLE_SHEETS_CREDENTIALS bukan JSON valid: {e}")
        return

    credentials = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )

    gc = gspread.authorize(credentials)
    spreadsheet = gc.open_by_key(sheet_id)
    worksheet = spreadsheet.sheet1
    worksheet.append_row(row_data, value_input_option="USER_ENTERED")
    logger.info(f"✅ Berhasil append ke Google Sheets: {row_data[:3]}...")


# ==============================================================================
# GOOGLE SHEETS: JADWAL SHIFT HELPERS
# ==============================================================================
SHIFT_SHEET_TAB = "JadwalShift"


def _get_gspread_client():
    """Return authorized gspread client, atau None jika credentials kosong."""
    import gspread
    from google.oauth2.service_account import Credentials

    creds_json = config.GOOGLE_SHEETS_CREDENTIALS
    sheet_id = config.GOOGLE_SHEETS_SPREADSHEET_ID

    if not creds_json or not sheet_id:
        return None, None

    creds_dict = json.loads(creds_json)
    credentials = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets"]
    )
    gc = gspread.authorize(credentials)
    spreadsheet = gc.open_by_key(sheet_id)
    return gc, spreadsheet


def _get_shift_worksheet():
    """Return worksheet 'JadwalShift', buat jika belum ada."""
    _, spreadsheet = _get_gspread_client()
    if spreadsheet is None:
        raise RuntimeError("Google Sheets credentials atau spreadsheet ID kosong.")

    try:
        ws = spreadsheet.worksheet(SHIFT_SHEET_TAB)
    except Exception:
        # Buat tab baru dengan header
        ws = spreadsheet.add_worksheet(title=SHIFT_SHEET_TAB, rows=500, cols=5)
        ws.append_row(["Tanggal", "Jam Mulai", "Jam Selesai", "Shift", "Teknisi"], value_input_option="USER_ENTERED")

    return ws


def _sheets_read_all_shifts() -> list:
    """Baca semua baris jadwal shift dari Google Sheets. Return list of dict."""
    ws = _get_shift_worksheet()
    rows = ws.get_all_values()

    results = []
    for row in rows[1:]:  # skip header
        if len(row) < 5 or not row[0]:
            continue
        try:
            tgl = dt.datetime.strptime(row[0].strip(), "%Y-%m-%d").date()
        except ValueError:
            continue
        results.append({
            "tanggal": tgl,
            "jam_mulai": row[1].strip(),
            "jam_selesai": row[2].strip(),
            "shift": row[3].strip(),
            "teknisi": row[4].strip(),
        })

    results.sort(key=lambda x: (x["tanggal"], x["jam_mulai"]))
    return results


def _sheets_read_shifts(tanggal_start: dt.date = None, tanggal_end: dt.date = None) -> list:
    """Baca jadwal shift dari Google Sheets dalam rentang tanggal."""
    all_rows = _sheets_read_all_shifts()
    filtered = []
    for r in all_rows:
        if tanggal_start and r["tanggal"] < tanggal_start:
            continue
        if tanggal_end and r["tanggal"] > tanggal_end:
            continue
        filtered.append(r)
    return filtered


def _sheets_upsert_shift(tanggal: dt.date, jam_mulai: str, jam_selesai: str, shift_name: str, teknisi: str):
    """Update baris jika tanggal+shift sudah ada, atau tambah baru."""
    ws = _get_shift_worksheet()
    rows = ws.get_all_values()
    target_tgl = tanggal.strftime("%Y-%m-%d")

    for idx, row in enumerate(rows[1:], start=2):  # row index 1-based di sheets, skip header
        if len(row) < 5:
            continue
        if row[0].strip() == target_tgl and row[3].strip().upper() == shift_name.upper():
            # Update baris existing
            ws.update(f"A{idx}:E{idx}", [[target_tgl, jam_mulai, jam_selesai, shift_name, teknisi]])
            return

    # Tidak ditemukan → append
    ws.append_row([target_tgl, jam_mulai, jam_selesai, shift_name, teknisi], value_input_option="USER_ENTERED")


def _sheets_add_shift(tanggal: dt.date, jam_mulai: str, jam_selesai: str, shift_name: str, teknisi: str):
    """Selalu tambah baris baru."""
    ws = _get_shift_worksheet()
    ws.append_row([tanggal.strftime("%Y-%m-%d"), jam_mulai, jam_selesai, shift_name, teknisi], value_input_option="USER_ENTERED")


def _sheets_delete_shifts(tanggal: dt.date = None, shift_name: str = None) -> int:
    """Hapus baris berdasarkan tanggal dan/atau shift. Return jumlah baris terhapus."""
    ws = _get_shift_worksheet()
    rows = ws.get_all_values()
    target_tgl = tanggal.strftime("%Y-%m-%d") if tanggal else None

    rows_to_delete = []
    for idx, row in enumerate(rows[1:], start=2):
        if len(row) < 5 or not row[0]:
            continue
        match = True
        if target_tgl and row[0].strip() != target_tgl:
            match = False
        if shift_name and row[3].strip().upper() != shift_name.upper():
            match = False
        if match:
            rows_to_delete.append(idx)

    # Hapus dari bawah ke atas
    for idx in reversed(rows_to_delete):
        ws.delete_rows(idx)

    return len(rows_to_delete)


def _sheets_update_time(shift_name: str, jam_mulai: str, jam_selesai: str, tanggal: dt.date = None) -> int:
    """Update jam mulai/selesai untuk shift (dan tanggal) tertentu. Return jumlah baris terupdate."""
    ws = _get_shift_worksheet()
    rows = ws.get_all_values()
    target_tgl = tanggal.strftime("%Y-%m-%d") if tanggal else None

    updated = 0
    for idx, row in enumerate(rows[1:], start=2):
        if len(row) < 5 or not row[0]:
            continue
        if row[3].strip().upper() != shift_name.upper():
            continue
        if target_tgl and row[0].strip() != target_tgl:
            continue
        # Update kolom B dan C
        ws.update(f"B{idx}:C{idx}", [[jam_mulai, jam_selesai]])
        updated += 1

    return updated


def get_current_shift_from_sheets() -> dict:
    """Membaca jadwal shift dari Google Sheets dan return shift + teknisi saat ini."""
    now_dt = dt.datetime.now(TZ)
    today = now_dt.date()
    current_time = now_dt.time()

    try:
        rows = _sheets_read_shifts(today, today)
    except Exception as e:
        logger.error(f"Gagal baca shift dari Google Sheets: {e}")
        return {"shift": "OFF_SHIFT", "technician": None}

    for r in rows:
        try:
            t_start = dt.datetime.strptime(r["jam_mulai"], "%H:%M").time()
            t_end = dt.datetime.strptime(r["jam_selesai"], "%H:%M").time()
        except ValueError:
            continue

        if t_start <= current_time < t_end:
            return {"shift": r["shift"], "technician": r["teknisi"]}

    return {"shift": "OFF_SHIFT", "technician": None}


def load_technicians() -> list:
    """Membaca daftar teknisi dari file JSON"""
    if not os.path.exists(TECH_FILE):
        logger.warning(f"File {TECH_FILE} tidak ditemukan. Otomasi pembagian tiket tidak dapat berjalan.")
        return []
    try:
        with open(TECH_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"Gagal membaca file {TECH_FILE}: {e}")
        return []


def get_current_shift_from_excel(excel_path: str = "jadwal_shift.xlsx") -> dict:
    """Membaca jadwal shift dan nama teknisi dari file Excel secara real-time"""
    now_dt = dt.datetime.now(TZ)
    today_str = now_dt.strftime("%Y-%m-%d")
    current_time = now_dt.time()

    try:
        if not os.path.exists(excel_path):
            logger.warning(f"File Excel {excel_path} tidak ditemukan.")
            return {"shift": "OFF_SHIFT", "technician": None}

        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5:
                continue

            tgl_val, jam_mulai_val, jam_selesai_val, nama_shift, teknisi = (
                row[0],
                row[1],
                row[2],
                row[3],
                row[4],
            )

            if isinstance(tgl_val, (dt.date, dt.datetime)):
                tgl_str = tgl_val.strftime("%Y-%m-%d")
            else:
                tgl_str = str(tgl_val).strip()

            if tgl_str == today_str:
                t_start = (
                    jam_mulai_val
                    if isinstance(jam_mulai_val, dt.time)
                    else dt.datetime.strptime(str(jam_mulai_val).strip(), "%H:%M").time()
                )
                t_end = (
                    jam_selesai_val
                    if isinstance(jam_selesai_val, dt.time)
                    else dt.datetime.strptime(str(jam_selesai_val).strip(), "%H:%M").time()
                )

                if t_start <= current_time < t_end:
                    return {
                        "shift": str(nama_shift).strip(),
                        "technician": str(teknisi).strip(),
                    }

    except Exception as e:
        logger.error(f"Gagal membaca file Excel shift: {e}")

    return {"shift": "OFF_SHIFT", "technician": None}


def load_assign_state() -> dict:
    """Membaca indeks giliran teknisi terakhir dari file JSON"""
    try:
        with open(SDP_ASSIGN_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_assigned": {}}


def save_assign_state(state: dict):
    """Menyimpan indeks giliran teknisi terbaru"""
    try:
        with open(SDP_ASSIGN_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, indent=2)
    except Exception as e:
        logger.error(f"Gagal menyimpan state assign: {e}")


def select_technician_round_robin(shift: str, on_duty_techs_raw: str = None) -> dict:
    """Mengambil teknisi berikutnya secara bergantian dan adil di shift yang sama"""
    all_techs = load_technicians()
    
    if on_duty_techs_raw:
        tech_names = [t.strip().lower() for t in on_duty_techs_raw.split(",")]
        on_duty_techs = [
            t for t in all_techs 
            if t.get("name", "").strip().lower() in tech_names or t.get("email", "").strip().lower() in tech_names
        ]
    else:
        on_duty_techs = [t for t in all_techs if t.get("shift") == shift]

    if not on_duty_techs:
        # Fallback jika nama dari excel tidak cocok dengan JSON, pakai object temporary
        if on_duty_techs_raw:
            first_tech = on_duty_techs_raw.split(",")[0].strip()
            return {"name": first_tech, "email": first_tech}
        raise ValueError(f"Tidak ada teknisi yang terdaftar untuk shift {shift}")

    state = load_assign_state()
    last_index = state.get("last_assigned", {}).get(shift, -1)
    next_index = (last_index + 1) % len(on_duty_techs)

    if "last_assigned" not in state:
        state["last_assigned"] = {}
    state["last_assigned"][shift] = next_index
    save_assign_state(state)

    selected_tech = on_duty_techs[next_index]
    logger.info(f"🔄 [ROUND-ROBIN] Ticket dialokasikan ke: {selected_tech['name']} (Shift {shift})")
    return selected_tech


async def auto_assign_sdp_tickets(context: ContextTypes.DEFAULT_TYPE):
    """
    Job Background:
    1. GET ticket dengan Status = Open HANYA untuk grup divisi di SDP_NOTIFY_GROUPS
    2. Cek shift & tentukan teknisi via Round-Robin
    3. Update Technician + Status = In Progress Investigation
    4. VERIFICATION & Simpan ke File Log
    """
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    # Prioritas: Google Sheets, fallback ke file Excel lokal
    shift_info = get_current_shift_from_sheets()
    if shift_info.get("shift") == "OFF_SHIFT":
        shift_info = get_current_shift_from_excel("jadwal_shift.xlsx")
    current_shift = shift_info.get("shift")
    on_duty_techs = shift_info.get("technician")

    if not current_shift or current_shift == "OFF_SHIFT" or not on_duty_techs:
        logger.info("Saat ini di luar jam shift / tidak ada teknisi piket.")
        return

    try:
        open_tickets = await asyncio.to_thread(
            sdp.list_requests, 15, "Open", config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal mengambil tiket Open untuk otomasi assign: {e}")
        return

    if not open_tickets:
        return

    for req in open_tickets:
        ticket_id = req.get("id")
        if not ticket_id:
            continue

        ticket_group = (req.get("group") or {}).get("name", "")

        try:
            assigned_tech = select_technician_round_robin(current_shift, on_duty_techs)

            # 1. Ambil data subcategory, category & department asli dari tiket
            subcat_raw = req.get("subcategory")
            subcat_name = subcat_raw.get("name", "") if isinstance(subcat_raw, dict) else (subcat_raw if isinstance(subcat_raw, str) else "")

            cat_raw = req.get("category")
            cat_name = cat_raw.get("name", "") if isinstance(cat_raw, dict) else (cat_raw if isinstance(cat_raw, str) else "")

            dept_raw = req.get("department")
            dept_name = dept_raw.get("name", "") if isinstance(dept_raw, dict) else (dept_raw if isinstance(dept_raw, str) else "")

            # 2. TAHAP 1 & 2: Tentukan subcategory target
            target_subcat = None
            if subcat_name.strip():
                # Jika subcategory tiket SUDAH TERISI, pakai nilai bawaannya
                target_subcat = subcat_name.strip()
            else:
                # TAHAP 1: Cek apakah nama department tiket cocok dengan key di config
                dept_map = getattr(config, "DEPARTMENT_SUBCATEGORY_MAP", {})
                for dept_key, mapped_subcat in dept_map.items():
                    # Pengecekan case-insensitive
                    if dept_key.lower() in dept_name.lower():
                        # TAHAP 2: Tetapkan subcategory dari mapping
                        target_subcat = mapped_subcat
                        break

            # FALLBACK: subcategory wajib diisi di SDP, jika tidak ada match gunakan default
            if not target_subcat:
                target_subcat = getattr(config, "SDP_DEFAULT_SUBCATEGORY", "General")

            # 3. Susun payload update dasar
            update_payload = {
                "request": {
                    "status": {"name": "In Progress Investigation"},
                    "technician": {
                        "email_id": assigned_tech["email"]
                    },
                    "subcategory": {"name": target_subcat}
                }
            }

            # Jika tiket memiliki category asli, sertakan agar SDP tidak mereset subcategory ke General
            if cat_name.strip():
                update_payload["request"]["category"] = {"name": cat_name.strip()}

            # 4. Fungsi eksekusi PUT ke API SDP
            def _do_update():
                url = f"{sdp.base_url}/api/v3/requests/{ticket_id}"
                data = {"input_data": json.dumps(update_payload)}
                res = sdp.session.put(url, data=data)
                return res.status_code == 200, res.json()

            success, res_json = await asyncio.to_thread(_do_update)

            if success:
                verified_req = await asyncio.to_thread(sdp.get_request, ticket_id)
                v_status = (verified_req.get("status") or {}).get("name", "")
                v_tech = (verified_req.get("technician") or {}).get("name", "")

                write_ticket_log(
                    ticket_id=str(ticket_id),
                    tech_name=v_tech,
                    tech_email=assigned_tech['email'],
                    status=v_status,
                    shift=current_shift,
                    is_success=True
                )
                logger.info(f"✅ Tiket #{ticket_id} berhasil di-assign ke {v_tech} ({current_shift})")
            else:
                write_ticket_log(
                    ticket_id=str(ticket_id),
                    tech_name=assigned_tech.get('name', 'N/A'),
                    tech_email=assigned_tech.get('email', 'N/A'),
                    status="In Progress Investigation",
                    shift=current_shift,
                    is_success=False,
                    error_msg=str(res_json)
                )
                logger.error(f"Gagal update tiket #{ticket_id}: {res_json}")

        except Exception as e:
            write_ticket_log(
                ticket_id=str(ticket_id),
                tech_name="N/A",
                tech_email="N/A",
                status="N/A",
                shift=current_shift,
                is_success=False,
                error_msg=str(e)
            )
            logger.exception(f"Error pada otomasi tiket #{ticket_id}: {e}")


# ==============================================================================
# LOGIKA SCRIPT UTAMA & TELEGRAM BOT HANDLERS
# ==============================================================================

def get_jira_client_for_user(user_id) -> JiraClient:
    uid = str(user_id)
    if uid in _JIRA_CLIENT_CACHE:
        return _JIRA_CLIENT_CACHE[uid]
    account = user_accounts.get_account(uid)
    if not account:
        return None
    client = JiraClient(email=account["email"], api_token=account["api_token"])
    _JIRA_CLIENT_CACHE[uid] = client
    return client


def invalidate_jira_cache(user_id):
    _JIRA_CLIENT_CACHE.pop(str(user_id), None)


async def ensure_jira_account(update: Update, context: ContextTypes.DEFAULT_TYPE):
    client = get_jira_client_for_user(update.effective_user.id)
    if client is None:
        await update.effective_message.reply_text(
            "Kamu belum daftarkan akun Jira pribadi di bot ini.\n"
            "Ketik /myjira untuk daftar dulu (email Jira + API token kamu sendiri), "
            "supaya logwork tercatat atas nama kamu sendiri di Jira."
        )
    return client


TIME_PATTERN = re.compile(r"^(\d+w\s*)?(\d+d\s*)?(\d+h\s*)?(\d+m\s*)?$", re.IGNORECASE)
ISSUE_KEY_RE = re.compile(r"\b([A-Za-z][A-Za-z0-9]+-\d+)\b")
JAM_RE = re.compile(r"\bjam\s+(\d{1,2})(?:[:.](\d{2}))?\s*(pagi|siang|sore|malam)\b", re.IGNORECASE)
DURASI_RE = re.compile(r"(?:selama|durasi)\s+([^,\.]+)", re.IGNORECASE)
DESKRIPSI_RE = re.compile(r"deskripsi(?:nya)?(?:\s+pekerjaan(?:nya|ya))?\s*[:\-]?\s*(.+)", re.IGNORECASE)


def _parse_jam_indo(hour: int, minute: int, period: str) -> dt.time:
    period = (period or "").lower()
    if period == "pagi":
        if hour == 12:
            hour = 0
    elif period == "siang":
        if hour < 12:
            hour += 12
    elif period in ("sore", "malam"):
        if hour == 12 and period == "malam":
            hour = 0
        elif hour < 12:
            hour += 12
    return dt.time(hour=hour % 24, minute=minute)


def parse_natural_log(text: str) -> dict:
    result = {}

    m = ISSUE_KEY_RE.search(text)
    if m:
        result["issue_key"] = m.group(1).upper()

    m = JAM_RE.search(text)
    if m:
        hour = int(m.group(1))
        minute = int(m.group(2)) if m.group(2) else 0
        result["start_time"] = _parse_jam_indo(hour, minute, m.group(3))

    m = DURASI_RE.search(text)
    if m:
        raw = m.group(1)
        raw = re.sub(r"\bjam\b", "h", raw, flags=re.IGNORECASE)
        raw = re.sub(r"\b(menit|mnt)\b", "m", raw, flags=re.IGNORECASE)
        tokens = re.findall(r"\d+\s*[hm]", raw, flags=re.IGNORECASE)
        if tokens:
            result["duration"] = " ".join(t.replace(" ", "") for t in tokens)

    m = DESKRIPSI_RE.search(text)
    if m:
        result["description"] = m.group(1).strip().rstrip(".")

    if re.search(r"\bkemarin\b", text, re.IGNORECASE):
        result["date_offset"] = -1

    return result


def _extract_project_prefix(text: str):
    text = text.strip()
    if re.fullmatch(r"[A-Za-z]{2,10}-", text):
        return text[:-1].upper()
    if re.fullmatch(r"[A-Z]{2,10}", text):
        return text
    return None


(
    LOG_ISSUE,
    LOG_TIME,
    LOG_DESC,
    LOG_DATE,
    LOG_CONFIRM,
    LOG_BULK_CONFIRM,
    PICK_ISSUE_FOR_EDIT,
    PICK_WORKLOG_EDIT,
    EDIT_TIME,
    EDIT_DESC,
    PICK_ISSUE_FOR_DELETE,
    PICK_WORKLOG_DELETE,
    CONFIRM_DELETE,
    ADD_GUIDE_TITLE,
    ADD_GUIDE_KEYWORDS,
    ADD_GUIDE_CONTENT,
    ADD_GUIDE_CONFIRM,
    DEL_GUIDE_PICK,
    DEL_GUIDE_CONFIRM,
    EDIT_GUIDE_PICK,
    EDIT_GUIDE_MENU,
    EDIT_GUIDE_TITLE,
    EDIT_GUIDE_KEYWORDS,
    EDIT_GUIDE_CONTENT,
    ADD_GUIDE_ACTION_ASK,
    ADD_GUIDE_ACTION_SCRIPT,
    ADD_GUIDE_ACTION_FLAG,
    ADD_GUIDE_ACTION_TYPE,
    EDIT_GUIDE_ACTION_SCRIPT,
    EDIT_GUIDE_ACTION_FLAG,
    EDIT_GUIDE_ACTION_TYPE,
    RUN_ACTION_PARAM,
    RUN_ACTION_CONFIRM,
    ADD_GUIDE_ACTION_MODE,
    EDIT_GUIDE_ACTION_MODE,
    MYJIRA_EMAIL,
    MYJIRA_TOKEN,
    MYJIRA_CONFIRM,
    DELRES_INPUT,
    DELRES_CONFIRM,
    RELVOUCHER_INPUT,
    RELVOUCHER_CONFIRM,
    CHECKSTOCK_INPUT,
    CHECKSTOCK_CONFIRM,
    CEKPROMO_INPUT,
    CEKAWB_INPUT,
    CEKAWB_QUERY_CONFIRM,
    AWBJNE_INPUT,
    UPDATESHIFT_UPLOAD,
) = range(49)


def _thread_id_from_update(update: Update):
    msg = update.effective_message
    if msg is not None and getattr(msg, "is_topic_message", False):
        return msg.message_thread_id
    return None


def _is_allowed_chat(chat_id) -> bool:
    return str(chat_id) in config.allowed_chat_ids()


def restricted(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        if not _is_allowed_chat(update.effective_chat.id):
            await update.effective_message.reply_text(
                "Maaf, saya tidak kenal dengan anda"
            )
            return ConversationHandler.END
        return await func(update, context, *args, **kwargs)

    return wrapper


def seconds_to_human(total_seconds: int) -> str:
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    parts = []
    if hours:
        parts.append(f"{hours}h")
    if minutes:
        parts.append(f"{minutes}m")
    return " ".join(parts) if parts else "0m"


@restricted
async def tools_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /tools — tampilkan daftar tools sebagai tombol."""
    buttons = [
        [
            InlineKeyboardButton("📦 Cek Stock", callback_data="tools_stock"),
            InlineKeyboardButton("🏷 Cek Promo", callback_data="tools_promo"),
        ],
        [
            InlineKeyboardButton("🚚 Cek AWB (OMS)", callback_data="tools_awb"),
            InlineKeyboardButton("🚚 AWB JNE", callback_data="tools_awbjne"),
        ],
        [
            InlineKeyboardButton("🗑 Delete Reservation", callback_data="tools_delreservation"),
            InlineKeyboardButton("🎟 Release Voucher", callback_data="tools_releasevoucher"),
        ],
        [
            InlineKeyboardButton("📊 Query Database", callback_data="tools_query"),
            InlineKeyboardButton("🎫 Cek Tiket SDP", callback_data="tools_sdticket"),
        ],
    ]
    await update.message.reply_text(
        "<b>🧰 Tools Tersedia</b>\n\n"
        "Pilih tool yang ingin digunakan:",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def tools_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback untuk tombol /tools — kirim instruksi pemakaian."""
    query = update.callback_query
    await query.answer()

    tool = query.data.replace("tools_", "")

    instructions = {
        "stock": (
            "📦 <b>Cek Stock (OAA)</b>\n\n"
            "Ketik:\n<code>/stock 8000044321 SS20</code>\n\n"
            "Atau format lengkap:\n"
            "<code>sku: 8000044321\nsource: SS20</code>"
        ),
        "promo": (
            "🏷 <b>Cek Promo SRP Price</b>\n\n"
            "Ketik:\n<code>/promo 8100102377 AZ02</code>\n\n"
            "Atau format lengkap:\n"
            "<code>sku: 8100102377\nbucode: AZ02</code>"
        ),
        "awb": (
            "🚚 <b>Cek AWB / Tracking Order (OMS)</b>\n\n"
            "Ketik:\n<code>/awb 3301352973 IBOX</code>\n\n"
            "Atau format lengkap:\n"
            "<code>ordernumber: 3301352973\nsource: IBOX</code>"
        ),
        "awbjne": (
            "🚚 <b>Cek AWB JNE</b>\n\n"
            "Ketik:\n<code>/awbjne 8402663858 0157352600237230</code>\n\n"
            "Atau format lengkap:\n"
            "<code>ordernumber: 8402663858\nawb: 0157352600237230</code>"
        ),
        "delreservation": (
            "🗑 <b>Delete Reservation</b>\n\n"
            "Ketik:\n<code>/delreservation</code>\n\n"
            "Lalu upload file CSV atau ketik manual:\n"
            "<code>bucode: E370\nordernumber: 8302562258\nsku: 8100258103\nqty: 1</code>"
        ),
        "releasevoucher": (
            "🎟 <b>Release Voucher</b>\n\n"
            "Ketik:\n<code>/releasevoucher</code>\n\n"
            "Lalu upload file CSV atau ketik ID:\n"
            "<code>462016\n462017\n462018</code>"
        ),
        "query": (
            "📊 <b>Query Database</b>\n\n"
            "Ketik:\n<code>/query</code> → pilih dari daftar\n"
            "<code>/query nama_file</code> → jalankan langsung"
        ),
        "sdticket": (
            "🎫 <b>Cek Tiket ServiceDesk</b>\n\n"
            "Ketik:\n<code>/sdticket 12345</code>\n\n"
            "Atau ketik bebas:\n<code>cek tiket #12345</code>"
        ),
    }

    text = instructions.get(tool, "Tool tidak dikenali.")
    await query.edit_message_text(text, parse_mode=ParseMode.HTML)


@restricted
async def log_activity_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /logactivity — kirim isi ticket_activity.log ke chat."""
    if not os.path.exists(TICKET_LOG_FILE):
        await update.message.reply_text("📋 File log kosong — belum ada aktivitas tiket.")
        return

    file_size = os.path.getsize(TICKET_LOG_FILE)

    if file_size == 0:
        await update.message.reply_text("📋 File log kosong — belum ada aktivitas tiket.")
        return

    # Kalau file kecil (<3500 chars), kirim sebagai text
    # Kalau besar, kirim sebagai file
    if file_size < 3500:
        with open(TICKET_LOG_FILE, "r", encoding="utf-8") as f:
            content = f.read()
        await update.message.reply_text(
            f"📋 <b>Ticket Activity Log</b>\n\n<pre>{html.escape(content)}</pre>",
            parse_mode=ParseMode.HTML,
        )
    else:
        # Kirim 50 baris terakhir sebagai text
        with open(TICKET_LOG_FILE, "r", encoding="utf-8") as f:
            lines = f.readlines()

        last_50 = lines[-50:]
        text_preview = "".join(last_50)
        if len(text_preview) > 3500:
            text_preview = text_preview[-3500:]

        await update.message.reply_text(
            f"📋 <b>Ticket Activity Log</b> (50 baris terakhir)\n\n<pre>{html.escape(text_preview)}</pre>",
            parse_mode=ParseMode.HTML,
        )

        # Kirim juga sebagai file lengkap
        with open(TICKET_LOG_FILE, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="ticket_activity.log",
                caption=f"📋 Log lengkap ({len(lines)} baris)",
            )


@restricted
async def chatid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    msg = update.message
    text = f"Chat ID di sini: `{chat.id}`\nTipe: {chat.type}"
    if msg.is_topic_message and msg.message_thread_id:
        text += f"\nTopic ID: `{msg.message_thread_id}`"
    await msg.reply_text(text, parse_mode=ParseMode.MARKDOWN)


@restricted
async def myjira_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""
    user_id = update.effective_user.id

    if arg == "remove":
        ok = user_accounts.remove_account(user_id)
        invalidate_jira_cache(user_id)
        await update.message.reply_text(
            "Akun Jira kamu sudah dihapus dari bot." if ok else "Kamu belum daftarkan akun Jira."
        )
        return ConversationHandler.END

    if arg == "status":
        account = user_accounts.get_account(user_id)
        if account:
            await update.message.reply_text(
                f"Akun Jira kamu terdaftar: {account['email']}\n\n"
                "Ganti akun: /myjira\nHapus: /myjira remove"
            )
        else:
            await update.message.reply_text(
                "Kamu belum daftarkan akun Jira. Ketik /myjira untuk daftar."
            )
        return ConversationHandler.END

    return await myjira_start(update, context)


async def myjira_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Daftarkan akun Jira pribadi kamu, supaya logwork/edit/delete tercatat atas nama kamu sendiri.\n\n"
        "Masukkan **email akun Jira** kamu:",
        parse_mode=ParseMode.MARKDOWN,
    )
    return MYJIRA_EMAIL


async def myjira_email(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email = update.message.text.strip()
    if "@" not in email:
        await update.message.reply_text("Sepertinya itu bukan email yang valid, coba lagi:")
        return MYJIRA_EMAIL
    context.user_data["myjira_email"] = email
    await update.message.reply_text(
        "Sekarang masukkan **API token** Jira kamu.\n\n"
        "Belum punya? Buat dulu di: https://id.atlassian.com/manage-profile/security/api-tokens\n"
        "(klik \"Create API token\", copy tokennya, lalu paste di sini)",
        parse_mode=ParseMode.MARKDOWN,
    )
    return MYJIRA_TOKEN


async def myjira_token(update: Update, context: ContextTypes.DEFAULT_TYPE):
    token = update.message.text.strip()
    if not token:
        await update.message.reply_text("Tokennya kosong, coba lagi:")
        return MYJIRA_TOKEN
    context.user_data["myjira_token"] = token
    email = context.user_data["myjira_email"]

    await update.message.reply_text("Mengecek akun ke Jira, mohon tunggu...")
    try:
        test_client = JiraClient(email=email, api_token=token)
        resp = await asyncio.to_thread(
            test_client._request, "GET", test_client._url("/rest/api/3/myself")
        )
        me = test_client._safe_json(resp)
        display_name = me.get("displayName", "")
    except Exception as e:
        await update.message.reply_text(
            f"Gagal verifikasi ke Jira:\n{e}\n\nCek lagi email/token-nya, atau /cancel untuk batal."
        )
        return MYJIRA_TOKEN

    context.user_data["myjira_display_name"] = display_name
    keyboard = [
        [
            InlineKeyboardButton("✅ Ya, simpan", callback_data="myjira_save_yes"),
            InlineKeyboardButton("Batal", callback_data="myjira_save_no"),
        ]
    ]
    await update.message.reply_text(
        f"Berhasil verifikasi! Akun Jira: {display_name} ({email})\n\nSimpan sebagai akun kamu?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return MYJIRA_CONFIRM


async def myjira_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "myjira_save_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    user_id = update.effective_user.id
    user_accounts.set_account(
        user_id, d["myjira_email"], d["myjira_token"], d.get("myjira_display_name", "")
    )
    invalidate_jira_cache(user_id)
    await query.edit_message_text(
        f"Akun Jira kamu ({d['myjira_email']}) berhasil disimpan. ✅\n"
        "Sekarang /log, /edit, /delete, /today, /week bakal pakai akun ini."
    )
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Halo! Saya itops_bot bantu logwork ke Jira dan Otomasi ServiceDesk.\n\n"
        "Perintah yang tersedia:\n"
        "/myjira - daftarkan akun Jira pribadi kamu (wajib sebelum /log dkk)\n"
        "/log - isi logwork baru\n"
        "/today - rekap logwork hari ini\n"
        "/week - rekap logwork minggu ini\n"
        "/tasks - lihat daftar task di suatu project (contoh: /tasks TDBU)\n"
        "/export - download report Excel task suatu project (contoh: /export TIC)\n"
        "/tools - tampilkan semua tools yang tersedia\n"
        "/sdtickets [filter] - lihat tiket ServiceDesk Plus\n"
        "/sdticket <id> - lihat detail 1 tiket ServiceDesk Plus\n"
        "/sdreminder <menit>|off|status - atur reminder berkala utk tiket Open\n"
        "/delreservation - delete/cancel reservation (file CSV atau ketik manual)\n"
        "/releasevoucher - release voucher/promo (file CSV atau ketik manual)\n"
        "/stock - cek stock OAA (ketik SKU & Source)\n"
        "/promo - cek promo SRP price (ketik SKU & BU code)\n"
        "/awb - cek AWB/tracking order (ketik order number & source)\n"
        "/awbjne - cek AWB JNE (ketik order number & AWB)\n"
        "/query - jalankan query database manual\n"
        "/logactivity - lihat log aktivitas tiket\n"
        "/edit - edit logwork yang sudah ada\n"
        "/delete - hapus logwork\n"
        "/guide <kata kunci> - munculkan guidance/panduan tersimpan\n"
        "/run <kata kunci> - langsung eksekusi action/script guidance\n"
        "/addguide - tambah guidance baru\n"
        "/listguide - lihat semua guidance tersimpan\n"
        "/cancel - batalkan proses yang sedang berjalan"
    )


@restricted
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("Oke, dibatalkan.")
    return ConversationHandler.END


MONTH_MAP = {
    "januari": 1, "jan": 1, "februari": 2, "feb": 2, "maret": 3, "mar": 3,
    "april": 4, "apr": 4, "mei": 5, "juni": 6, "jun": 6, "juli": 7, "jul": 7,
    "agustus": 8, "agt": 8, "agu": 8, "september": 9, "sep": 9, "oktober": 10, "okt": 10,
    "november": 11, "nov": 11, "desember": 12, "des": 12
}


def parse_bulk_log_text(text: str) -> list:
    items = []
    lines = text.splitlines()
    current_year = dt.datetime.now(TZ).year

    for line in lines:
        line_str = line.strip()
        if not line_str:
            continue

        match_key = ISSUE_KEY_RE.search(line_str)
        if not match_key:
            continue
        issue_key = match_key.group(1).upper()

        dur_match = re.search(r'(?:selama|durasi)?\s*(\d+\s*(?:jam|h|menit|mnt|m))+', line_str, re.IGNORECASE)
        time_spent = "1h"
        if dur_match:
            raw_dur = dur_match.group(0)
            raw_dur = re.sub(r'\bjam\b', 'h', raw_dur, flags=re.IGNORECASE)
            raw_dur = re.sub(r'\b(menit|mnt)\b', 'm', raw_dur, flags=re.IGNORECASE)
            tokens = re.findall(r'\d+\s*[hm]', raw_dur, flags=re.IGNORECASE)
            if tokens:
                time_spent = " ".join(t.replace(" ", "") for t in tokens)

        date_obj = dt.datetime.now(TZ).date()
        date_match = re.search(r'(\d{1,2})\s+([a-zA-Z]+)(?:\s+(\d{4}))?', line_str, re.IGNORECASE)
        if date_match:
            day = int(date_match.group(1))
            month_str = date_match.group(2).lower()
            year = int(date_match.group(3)) if date_match.group(3) else current_year
            month = MONTH_MAP.get(month_str, date_obj.month)
            try:
                date_obj = dt.date(year, month, day)
            except ValueError:
                pass

        time_match = re.search(r'jam\s+(\d{1,2})[:.](\d{2})', line_str, re.IGNORECASE)
        start_time = None
        if time_match:
            start_time = dt.time(int(time_match.group(1)), int(time_match.group(2)))

        quotes = re.findall(r'"([^"]*)"', line_str)
        if quotes:
            comment = quotes[0]
        else:
            clean_text = re.sub(r'task\s+[A-Za-z0-9-]+', '', line_str, flags=re.IGNORECASE)
            clean_text = re.sub(r'pada tanggal.*', '', clean_text, flags=re.IGNORECASE)
            clean_text = re.sub(r'selama.*', '', clean_text, flags=re.IGNORECASE)
            comment = clean_text.strip(" .a-b-c-1-2-3-")

        items.append({
            "issue_key": issue_key,
            "comment": comment or "Worklog update",
            "time_spent": time_spent,
            "date": date_obj,
            "start_time": start_time
        })

    return items


async def show_bulk_log_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    items = context.user_data.get("bulk_items", [])
    if not items:
        await update.message.reply_text("Tidak ada data logwork yang berhasil dibaca. Coba lagi.")
        return ConversationHandler.END

    lines = [f"📋 <b>Daftar {len(items)} Logwork yang akan dikirim:</b>\n"]
    for idx, item in enumerate(items, 1):
        jam_str = f" jam {item['start_time'].strftime('%H:%M')}" if item.get('start_time') else ""
        lines.append(
            f"<b>{idx}. {item['issue_key']}</b> - {item['time_spent']}\n"
            f"    • Deskripsi: <i>{item['comment']}</i>\n"
            f"    • Tanggal: {item['date'].strftime('%d %b %Y')}{jam_str}\n"
        )

    lines.append("Kirim semua logwork ini ke Jira?")

    keyboard = [
        [
            InlineKeyboardButton("✅ Submit Semua", callback_data="bulk_confirm_yes"),
            InlineKeyboardButton("❌ Batal", callback_data="bulk_confirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text("\n".join(lines), reply_markup=markup, parse_mode=ParseMode.HTML)
    return LOG_BULK_CONFIRM


async def bulk_log_confirm_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "bulk_confirm_no":
        await query.edit_message_text("❌ Proses logwork dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    items = context.user_data.get("bulk_items", [])
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text(f"⏳ Sedang mengirim {len(items)} logwork ke Jira, mohon tunggu...")

    success_count = 0
    failed_items = []

    for item in items:
        try:
            await asyncio.to_thread(
                jira_client.add_worklog,
                item["issue_key"],
                item["time_spent"],
                item["comment"],
                item["date"],
                item.get("start_time"),
            )
            success_count += 1
        except JiraError as e:
            failed_items.append(f"{item['issue_key']}: {e}")

    result_text = f"✅ <b>Berhasil mengirim {success_count} dari {len(items)} logwork ke Jira!</b>"
    if failed_items:
        result_text += "\n\n⚠️ <b>Gagal pada task berikut:</b>\n" + "\n".join(failed_items)

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text=result_text,
        parse_mode=ParseMode.HTML,
        message_thread_id=_thread_id_from_update(update)
    )

    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def log_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    text = update.message.text or ""
    after = re.sub(r"^/log(@\w+)?\s*", "", text, flags=re.IGNORECASE).strip()

    if "\n" in after or len(re.findall(r'[A-Za-z]+-\d+', after)) > 1:
        parsed_items = parse_bulk_log_text(after)
        if parsed_items:
            context.user_data["bulk_items"] = parsed_items
            return await show_bulk_log_confirm(update, context)

    if after:
        context.user_data.update(parse_natural_log(after))
    return await advance_log_flow(update, context)


async def advance_log_flow(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    msg = update.effective_message

    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END

    if not d.get("issue_key"):
        await msg.reply_text("Isi ke issue Jira mana? (contoh: PROJ-123)")
        return LOG_ISSUE

    if "issue_title" not in d:
        try:
            d["issue_title"] = await asyncio.to_thread(jira_client.get_issue_summary_title, d["issue_key"])
        except JiraError:
            await msg.reply_text(
                f"Issue {d['issue_key']} tidak ditemukan atau tidak bisa diakses. "
                "Coba ketik issue key yang benar (contoh: PROJ-123), atau /cancel."
            )
            d.pop("issue_key", None)
            return LOG_ISSUE

    if not d.get("time_spent"):
        if d.get("duration"):
            d["time_spent"] = d["duration"]
        else:
            await msg.reply_text(
                f'Issue: {d["issue_key"]} - "{d["issue_title"]}"\n\n'
                "Berapa lama waktu yang dihabiskan? (contoh: 2h, 1h 30m, 45m)"
            )
            return LOG_TIME

    if not d.get("comment"):
        if d.get("description"):
            d["comment"] = d["description"]
        else:
            await msg.reply_text("Deskripsi pekerjaannya apa? (bisa singkat saja)")
            return LOG_DESC

    if not d.get("date"):
        if d.get("date_offset"):
            today = dt.datetime.now(TZ).date()
            d["date"] = today + dt.timedelta(days=d["date_offset"])
        else:
            keyboard = [
                [
                    InlineKeyboardButton("Hari ini", callback_data="date_today"),
                    InlineKeyboardButton("Kemarin", callback_data="date_yesterday"),
                ],
                [InlineKeyboardButton("Ketik tanggal manual", callback_data="date_manual")],
            ]
            await msg.reply_text(
                "Tanggal logwork-nya kapan?", reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return LOG_DATE

    return await log_show_confirm(update, context)


async def log_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["issue_key"] = update.message.text.strip().upper()
    context.user_data.pop("issue_title", None)
    return await advance_log_flow(update, context)


async def log_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    time_text = update.message.text.strip()
    if not TIME_PATTERN.match(time_text.replace(" ", "")) and not TIME_PATTERN.match(time_text):
        await update.message.reply_text(
            "Format waktu kurang tepat. Gunakan contoh seperti: 2h, 1h 30m, 45m, 1d. Coba lagi:"
        )
        return LOG_TIME
    context.user_data["time_spent"] = time_text
    return await advance_log_flow(update, context)


async def log_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["comment"] = update.message.text.strip()
    return await advance_log_flow(update, context)


async def log_date_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    today = dt.datetime.now(TZ).date()
    if query.data == "date_today":
        context.user_data["date"] = today
        return await log_show_confirm(update, context)
    elif query.data == "date_yesterday":
        context.user_data["date"] = today - dt.timedelta(days=1)
        return await log_show_confirm(update, context)
    else:
        await query.edit_message_text("Ketik tanggalnya, format: YYYY-MM-DD (contoh: 2026-08-01)")
        return LOG_DATE


async def log_date_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        date_obj = dt.datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text("Format tanggal salah. Contoh yang benar: 2026-08-01")
        return LOG_DATE
    context.user_data["date"] = date_obj
    return await log_show_confirm(update, context)


async def log_show_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    tanggal_line = d["date"].isoformat()
    if d.get("start_time"):
        tanggal_line += f" (jam {d['start_time'].strftime('%H:%M')})"
    text = (
        "Konfirmasi logwork berikut:\n\n"
        f"Issue: {d['issue_key']} - \"{d['issue_title']}\"\n"
        f"Waktu: {d['time_spent']}\n"
        f"Deskripsi: {d['comment']}\n"
        f"Tanggal: {tanggal_line}\n\n"
        "Kirim ke Jira sekarang?"
    )
    keyboard = [
        [
            InlineKeyboardButton("Ya, kirim", callback_data="confirm_yes"),
            InlineKeyboardButton("Batal", callback_data="confirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    return LOG_CONFIRM


async def log_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "confirm_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(
            jira_client.add_worklog,
            d["issue_key"],
            d["time_spent"],
            d["comment"],
            d["date"],
            d.get("start_time"),
        )
        await query.edit_message_text(
            f"Berhasil dicatat ke {d['issue_key']} ({d['time_spent']}). 👍"
        )
    except JiraError as e:
        await query.edit_message_text(f"Gagal mengirim ke Jira:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


def format_summary(entries: list, heading: str) -> str:
    if not entries:
        return f"{heading}\n\nBelum ada logwork tercatat."
    lines = [heading, ""]
    grand_total = 0
    for item in entries:
        total = seconds_to_human(item["total_seconds"])
        grand_total += item["total_seconds"]
        lines.append(f"• {item['issue_key']} - \"{item['summary']}\" — {total}")
        for e in item["entries"]:
            comment = f" ({e['comment']})" if e["comment"] else ""
            lines.append(f"    - {e['date']} · {e['time_spent']}{comment}")
    lines.append("")
    lines.append(f"Total: {seconds_to_human(grand_total)}")
    return "\n".join(lines)


@restricted
async def today_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return
    today = dt.datetime.now(TZ).date()
    try:
        entries = await asyncio.to_thread(jira_client.get_summary, today, today)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data dari Jira:\n{e}")
        return
    await update.message.reply_text(
        _safe_telegram_text(format_summary(entries, f"Rekap hari ini ({today.isoformat()}):"))
    )


@restricted
async def week_summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return
    today = dt.datetime.now(TZ).date()
    start = today - dt.timedelta(days=today.weekday())
    end = today
    try:
        entries = await asyncio.to_thread(jira_client.get_summary, start, end)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data dari Jira:\n{e}")
        return
    await update.message.reply_text(
        _safe_telegram_text(
            format_summary(entries, f"Rekap minggu ini ({start.isoformat()} s/d {end.isoformat()}):")
        )
    )


@restricted
async def list_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if text.lower().startswith("/tasks"):
        parts = text.split(maxsplit=1)
        if len(parts) < 2 or not parts[1].strip():
            await update.message.reply_text(
                "Ketik kode project-nya, contoh: /tasks TDBU\n"
                "Atau langsung ketik kode project-nya saja, contoh: TDBU- atau TIC"
            )
            return
        project_key = re.sub(r"-$", "", parts[1].strip()).upper()
    else:
        project_key = _extract_project_prefix(text)
        if not project_key:
            return

    try:
        issues = await asyncio.to_thread(jira.search_project_issues, project_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil daftar task project {project_key}:\n{e}")
        return

    if not issues:
        await update.message.reply_text(
            f"Tidak ada task ditemukan untuk project {project_key}.\n"
            "Cek lagi kode project-nya, atau pastikan kamu punya akses ke project tersebut."
        )
        return

    lines = [f"📋 Task terbaru di project {project_key} (maks 50, urut yang terakhir diupdate):", ""]
    for issue in issues:
        key = issue["key"]
        summary = issue["fields"].get("summary", "")
        status = (issue["fields"].get("status") or {}).get("name", "")
        lines.append(f"• {key} - {summary} [{status}]")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines)))


def _safe_telegram_text(text: str, limit: int = 3800) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n\n... (dipotong, terlalu banyak data -- coba filter lebih spesifik)"


def _strip_html(text: str) -> str:
    if not text:
        return ""
    
    decoded_text = html.unescape(text)
    decoded_text = re.sub(r"(?i)<(br|p|div|li)\b[^>]*>", "\n", decoded_text)
    clean_text = re.sub(r"<[^>]+>", "", decoded_text)
    lines = [re.sub(r"[ \t]+", " ", line).strip() for line in clean_text.splitlines()]
    return "\n".join(line for line in lines if line)


def _format_sdp_request_line(req: dict) -> str:
    req_id = req.get("id", "")
    subject = req.get("subject", "")
    status = (req.get("status") or {}).get("name", "")
    requester = (req.get("requester") or {}).get("name", "")
    
    description = _strip_html(req.get("description", ""))
    desc_line = f"\nDescription:\n{description}" if description else ""
    
    return (
        f"Ticket #{req_id}\n"
        f"Subject: {subject}\n"
        f"Requester: {requester}\n"
        f"Status: {status}"
        f"{desc_line}"
    )


SDP_FILTER_RE = re.compile(r"(status|group)\s*:\s*(.+?)(?=\s+(?:status|group)\s*:|$)", re.IGNORECASE)


def _parse_sdp_filters(text: str) -> dict:
    filters = {}
    for m in SDP_FILTER_RE.finditer(text):
        filters[m.group(1).lower()] = m.group(2).strip()
    return filters


@restricted
async def sdtickets_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not sdp:
        await update.message.reply_text(
            "ServiceDesk Plus belum dikonfigurasi. Isi dulu SDP_BASE_URL dan SDP_API_KEY di file .env, "
            "lalu restart bot."
        )
        return
    parts = update.message.text.split(maxsplit=1)
    raw_filter = parts[1].strip() if len(parts) > 1 else ""
    filters = _parse_sdp_filters(raw_filter)
    status_filter = filters.get("status")
    group_filter = filters.get("group")

    used_default_groups = False
    if group_filter is None and config.SDP_NOTIFY_GROUPS:
        group_filter = config.SDP_NOTIFY_GROUPS
        used_default_groups = True
    elif group_filter and group_filter.strip().lower() == "all":
        group_filter = None

    try:
        requests_list = await asyncio.to_thread(
            sdp.list_requests, 15, status_filter, group_filter
        )
    except SDPError as e:
        await update.message.reply_text(f"Gagal mengambil data dari ServiceDesk Plus:\n{e}")
        return

    if not requests_list:
        await update.message.reply_text("Tidak ada tiket ditemukan dengan filter tersebut.")
        return

    filter_bits = []
    if status_filter:
        filter_bits.append(f"status: {status_filter}")
    if used_default_groups:
        filter_bits.append(f"group kamu: {', '.join(config.SDP_NOTIFY_GROUPS)}")
    elif group_filter:
        filter_bits.append(f"group: {group_filter}")
    filter_note = f" ({', '.join(filter_bits)})" if filter_bits else ""

    lines = [f"🎫 15 tiket ServiceDesk Plus terbaru{filter_note}:", ""]
    for req in requests_list:
        lines.append(_format_sdp_request_line(req))
        lines.append("")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines).rstrip()))


# Regex untuk menangkap pesan natural "cek tiket #12356" / "cek ticket 12356" / "cek tiket#12356"
CEK_TIKET_RE = re.compile(r"\b(?:cek|check|lihat|status)\s+(?:tiket|ticket|tik)\s*#?\s*(\d+)", re.IGNORECASE)

# Regex untuk mendeteksi text bebas yang mengandung format input tools
NATURAL_TOOLS_RE = re.compile(
    r"(?:ordernumber|order|no)\s*[:=]\s*\S+.*(?:awb|resi|source|src)\s*[:=]\s*\S+"
    r"|(?:sku|artikle?)\s*[:=]\s*\S+.*(?:bucode|bu|source|store)\s*[:=]\s*\S+",
    re.IGNORECASE | re.DOTALL
)


async def natural_tools_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler text bebas yang mendeteksi format input tools dan langsung eksekusi."""
    if not _is_allowed_chat(update.effective_chat.id):
        return
    await _try_execute_tools(update, context, update.message.text.strip())


async def _try_execute_tools(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str) -> bool:
    """Deteksi format input tools dari `text` dan langsung eksekusi.

    Return True kalau ada tool yang terdeteksi & dijalankan, False kalau tidak
    cocok dengan format tool manapun.
    """
    text = (text or "").strip()

    # Coba detect Delete Reservation (bucode + ordernumber + sku + qty)
    # Harus dicek SEBELUM promo karena keduanya punya sku + bucode
    delres_data = _parse_delres_text(text)
    if delres_data:
        # Ada data delreservation, arahkan ke flow konfirmasi
        csv_path = _build_delres_csv(delres_data)
        context.user_data["delres_csv"] = csv_path
        context.user_data["delres_source"] = "text"
        context.user_data["delres_items"] = delres_data

        preview_lines = []
        for i, item in enumerate(delres_data, 1):
            skus = ", ".join(item["itemCodes"])
            preview_lines.append(
                f"{i}. {item['transactionNumber']} | BU: {item['businessUnitCode']} | "
                f"SKU: {skus} | Qty: {item.get('qty', 1)}"
            )

        await update.message.reply_text(
            f"<b>🗑 Konfirmasi Delete Reservation</b>\n\n"
            + "\n".join(preview_lines)
            + "\n\nEksekusi sekarang?",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Eksekusi", callback_data="delres_yes"),
                    InlineKeyboardButton("❌ Batal", callback_data="delres_no"),
                ]
            ]),
        )
        return True

    # Coba detect AWB JNE (ordernumber + awb)
    awbjne_data = _parse_awbjne_text(text)
    if awbjne_data:
        context.user_data["awbjne_items"] = awbjne_data
        await _awbjne_execute_multi(update, context)
        return True

    # Coba detect AWB OMS (ordernumber + source)
    cekawb_data = _parse_cekawb_text(text)
    if cekawb_data:
        context.user_data["cekawb_data"] = cekawb_data
        await _cekawb_execute(update, context)
        return True

    # Coba detect Promo (sku + bucode, TANPA ordernumber)
    promo_data = _parse_cekpromo_text(text)
    if promo_data and not re.search(r"(?:ordernumber|order|no)\s*[:=]", text, re.IGNORECASE):
        context.user_data["cekpromo_data"] = promo_data
        await _cekpromo_execute(update, context)
        return True

    # Coba detect Stock (sku/article + source)
    stock_data = _parse_checkstock_text(text)
    if stock_data:
        context.user_data["checkstock_data"] = stock_data
        await _checkstock_show_confirm(update, context)
        return True

    return False


async def _format_ticket_detail(request_id: str, req: dict) -> str:
    """Format detail tiket SDP. Jika status Resolved/Closed, tampilkan juga notes resolusi."""
    subject = req.get("subject", "")
    status = (req.get("status") or {}).get("name", "")
    requester = (req.get("requester") or {}).get("name", "")
    technician = (req.get("technician") or {}).get("name", "")
    priority = (req.get("priority") or {}).get("name", "")
    group = (req.get("group") or {}).get("name", "")
    category = (req.get("category") or {}).get("name", "")
    subcategory = (req.get("subcategory") or {}).get("name", "")
    department = (req.get("department") or {}).get("name", "")
    created = (req.get("created_time") or {}).get("display_value", "")
    due_date = (req.get("due_by_time") or {}).get("display_value", "")

    lines = [
        f"🎫 <b>Tiket #{request_id}</b>",
        "",
        f"<b>Subjek:</b> {html.escape(subject)}",
        f"<b>Status:</b> {status}",
        f"<b>Requester:</b> {requester}",
    ]
    if technician:
        lines.append(f"<b>Technician:</b> {technician}")
    if group:
        lines.append(f"<b>Group:</b> {group}")
    if priority:
        lines.append(f"<b>Priority:</b> {priority}")
    if department:
        lines.append(f"<b>Department:</b> {department}")
    if category:
        lines.append(f"<b>Category:</b> {category}")
    if subcategory:
        lines.append(f"<b>Subcategory:</b> {subcategory}")
    if created:
        lines.append(f"<b>Dibuat:</b> {created}")
    if due_date:
        lines.append(f"<b>Due Date:</b> {due_date}")

    # Tampilkan deskripsi tiket
    description = _strip_html(req.get("description", ""))
    if description:
        lines.append(f"\n<b>Deskripsi:</b>\n{html.escape(description)}")

    # Jika tiket sudah Resolved/Closed, ambil notes resolusi
    if status.lower() in ("resolved", "closed"):
        try:
            notes = await asyncio.to_thread(sdp.get_request_notes, request_id)
            if notes:
                lines.append("")
                lines.append("<b>📋 Resolution Notes:</b>")
                for note in notes:
                    note_by = (note.get("created_by") or {}).get("name", "")
                    note_desc = _strip_html(note.get("description", ""))
                    if not note_desc:
                        continue
                    if note_by:
                        lines.append(f"\n<b>by {html.escape(note_by)}</b>")
                    lines.append(html.escape(note_desc))
        except SDPError:
            pass  # Gagal ambil notes, tidak fatal -- tampilkan info dasar saja

    return "\n".join(lines)


async def cek_tiket_natural(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler pesan natural 'cek tiket #12356' untuk menampilkan detail tiket ServiceDesk Plus."""
    if not _is_allowed_chat(update.effective_chat.id):
        return
    if not sdp:
        await update.message.reply_text(
            "ServiceDesk Plus belum dikonfigurasi. Isi dulu SDP_BASE_URL dan SDP_API_KEY di file .env, "
            "lalu restart bot."
        )
        return

    match = CEK_TIKET_RE.search(update.message.text)
    if not match:
        return
    request_id = match.group(1)

    try:
        req = await asyncio.to_thread(sdp.get_request, request_id)
    except SDPError as e:
        await update.message.reply_text(f"Gagal mengambil tiket #{request_id}:\n{e}")
        return

    text = await _format_ticket_detail(request_id, req)
    await update.message.reply_text(_safe_telegram_text(text), parse_mode=ParseMode.HTML)


@restricted
async def sdticket_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk /sdticket <id>"""
    if not sdp:
        await update.message.reply_text(
            "ServiceDesk Plus belum dikonfigurasi. Isi dulu SDP_BASE_URL dan SDP_API_KEY di file .env, "
            "lalu restart bot."
        )
        return
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await update.message.reply_text("Ketik: /sdticket <ID tiket>\ncontoh: /sdticket 12345")
        return
    request_id = parts[1].strip().lstrip("#")

    try:
        req = await asyncio.to_thread(sdp.get_request, request_id)
    except SDPError as e:
        await update.message.reply_text(f"Gagal mengambil tiket {request_id}:\n{e}")
        return

    text = await _format_ticket_detail(request_id, req)
    await update.message.reply_text(_safe_telegram_text(text), parse_mode=ParseMode.HTML)


@restricted
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    if len(parts) < 2 or not parts[1].strip():
        await update.message.reply_text("Ketik: /export <KODE_PROJECT>\ncontoh: /export TIC")
        return
    project_key = re.sub(r"-$", "", parts[1].strip()).upper()

    await update.message.reply_text(
        f"Sedang mengambil data project {project_key} dari Jira (termasuk riwayat status "
        "Dev/QA Done), mohon tunggu sebentar..."
    )
    try:
        rows = await asyncio.to_thread(jira.get_project_export, project_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil data:\n{e}")
        return

    if not rows:
        await update.message.reply_text(
            f"Tidak ada task ditemukan di project {project_key}. "
            "Cek lagi kode project-nya, atau pastikan kamu punya akses."
        )
        return

    filepath = build_export_excel(project_key, rows)
    try:
        with open(filepath, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=os.path.basename(filepath),
                caption=f"📊 Report project {project_key} ({len(rows)} task)",
            )
    finally:
        try:
            os.remove(filepath)
        except OSError:
            pass


@restricted
async def edit_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END
    await update.message.reply_text("Issue Jira mana yang mau diedit worklog-nya? (contoh: PROJ-123)")
    return PICK_ISSUE_FOR_EDIT


async def edit_pick_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    issue_key = update.message.text.strip().upper()
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        return ConversationHandler.END
    try:
        worklogs = await asyncio.to_thread(jira_client.get_my_recent_worklogs, issue_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil worklog:\n{e}")
        return PICK_ISSUE_FOR_EDIT
    if not worklogs:
        await update.message.reply_text(
            f"Tidak ada worklog kamu di {issue_key} dalam 14 hari terakhir. Coba issue lain atau /cancel."
        )
        return PICK_ISSUE_FOR_EDIT
    context.user_data["issue_key"] = issue_key
    buttons = []
    for w in worklogs:
        label = f"{w['started'][:10]} · {w.get('timeSpent', '')}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"wl_{w['id']}")])
    await update.message.reply_text(
        "Pilih worklog yang mau diedit:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return PICK_WORKLOG_EDIT


async def edit_pick_worklog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    worklog_id = query.data.replace("wl_", "")
    context.user_data["worklog_id"] = worklog_id
    await query.edit_message_text(
        "Masukkan waktu baru (contoh: 2h 30m), atau ketik - untuk tidak mengubah waktu:"
    )
    return EDIT_TIME


async def edit_time(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["new_time"] = None if text == "-" else text
    await update.message.reply_text(
        "Masukkan deskripsi baru, atau ketik - untuk tidak mengubah deskripsi:"
    )
    return EDIT_DESC


async def edit_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    new_comment = None if text == "-" else text
    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(
            jira_client.update_worklog, d["issue_key"], d["worklog_id"], d.get("new_time"), new_comment
        )
        await update.message.reply_text("Worklog berhasil diperbarui. ✅")
    except JiraError as e:
        await update.message.reply_text(f"Gagal update worklog:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def delete_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    jira_client = await ensure_jira_account(update, context)
    if not jira_client:
        return ConversationHandler.END
    await update.message.reply_text("Issue Jira mana yang worklog-nya mau dihapus? (contoh: PROJ-123)")
    return PICK_ISSUE_FOR_DELETE


async def delete_pick_issue(update: Update, context: ContextTypes.DEFAULT_TYPE):
    issue_key = update.message.text.strip().upper()
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await update.message.reply_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        return ConversationHandler.END
    try:
        worklogs = await asyncio.to_thread(jira_client.get_my_recent_worklogs, issue_key)
    except JiraError as e:
        await update.message.reply_text(f"Gagal mengambil worklog:\n{e}")
        return PICK_ISSUE_FOR_DELETE
    if not worklogs:
        await update.message.reply_text(
            f"Tidak ada worklog kamu di {issue_key} dalam 14 hari terakhir. Coba issue lain atau /cancel."
        )
        return PICK_ISSUE_FOR_DELETE
    context.user_data["issue_key"] = issue_key
    buttons = []
    for w in worklogs:
        label = f"{w['started'][:10]} · {w.get('timeSpent', '')}"
        buttons.append([InlineKeyboardButton(label, callback_data=f"wl_{w['id']}")])
    await update.message.reply_text(
        "Pilih worklog yang mau dihapus:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return PICK_WORKLOG_DELETE


async def delete_pick_worklog(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    worklog_id = query.data.replace("wl_", "")
    context.user_data["worklog_id"] = worklog_id
    keyboard = [
        [
            InlineKeyboardButton("Ya, hapus", callback_data="del_yes"),
            InlineKeyboardButton("Batal", callback_data="del_no"),
        ]
    ]
    await query.edit_message_text(
        "Yakin mau hapus worklog ini?", reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return CONFIRM_DELETE


async def delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "del_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    d = context.user_data
    jira_client = get_jira_client_for_user(update.effective_user.id)
    if not jira_client:
        await query.edit_message_text("Akun Jira kamu tidak ditemukan lagi, ketik /myjira dulu.")
        context.user_data.clear()
        return ConversationHandler.END
    try:
        await asyncio.to_thread(jira_client.delete_worklog, d["issue_key"], d["worklog_id"])
        await query.edit_message_text("Worklog berhasil dihapus. 🗑️")
    except JiraError as e:
        await update.message.reply_text(f"Gagal menghapus worklog:\n{e}")
    context.user_data.clear()
    return ConversationHandler.END


async def send_guidance_entry(update: Update, context: ContextTypes.DEFAULT_TYPE, entry: dict):
    caption = f"📘 {entry['title']}\n\n{entry['content']}".strip()
    attachment = entry.get("attachment")
    action = entry.get("action")
    chat_id = update.effective_chat.id

    action_markup = None
    if action:
        action_markup = InlineKeyboardMarkup(
            [[InlineKeyboardButton("⚙️ Jalankan Script", callback_data=f"runaction_{entry['id']}")]]
        )

    if attachment:
        doc_target = attachment.get("file_id") or attachment.get("file_path")
        thread_id = _thread_id_from_update(update)

        if doc_target:
            try:
                if isinstance(doc_target, str) and os.path.exists(doc_target):
                    with open(doc_target, "rb") as f:
                        await context.bot.send_document(
                            chat_id=chat_id,
                            document=f,
                            filename=attachment.get("file_name"),
                            caption=caption[:1024] if caption else None,
                            message_thread_id=thread_id,
                        )
                else:
                    await context.bot.send_document(
                        chat_id=chat_id,
                        document=doc_target,
                        filename=attachment.get("file_name"),
                        caption=caption[:1024] if caption else None,
                        message_thread_id=thread_id,
                    )
            except BadRequest as e:
                logger.error(f"Gagal mengirim dokumen guidance: {e}")
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="⚠️ File/dokumen panduan tidak dapat dikirim karena file_id atau path tidak valid.",
                    message_thread_id=thread_id,
                )

            if len(caption) > 1024:
                await context.bot.send_message(chat_id=chat_id, text=caption, message_thread_id=thread_id)
            if action_markup:
                await context.bot.send_message(
                    chat_id=chat_id,
                    text="Guidance ini bisa langsung dieksekusi:",
                    reply_markup=action_markup,
                    message_thread_id=thread_id,
                )
            return

    if update.callback_query:
        await update.callback_query.edit_message_text(caption, reply_markup=action_markup)
    else:
        await update.message.reply_text(caption, reply_markup=action_markup)


async def send_guidance_matches(update: Update, context: ContextTypes.DEFAULT_TYPE, query: str, matches: list):
    if len(matches) == 1:
        await send_guidance_entry(update, context, matches[0])
        return
    buttons = [
        [InlineKeyboardButton(m["title"], callback_data=f"gid_{m['id']}")] for m in matches[:8]
    ]
    await update.message.reply_text(
        f"Ada beberapa guidance yang cocok dengan '{query}', pilih salah satu:",
        reply_markup=InlineKeyboardMarkup(buttons),
    )


@restricted
async def guide_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    query = parts[1].strip() if len(parts) > 1 else ""
    if not query:
        await update.message.reply_text(
            "Ketik: /guide <kata kunci>\ncontoh: /guide koneksi jaringan error"
        )
        return
    matches = guidance_store.find_matches(query)
    if not matches:
        await update.message.reply_text(
            f"Guidance untuk '{query}' tidak ditemukan.\n"
            "Ketik /listguide untuk lihat semua topik yang tersedia, "
            "atau /addguide untuk menambahkan yang baru."
        )
        return
    await send_guidance_matches(update, context, query, matches)


async def guide_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("gid_", "")
    entry = guidance_store.get_guidance(gid)
    if not entry:
        await query.edit_message_text("Guidance tidak ditemukan (mungkin sudah dihapus).")
        return
    await send_guidance_entry(update, context, entry)


def _action_param_list(action: dict) -> list:
    if action.get("params"):
        return action["params"]
    flag = action.get("param_flag")
    if flag:
        return [{"flag": flag, "type": action.get("param_type", "text")}]
    return []


async def _ask_next_action_param(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    flags = d["run_flags"]
    idx = d["run_flag_index"]

    if idx >= len(flags):
        return await _show_run_confirm(update, context)

    current = flags[idx]
    flag_name = current.get("flag", "")
    ptype = current.get("type", "text")
    if ptype == "file":
        text = f"Upload file untuk parameter {flag_name}:"
    else:
        text = f"Masukkan nilai untuk parameter {flag_name}\n(kalau lebih dari 1 nilai, pisahkan dengan koma)"

    if update.callback_query:
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text=text, message_thread_id=_thread_id_from_update(update)
        )
    else:
        await update.effective_message.reply_text(text)
    return RUN_ACTION_PARAM


async def _show_run_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    lines = ["Konfirmasi:", "", f"Script: {d['run_script_path']}"]
    for flag_name, info in d["run_values"].items():
        lines.append(f"{flag_name}: {info['display']}")
    lines.append("")
    lines.append("Jalankan sekarang?")

    keyboard = [
        [
            InlineKeyboardButton("✅ Ya, jalankan", callback_data="runconfirm_yes"),
            InlineKeyboardButton("Batal", callback_data="runconfirm_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="\n".join(lines),
            reply_markup=markup,
            message_thread_id=_thread_id_from_update(update),
        )
    else:
        await update.effective_message.reply_text("\n".join(lines), reply_markup=markup)
    return RUN_ACTION_CONFIRM


async def _start_run_action(update: Update, context: ContextTypes.DEFAULT_TYPE, entry: dict):
    action = entry["action"]
    flags = _action_param_list(action)
    context.user_data.clear()
    context.user_data["run_guide_id"] = entry["id"]
    context.user_data["run_script_path"] = action.get("script_path", "")
    context.user_data["run_flags"] = flags

    paired = action.get("input_mode") == "paired" and len(flags) > 1
    context.user_data["run_paired_mode"] = paired

    if paired:
        flag_names = ", ".join(f["flag"].lstrip("-") for f in flags)
        example = "\n".join(
            "8000044321, SS20" if i == 0 else "8000044322, SS21" for i in range(2)
        )
        text = (
            f"Masukkan data, satu baris per kombinasi ({flag_names}):\n\n"
            f"Contoh:\n{example}"
        )
        msg = update.effective_message
        if update.callback_query:
            await context.bot.send_message(
                chat_id=update.effective_chat.id, text=text, message_thread_id=_thread_id_from_update(update)
            )
        else:
            await msg.reply_text(text)
        return RUN_ACTION_PARAM

    context.user_data["run_flag_index"] = 0
    context.user_data["run_values"] = {}
    return await _ask_next_action_param(update, context)


async def run_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    parts = update.message.text.split(maxsplit=1)
    query = parts[1].strip() if len(parts) > 1 else ""
    if not query:
        await update.message.reply_text("Ketik: /run <kata kunci guidance>\ncontoh: /run delete reservation")
        return ConversationHandler.END

    matches = guidance_store.find_matches(query)
    if not matches:
        await update.message.reply_text(
            f"Guidance untuk '{query}' tidak ditemukan.\nKetik /listguide untuk lihat semua topik."
        )
        return ConversationHandler.END

    matches_with_action = [m for m in matches if m.get("action")]
    if not matches_with_action:
        await update.message.reply_text(
            f"Guidance '{matches[0]['title']}' ditemukan, tapi belum punya action/script.\n"
            "Tambahkan lewat /editguide dulu (tombol '⚙️ Tambah Action'), "
            "atau pakai /guide untuk lihat isinya saja."
        )
        return ConversationHandler.END

    if len(matches_with_action) > 1:
        buttons = [
            [InlineKeyboardButton(f"▶️ {m['title']}", callback_data=f"runaction_{m['id']}")]
            for m in matches_with_action[:8]
        ]
        await update.message.reply_text(
            f"Ada beberapa guidance dengan action yang cocok dengan '{query}', pilih salah satu:",
            reply_markup=InlineKeyboardMarkup(buttons),
        )
        return ConversationHandler.END

    entry = matches_with_action[0]
    await update.message.reply_text(f"Jalankan: {entry['title']}")
    return await _start_run_action(update, context, entry)


async def run_action_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if not _is_allowed_chat(update.effective_chat.id):
        return ConversationHandler.END
    gid = query.data.replace("runaction_", "")
    entry = guidance_store.get_guidance(gid)
    thread_id = _thread_id_from_update(update)
    if not entry or not entry.get("action"):
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text="Action tidak ditemukan.", message_thread_id=thread_id
        )
        return ConversationHandler.END

    await context.bot.send_message(
        chat_id=update.effective_chat.id, text=f"Jalankan: {entry['title']}", message_thread_id=thread_id
    )
    return await _start_run_action(update, context, entry)


async def run_action_param(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    flags = d.get("run_flags")
    if not flags:
        await update.message.reply_text("Sesi eksekusi tidak ditemukan lagi, dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    if d.get("run_paired_mode"):
        text = (update.message.text or "").strip()
        if not text:
            await update.message.reply_text("Datanya kosong, coba lagi (kirim ulang semua baris):")
            return RUN_ACTION_PARAM
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        values = {f["flag"]: [] for f in flags}
        for line in lines:
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != len(flags):
                flag_names = ", ".join(f["flag"].lstrip("-") for f in flags)
                await update.message.reply_text(
                    f"Baris '{line}' jumlah nilainya ({len(parts)}) tidak sesuai jumlah "
                    f"parameter ({len(flags)}: {flag_names}). Coba lagi, kirim ulang semua baris:"
                )
                return RUN_ACTION_PARAM
            for f, val in zip(flags, parts):
                values[f["flag"]].append(val)
        d["run_values"] = {
            flag_name: {"type": "text", "value": vals, "display": ", ".join(vals)}
            for flag_name, vals in values.items()
        }
        return await _show_run_confirm(update, context)

    idx = d["run_flag_index"]
    current = flags[idx]
    flag_name = current.get("flag", "")
    ptype = current.get("type", "text")

    if ptype == "file":
        if not update.message.document:
            await update.message.reply_text("Mohon upload file yang dibutuhkan, atau ketik /cancel.")
            return RUN_ACTION_PARAM
        doc = update.message.document
        tg_file = await doc.get_file()
        local_path = os.path.join(tempfile.gettempdir(), doc.file_name)
        await tg_file.download_to_drive(local_path)
        d["run_values"][flag_name] = {"type": "file", "value": local_path, "display": doc.file_name}
    else:
        text = (update.message.text or "").strip()
        if not text:
            await update.message.reply_text("Nilainya kosong, coba lagi:")
            return RUN_ACTION_PARAM
        values = [v.strip() for v in text.split(",") if v.strip()]
        d["run_values"][flag_name] = {"type": "text", "value": values, "display": text}

    d["run_flag_index"] += 1
    return await _ask_next_action_param(update, context)


def _filter_item_detail_lines(text: str) -> str:
    item_line_re = re.compile(r"^\s*\[\d+\]")
    kept_lines = [line for line in text.splitlines() if not item_line_re.match(line)]
    result_lines = []
    prev_blank = False
    for line in kept_lines:
        is_blank = not line.strip()
        if is_blank and prev_blank:
            continue
        result_lines.append(line)
        prev_blank = is_blank
    return "\n".join(result_lines).strip()


async def run_action_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "runconfirm_no":
        await query.edit_message_text("Dibatalkan, script tidak dijalankan.")
        context.user_data.clear()
        return ConversationHandler.END

    d = context.user_data
    script_path = d.get("run_script_path", "")
    run_values = d.get("run_values", {})
    if not script_path:
        await query.edit_message_text("Sesi eksekusi tidak ditemukan lagi, dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text(f"⏳ Menjalankan {os.path.basename(script_path)}, mohon tunggu...")

    cmd = ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script_path]
    for flag_name, info in run_values.items():
        cmd.append(flag_name)
        if info["type"] == "file":
            cmd.append(info["value"])
        else:
            cmd.extend(info["value"])

    chat_id = update.effective_chat.id
    thread_id = _thread_id_from_update(update)
    script_dir = os.path.dirname(script_path) or None
    try:
        result = await asyncio.to_thread(
            subprocess.run, cmd, capture_output=True, text=True, timeout=300, cwd=script_dir
        )
        output = (result.stdout or "").strip()
        err = (result.stderr or "").strip()
        output = _filter_item_detail_lines(output)
        combined = output + (f"\n\n[stderr]\n{err}" if err else "")
        combined = combined or "(tidak ada output dari script)"
        if len(combined) > 3500:
            combined = combined[:3500] + "\n... (output dipotong, terlalu panjang)"
        icon = "✅" if result.returncode == 0 else "⚠️"
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"{icon} Selesai (exit code {result.returncode}):\n\n{combined}",
            message_thread_id=thread_id,
        )
    except subprocess.TimeoutExpired:
        await context.bot.send_message(
            chat_id=chat_id,
            text="⚠️ Script berjalan lebih dari 5 menit, dihentikan paksa (timeout).",
            message_thread_id=thread_id,
        )
    except FileNotFoundError:
        await context.bot.send_message(
            chat_id=chat_id,
            text=f"⚠️ Gagal menjalankan: file script tidak ditemukan di path:\n{script_path}",
            message_thread_id=thread_id,
        )
    except Exception as e:
        await context.bot.send_message(
            chat_id=chat_id, text=f"⚠️ Gagal menjalankan script:\n{e}", message_thread_id=thread_id
        )

    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def list_guide(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text(
            "Belum ada guidance tersimpan. Tambahkan lewat /addguide, "
            "atau isi langsung file guidance.json."
        )
        return
    lines = ["📚 Daftar guidance yang tersimpan:", ""]
    for item in data:
        kw = ", ".join(item.get("keywords", [])[:3])
        lines.append(f"• [{item['id']}] {item['title']}")
        if kw:
            lines.append(f"    kata kunci: {kw}")
        lines.append("")
    await update.message.reply_text(_safe_telegram_text("\n".join(lines).rstrip()))


# ==============================================================================
# WEB SEARCH (Serper.dev) — general query handler untuk /tanyabot
# ==============================================================================
# Kata pemicu yang menandakan pertanyaan umum/rekomendasi (general intent).
_GENERAL_INTENT_RE = re.compile(
    r"\b(rekomendasi|saran|tempat|wisata|kuliner|makan|jalan[- ]?jalan|liburan|"
    r"long ?weekend|berita|kabar|terbaru|info|cari|carikan|apa itu|siapa|kenapa|"
    r"bagaimana|gimana|tips|cara|resep|review|harga|jadwal|film|lagu|hotel|"
    r"restoran|cafe|kafe|destinasi|libur|weekend|trending|populer|news)\b",
    re.IGNORECASE,
)


def _web_search(query: str, num: int = 5) -> list:
    """Cari via Serper.dev (Google Search API). Return list of dict:
    [{"title", "snippet", "link"}]. Raise Exception kalau gagal / key kosong."""
    api_key = config.SERPER_API_KEY
    if not api_key:
        raise RuntimeError("SERPER_API_KEY belum diset.")

    resp = requests.post(
        "https://google.serper.dev/search",
        headers={"X-API-KEY": api_key, "Content-Type": "application/json"},
        json={"q": query, "gl": config.SERPER_GL, "hl": config.SERPER_HL, "num": num},
        timeout=12,
    )
    resp.raise_for_status()
    data = resp.json()

    results = []

    # Answer box (jawaban langsung dari Google) — kalau ada
    ab = data.get("answerBox")
    if ab:
        ans = ab.get("answer") or ab.get("snippet") or ""
        if ans:
            results.append({
                "title": ab.get("title", "Jawaban"),
                "snippet": ans,
                "link": ab.get("link", ""),
            })

    for item in data.get("organic", [])[:num]:
        results.append({
            "title": item.get("title", "").strip(),
            "snippet": item.get("snippet", "").strip(),
            "link": item.get("link", "").strip(),
        })

    return results[:num]


def _format_web_results(query: str, results: list) -> str:
    """Format hasil pencarian jadi list ringkas untuk Telegram."""
    lines = [
        f"🔎 <b>Hasil pencarian untuk:</b> {html.escape(query)}",
        "━━━━━━━━━━━━━━━━━━━━━━",
        "",
    ]
    for i, r in enumerate(results, 1):
        title = r.get("title") or "(tanpa judul)"
        snippet = r.get("snippet") or ""
        link = r.get("link") or ""
        block = f"<b>{i}. {html.escape(title)}</b>"
        if snippet:
            # Batasi panjang snippet biar ringkas
            if len(snippet) > 200:
                snippet = snippet[:200].rstrip() + "..."
            block += f"\n{html.escape(snippet)}"
        if link:
            block += f"\n🔗 <a href=\"{html.escape(link)}\">{html.escape(link)}</a>"
        lines.append(block)
        lines.append("")

    lines.append("<i>Sumber: Google Search via Serper.dev</i>")
    return "\n".join(lines).rstrip()


async def _answer_kb_or_fallback(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    """Cek KB → jawab. Kalau tidak ketemu, cek guidance. Kalau tetap tidak,
    fallback + simpan pertanyaan ke pending."""
    # 1. Cek Bank Data / Knowledge Base
    faq = kb_store.find_answer(text)
    if faq:
        await update.message.reply_text(
            f"💡 <b>{html.escape(faq.get('question', ''))}</b>\n\n"
            f"{html.escape(faq.get('answer', ''))}",
            parse_mode=ParseMode.HTML,
        )
        return

    # 2. Cek guidance store
    matches = guidance_store.find_matches(text)
    if matches:
        await send_guidance_matches(update, context, text, matches)
        return

    # 3. Cek format tools (cek stock/promo/awb/delreservation/dll) -> eksekusi
    try:
        if await _try_execute_tools(update, context, text):
            return
    except Exception as e:
        logger.error(f"Gagal eksekusi tools dari /tanyabot: {e}")

    # 4. Cek natural "cek tiket #12345"
    tiket_match = CEK_TIKET_RE.search(text)
    if tiket_match and sdp:
        request_id = tiket_match.group(1)
        try:
            req = await asyncio.to_thread(sdp.get_request, request_id)
            detail = await _format_ticket_detail(request_id, req)
            await update.message.reply_text(detail, parse_mode=ParseMode.HTML)
            return
        except Exception as e:
            logger.info(f"Gagal ambil tiket #{request_id} dari /tanyabot: {e}")

    # 5. General query / rekomendasi umum -> Web Search (Serper.dev)
    #    Dipicu kalau ada kata general intent ATAU pertanyaan cukup panjang.
    is_general = bool(_GENERAL_INTENT_RE.search(text)) or len(text.split()) >= 4
    if config.WEB_SEARCH_ENABLED and config.SERPER_API_KEY and is_general:
        try:
            await update.message.reply_text("🔎 Mencari informasi di web...")
            results = await asyncio.to_thread(_web_search, text, 5)
            if results:
                await update.message.reply_text(
                    _format_web_results(text, results),
                    parse_mode=ParseMode.HTML,
                    disable_web_page_preview=True,
                )
                return
        except Exception as e:
            logger.info(f"Web search gagal untuk '{text}': {e}")

    # 6. Tidak ketemu: fallback + simpan pertanyaan ke pending (auto)
    user = update.effective_user
    user_name = user.full_name if user else ""
    user_id = user.id if user else ""
    try:
        kb_store.add_pending(text, user_id, user_name)
    except Exception as e:
        logger.error(f"Gagal simpan pertanyaan pending KB: {e}")

    await update.message.reply_text(
        "Maaf, saya belum punya jawaban untuk pertanyaan ini.\n"
        "Pertanyaan kamu saya pelajari lebih lanjut. Terima kasih!"
    )


@restricted
async def tanya_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /tanyabot <pertanyaan> — tanya ke Bank Data secara eksplisit."""
    question = update.message.text.partition(" ")[2].strip()
    if not question:
        await update.message.reply_text(
            "Tulis pertanyaannya setelah /tanyabot.\n"
            "Contoh: <code>/tanyabot server apa yang sedang issue?</code>",
            parse_mode=ParseMode.HTML,
        )
        return
    await _answer_kb_or_fallback(update, context, question)


async def guide_free_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _is_allowed_chat(update.effective_chat.id):
        return

    chat_type = update.effective_chat.type  # 'private', 'group', 'supergroup'
    text = update.message.text or ""

    # Di GRUP: bot hanya merespons kalau di-mention (@bot) atau reply ke pesan bot.
    # Ini mencegah bot spam menjawab semua obrolan di grup notifikasi.
    if chat_type in ("group", "supergroup"):
        bot_username = (context.bot.username or "").lower()
        is_mention = bot_username and f"@{bot_username}" in text.lower()
        is_reply_to_bot = (
            update.message.reply_to_message is not None
            and update.message.reply_to_message.from_user is not None
            and update.message.reply_to_message.from_user.id == context.bot.id
        )
        if not (is_mention or is_reply_to_bot):
            # Diam saja untuk obrolan biasa; user tetap bisa pakai /tanyabot
            return
        # Bersihkan mention dari teks pertanyaan
        if is_mention:
            text = text.replace(f"@{context.bot.username}", "").strip()

    if not text:
        return

    await _answer_kb_or_fallback(update, context, text)


# ==============================================================================
# KNOWLEDGE BASE / BANK DATA - ADMIN COMMANDS
# ==============================================================================

@restricted
async def addfaq_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /addfaq <pertanyaan> | <jawaban> [| keyword1, keyword2]"""
    raw = update.message.text.partition(" ")[2].strip()  # buang '/addfaq'
    if not raw or "|" not in raw:
        await update.message.reply_text(
            "<b>📚 Tambah FAQ ke Bank Data</b>\n\n"
            "Format:\n"
            "<code>/addfaq pertanyaan | jawaban</code>\n"
            "<code>/addfaq pertanyaan | jawaban | keyword1, keyword2</code>\n\n"
            "Contoh:\n"
            "<code>/addfaq Server apa yang issue? | Server jaringan lokal | server, jaringan, down</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    parts = [p.strip() for p in raw.split("|")]
    question = parts[0]
    answer = parts[1] if len(parts) > 1 else ""
    keywords = None
    if len(parts) > 2 and parts[2]:
        keywords = [k.strip() for k in parts[2].split(",") if k.strip()]

    if not question or not answer:
        await update.message.reply_text(
            "⚠️ Pertanyaan dan jawaban tidak boleh kosong. Format:\n"
            "<code>/addfaq pertanyaan | jawaban</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        entry = kb_store.add_faq(question, answer, keywords)
        kw_text = ", ".join(entry["keywords"]) if entry["keywords"] else "(otomatis dari pertanyaan)"
        await update.message.reply_text(
            f"✅ <b>FAQ berhasil ditambahkan!</b>\n\n"
            f"🆔 ID: <code>{entry['id']}</code>\n"
            f"❓ Pertanyaan: {html.escape(entry['question'])}\n"
            f"💬 Jawaban: {html.escape(entry['answer'])}\n"
            f"🔑 Keywords: {html.escape(kw_text)}",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal tambah FAQ:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )


@restricted
async def listfaq_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /listfaq — tampilkan semua FAQ aktif."""
    faqs = kb_store.list_faqs()
    if not faqs:
        await update.message.reply_text("📭 Belum ada FAQ di Bank Data. Tambah dengan /addfaq.")
        return

    lines = ["<b>📚 Daftar FAQ Aktif</b>", "━━━━━━━━━━━━━━━━━━━━━━", ""]
    for f in faqs:
        kw = ", ".join(f.get("keywords", []))
        lines.append(
            f"🆔 <b>{f.get('id')}</b> — {html.escape(f.get('question', ''))}\n"
            f"   💬 {html.escape(f.get('answer', ''))}\n"
            f"   🔑 {html.escape(kw)}"
        )
        lines.append("")
    lines.append("Hapus dengan: <code>/delfaq &lt;id&gt;</code>")

    text = "\n".join(lines)
    # Potong kalau kepanjangan (batas telegram ~4096)
    if len(text) > 3900:
        text = text[:3900] + "\n\n... (terpotong)"
    await update.message.reply_text(text, parse_mode=ParseMode.HTML)


@restricted
async def delfaq_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /delfaq <id> — hapus FAQ."""
    if not context.args:
        await update.message.reply_text(
            "Format: <code>/delfaq &lt;id&gt;</code>\nContoh: <code>/delfaq 3</code>",
            parse_mode=ParseMode.HTML,
        )
        return
    try:
        faq_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("⚠️ ID harus angka. Contoh: <code>/delfaq 3</code>", parse_mode=ParseMode.HTML)
        return

    if kb_store.delete_faq(faq_id):
        await update.message.reply_text(f"✅ FAQ ID <code>{faq_id}</code> berhasil dihapus.", parse_mode=ParseMode.HTML)
    else:
        await update.message.reply_text(f"ℹ️ FAQ ID <code>{faq_id}</code> tidak ditemukan.", parse_mode=ParseMode.HTML)


@restricted
async def pending_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /pending — tampilkan pertanyaan user yang belum terjawab."""
    pendings = kb_store.list_pending(only_unanswered=True)
    if not pendings:
        await update.message.reply_text("✅ Tidak ada pertanyaan pending. Semua sudah terjawab!")
        return

    lines = ["<b>📥 Pertanyaan Pending (belum terjawab)</b>", "━━━━━━━━━━━━━━━━━━━━━━", ""]
    for p in pendings:
        uname = p.get("user_name") or p.get("user_id", "")
        lines.append(
            f"🆔 <b>{p.get('id')}</b> — dari {html.escape(str(uname))}\n"
            f"   ❓ {html.escape(p.get('question', ''))}\n"
            f"   🕐 {p.get('created_at', '')}"
        )
        lines.append("")
    lines.append("Jawab dengan: <code>/answerfaq &lt;id&gt; | &lt;jawaban&gt;</code>")

    text = "\n".join(lines)
    if len(text) > 3900:
        text = text[:3900] + "\n\n... (terpotong)"
    await update.message.reply_text(text, parse_mode=ParseMode.HTML)


@restricted
async def answerfaq_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /answerfaq <id> | <jawaban> — jawab pertanyaan pending.

    Jawaban akan otomatis dipromosikan menjadi FAQ aktif, dan (jika di grup
    berbeda) tidak dikirim ulang ke user. User bisa tanya lagi untuk dapat jawaban.
    """
    raw = update.message.text.partition(" ")[2].strip()  # buang '/answerfaq'
    if not raw or "|" not in raw:
        await update.message.reply_text(
            "Format:\n<code>/answerfaq &lt;id&gt; | &lt;jawaban&gt;</code>\n\n"
            "Contoh:\n<code>/answerfaq 101 | Silakan restart aplikasi lalu coba lagi.</code>\n\n"
            "Lihat daftar pending dengan /pending",
            parse_mode=ParseMode.HTML,
        )
        return

    id_part, _, answer = raw.partition("|")
    id_part = id_part.strip()
    answer = answer.strip()

    try:
        pending_id = int(id_part)
    except ValueError:
        await update.message.reply_text("⚠️ ID harus angka. Lihat /pending untuk daftar ID.", parse_mode=ParseMode.HTML)
        return

    if not answer:
        await update.message.reply_text("⚠️ Jawaban tidak boleh kosong.", parse_mode=ParseMode.HTML)
        return

    result = kb_store.answer_pending(pending_id, answer)
    if result is None:
        await update.message.reply_text(
            f"ℹ️ Pertanyaan pending ID <code>{pending_id}</code> tidak ditemukan.",
            parse_mode=ParseMode.HTML,
        )
        return

    pending = result["pending"]
    faq = result["faq"]

    # Konfirmasi ke admin
    await update.message.reply_text(
        f"✅ <b>Pertanyaan terjawab & ditambahkan ke FAQ!</b>\n\n"
        f"❓ {html.escape(pending.get('question', ''))}\n"
        f"💬 {html.escape(answer)}\n\n"
        f"FAQ baru dibuat dengan ID <code>{faq.get('id')}</code>.",
        parse_mode=ParseMode.HTML,
    )

    # Coba kirim jawaban langsung ke user penanya (best-effort)
    user_id = pending.get("user_id")
    if user_id:
        try:
            await context.bot.send_message(
                chat_id=int(user_id),
                text=(
                    f"💡 <b>Jawaban untuk pertanyaan kamu:</b>\n\n"
                    f"❓ {html.escape(pending.get('question', ''))}\n\n"
                    f"{html.escape(answer)}"
                ),
                parse_mode=ParseMode.HTML,
            )
        except Exception as e:
            logger.info(f"Tidak bisa kirim jawaban langsung ke user {user_id}: {e}")


@restricted
async def addguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "Judul guidance-nya apa?\ncontoh: Cara mengatasi koneksi jaringan error"
    )
    return ADD_GUIDE_TITLE


async def addguide_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["guide_title"] = update.message.text.strip()
    await update.message.reply_text(
        "Kata kunci pemicunya apa saja? Pisahkan dengan koma.\n"
        "Ini dipakai supaya bot bisa mengenali kapan harus munculkan guidance ini "
        "(lewat /guide maupun kalau kamu ketik bebas).\n\n"
        "Contoh: koneksi jaringan error, jaringan error, network issue"
    )
    return ADD_GUIDE_KEYWORDS


async def addguide_keywords(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keywords = [k.strip() for k in update.message.text.split(",") if k.strip()]
    if not keywords:
        await update.message.reply_text("Kata kuncinya kosong, coba lagi (pisahkan dengan koma):")
        return ADD_GUIDE_KEYWORDS
    context.user_data["guide_keywords"] = keywords
    await update.message.reply_text(
        "Sekarang kirim isi/detail guidance-nya (boleh panjang & multi-baris).\n"
        "Atau kalau ada file pendukung (script, dokumen, dll), langsung upload file-nya "
        "di sini juga boleh — caption di file-nya akan dipakai sebagai deskripsi."
    )
    return ADD_GUIDE_CONTENT


async def addguide_content(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document:
        doc = update.message.document
        context.user_data["guide_content"] = (update.message.caption or "").strip()
        context.user_data["guide_attachment"] = {
            "file_id": doc.file_id,
            "file_name": doc.file_name,
        }
    else:
        context.user_data["guide_content"] = update.message.text.strip()
        context.user_data["guide_attachment"] = None

    context.user_data["guide_action"] = None
    keyboard = [
        [
            InlineKeyboardButton("Ya", callback_data="ga_yes"),
            InlineKeyboardButton("Tidak", callback_data="ga_no"),
        ]
    ]
    await update.message.reply_text(
        "Apakah guidance ini bisa dieksekusi otomatis oleh bot (menjalankan script "
        "PowerShell di laptop kamu)?",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return ADD_GUIDE_ACTION_ASK


async def addguide_action_ask(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "ga_no":
        return await addguide_show_preview(update, context)
    await query.edit_message_text(
        "Masukkan path lengkap file script-nya di laptop kamu.\n"
        r"Contoh: D:\promovoucher\cancel_reservation.ps1"
    )
    return ADD_GUIDE_ACTION_SCRIPT


async def addguide_action_script(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["action_script_path"] = update.message.text.strip()
    await update.message.reply_text(
        "Nama parameter/flag-nya apa? (sesuai yang dipakai script)\n"
        "Kalau lebih dari 1 parameter, pisahkan dengan koma.\n\n"
        "Contoh 1 parameter: -CsvFile\n"
        "Contoh banyak parameter: -ArticleId, -SourceId"
    )
    return ADD_GUIDE_ACTION_FLAG


async def addguide_action_flag(update: Update, context: ContextTypes.DEFAULT_TYPE):
    flags = []
    for part in update.message.text.split(","):
        f = part.strip()
        if not f:
            continue
        if not f.startswith("-"):
            f = "-" + f
        flags.append(f)
    if not flags:
        await update.message.reply_text("Nama parameternya kosong, coba lagi:")
        return ADD_GUIDE_ACTION_FLAG

    context.user_data["action_flags"] = flags

    if len(flags) == 1:
        keyboard = [
            [
                InlineKeyboardButton("Teks bebas", callback_data="atype_text"),
                InlineKeyboardButton("Upload file", callback_data="atype_file"),
            ]
        ]
        await update.message.reply_text(
            "Nilai parameter ini nanti diminta dalam bentuk apa?",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return ADD_GUIDE_ACTION_TYPE

    keyboard = [
        [InlineKeyboardButton("Satu-satu per parameter", callback_data="amode_separate")],
        [InlineKeyboardButton("Baris per baris (berpasangan)", callback_data="amode_paired")],
    ]
    await update.message.reply_text(
        "Cara input datanya nanti gimana?\n\n"
        "• Satu-satu per parameter: bot tanya tiap parameter terpisah "
        f"({', '.join(flags)}), boleh multi-nilai dipisah koma tiap parameter.\n"
        "• Baris per baris (berpasangan): kamu isi sekali jalan, satu baris = "
        f"satu kombinasi ({', '.join(f.lstrip('-') for f in flags)}), contoh:\n"
        "8000044321, SS20\n8000044322, SS21",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return ADD_GUIDE_ACTION_MODE


async def addguide_action_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    flags = context.user_data.get("action_flags", [])
    input_mode = "paired" if query.data == "amode_paired" else "separate"
    action = {
        "script_path": context.user_data.get("action_script_path", ""),
        "params": [{"flag": f, "type": "text"} for f in flags],
    }
    if input_mode == "paired":
        action["input_mode"] = "paired"
    context.user_data["guide_action"] = action
    return await addguide_show_preview(update, context)


async def addguide_action_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    param_type = "file" if query.data == "atype_file" else "text"
    flags = context.user_data.get("action_flags", [])
    context.user_data["guide_action"] = {
        "script_path": context.user_data.get("action_script_path", ""),
        "params": [{"flag": flags[0], "type": param_type}] if flags else [],
    }
    return await addguide_show_preview(update, context)


async def addguide_show_preview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    file_line = f"\nFile: {d['guide_attachment']['file_name']}" if d.get("guide_attachment") else ""
    action = d.get("guide_action")
    action_line = ""
    if action:
        param_bits = [
            f"{p['flag']} ({'file' if p.get('type') == 'file' else 'teks'})"
            for p in action.get("params", [])
        ]
        action_line = f"\nAction: {action['script_path']} | " + ", ".join(param_bits)
        if action.get("input_mode") == "paired":
            action_line += " | mode: baris berpasangan"
    preview = (
        "Konfirmasi guidance berikut:\n\n"
        f"Judul: {d['guide_title']}\n"
        f"Kata kunci: {', '.join(d['guide_keywords'])}\n"
        f"{file_line}"
        f"{action_line}\n"
        f"Isi:\n{d['guide_content'] or '(tidak ada teks deskripsi)'}\n\n"
        "Simpan guidance ini?"
    )
    keyboard = [
        [
            InlineKeyboardButton("Ya, simpan", callback_data="guide_save_yes"),
            InlineKeyboardButton("Batal", callback_data="guide_save_no"),
        ]
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(preview, reply_markup=markup)
    else:
        await update.message.reply_text(preview, reply_markup=markup)
    return ADD_GUIDE_CONFIRM


async def addguide_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "guide_save_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    d = context.user_data
    entry = guidance_store.add_guidance(
        d["guide_title"],
        d["guide_keywords"],
        d["guide_content"],
        d.get("guide_attachment"),
        d.get("guide_action"),
    )
    await query.edit_message_text(f"Guidance '{entry['title']}' berhasil disimpan. ✅ (ID: {entry['id']})")
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def delguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text("Belum ada guidance tersimpan.")
        return ConversationHandler.END
    buttons = [
        [InlineKeyboardButton(f"[{d['id']}] {d['title']}", callback_data=f"delg_{d['id']}")]
        for d in data
    ]
    await update.message.reply_text(
        "Pilih guidance yang mau dihapus:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return DEL_GUIDE_PICK


async def delguide_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("delg_", "")
    context.user_data["del_guide_id"] = gid
    entry = guidance_store.get_guidance(gid)
    title = entry["title"] if entry else gid
    keyboard = [
        [
            InlineKeyboardButton("Ya, hapus", callback_data="delg_yes"),
            InlineKeyboardButton("Batal", callback_data="delg_no"),
        ]
    ]
    await query.edit_message_text(
        f"Yakin mau hapus guidance '{title}'?", reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return DEL_GUIDE_CONFIRM


async def delguide_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "delg_no":
        await query.edit_message_text("Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END
    gid = context.user_data.get("del_guide_id")
    ok = guidance_store.delete_guidance(gid)
    await query.edit_message_text(
        "Guidance berhasil dihapus. 🗑️" if ok else "Gagal menghapus (mungkin sudah dihapus)."
    )
    context.user_data.clear()
    return ConversationHandler.END


@restricted
async def editguide_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    data = guidance_store.load_guidance()
    if not data:
        await update.message.reply_text("Belum ada guidance tersimpan.")
        return ConversationHandler.END
    buttons = [
        [InlineKeyboardButton(f"[{d['id']}] {d['title']}", callback_data=f"editg_{d['id']}")]
        for d in data
    ]
    await update.message.reply_text(
        "Pilih guidance yang mau diedit:", reply_markup=InlineKeyboardMarkup(buttons)
    )
    return EDIT_GUIDE_PICK


async def editguide_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    gid = query.data.replace("editg_", "")
    entry = guidance_store.get_guidance(gid)
    if not entry:
        await query.edit_message_text("Guidance tidak ditemukan (mungkin sudah dihapus).")
        return ConversationHandler.END
    context.user_data["edit_guide_id"] = gid
    context.user_data["edit_title"] = entry["title"]
    context.user_data["edit_keywords"] = entry.get("keywords", [])
    context.user_data["edit_content"] = entry.get("content", "")
    context.user_data["edit_attachment"] = entry.get("attachment")
    context.user_data["edit_action"] = entry.get("action")
    return await show_edit_menu(update, context)


async def show_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    d = context.user_data
    file_line = ""
    if d.get("edit_attachment"):
        file_line = f"\nFile terlampir: {d['edit_attachment']['file_name']}"
    content_preview = d["edit_content"][:200] or "(kosong)"
    action = d.get("edit_action")
    action_line = ""
    if action:
        param_bits = [
            f"{p['flag']} ({'file' if p.get('type') == 'file' else 'teks'})"
            for p in _action_param_list(action)
        ]
        action_line = f"\nAction: {action.get('script_path','')} | " + ", ".join(param_bits)
        if action.get("input_mode") == "paired":
            action_line += " | mode: baris berpasangan"
    text = (
        "Mengedit guidance:\n\n"
        f"Judul: {d['edit_title']}\n"
        f"Kata kunci: {', '.join(d['edit_keywords'])}"
        f"{file_line}"
        f"{action_line}\n"
        f"Isi: {content_preview}\n\n"
        "Mau ubah bagian mana?"
    )
    keyboard = [
        [
            InlineKeyboardButton("✏️ Judul", callback_data="ef_title"),
            InlineKeyboardButton("✏️ Kata kunci", callback_data="ef_keywords"),
        ],
        [InlineKeyboardButton("✏️ Isi / File", callback_data="ef_content")],
        [
            InlineKeyboardButton(
                "⚙️ Hapus Action" if action else "⚙️ Tambah Action", callback_data="ef_action"
            )
        ],
        [
            InlineKeyboardButton("✅ Selesai & Simpan", callback_data="ef_done"),
            InlineKeyboardButton("Batal", callback_data="ef_cancel"),
        ],
    ]
    markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(text, reply_markup=markup)
    else:
        await update.message.reply_text(text, reply_markup=markup)
    return EDIT_GUIDE_MENU


async def editguide_menu_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    choice = query.data

    if choice == "ef_cancel":
        await query.edit_message_text("Dibatalkan, tidak ada perubahan yang disimpan.")
        context.user_data.clear()
        return ConversationHandler.END

    if choice == "ef_done":
        d = context.user_data
        keywords = d["edit_keywords"]
        if isinstance(keywords, str):
            keywords = [k.strip() for k in keywords.split(",") if k.strip()]
        updates = {
            "title": d["edit_title"],
            "keywords": [k.strip().lower() for k in keywords],
            "content": d["edit_content"],
            "attachment": d.get("edit_attachment"),
            "action": d.get("edit_action"),
        }
        entry = guidance_store.update_guidance(d["edit_guide_id"], updates)
        if entry:
            await query.edit_message_text(f"Guidance '{entry['title']}' berhasil diperbarui. ✅")
        else:
            await query.edit_message_text("Gagal menyimpan perubahan (guidance mungkin sudah dihapus).")
        context.user_data.clear()
        return ConversationHandler.END

    if choice == "ef_action":
        if context.user_data.get("edit_action"):
            context.user_data["edit_action"] = None
            return await show_edit_menu(update, context)
        await query.edit_message_text(
            "Masukkan path lengkap file script-nya di laptop kamu.\n"
            r"Contoh: D:\promovoucher\cancel_reservation.ps1"
        )
        return EDIT_GUIDE_ACTION_SCRIPT

    if choice == "ef_title":
        await query.edit_message_text("Ketik judul baru:")
        return EDIT_GUIDE_TITLE
    if choice == "ef_keywords":
        await query.edit_message_text("Ketik kata kunci baru, pisahkan dengan koma:")
        return EDIT_GUIDE_KEYWORDS
    if choice == "ef_content":
        await query.edit_message_text(
            "Kirim isi/detail baru (teks biasa, atau upload file kalau ada)."
        )
        return EDIT_GUIDE_CONTENT


async def editguide_title(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["edit_title"] = update.message.text.strip()
    return await show_edit_menu(update, context)


async def editguide_keywords(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keywords = [k.strip() for k in update.message.text.split(",") if k.strip()]
    if not keywords:
        await update.message.reply_text("Kata kuncinya kosong, coba lagi (pisahkan dengan koma):")
        return EDIT_GUIDE_KEYWORDS
    context.user_data["edit_keywords"] = keywords
    return await show_edit_menu(update, context)


async def editguide_content(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.document:
        doc = update.message.document
        context.user_data["edit_content"] = (update.message.caption or "").strip()
        context.user_data["edit_attachment"] = {"file_id": doc.file_id, "file_name": doc.file_name}
    else:
        context.user_data["edit_content"] = update.message.text.strip()
        context.user_data["edit_attachment"] = None
    return await show_edit_menu(update, context)


async def editguide_action_script(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["_edit_action_script_path"] = update.message.text.strip()
    await update.message.reply_text(
        "Nama parameter/flag-nya apa? (sesuai yang dipakai script)\n"
        "Kalau lebih dari 1 parameter, pisahkan dengan koma.\n\n"
        "Contoh 1 parameter: -CsvFile\n"
        "Contoh banyak parameter: -ArticleId, -SourceId"
    )
    return EDIT_GUIDE_ACTION_FLAG


async def editguide_action_flag(update: Update, context: ContextTypes.DEFAULT_TYPE):
    flags = []
    for part in update.message.text.split(","):
        f = part.strip()
        if not f:
            continue
        if not f.startswith("-"):
            f = "-" + f
        flags.append(f)
    if not flags:
        await update.message.reply_text("Nama parameternya kosong, coba lagi:")
        return EDIT_GUIDE_ACTION_FLAG

    context.user_data["_edit_action_flags"] = flags

    if len(flags) == 1:
        keyboard = [
            [
                InlineKeyboardButton("Teks bebas", callback_data="etype_text"),
                InlineKeyboardButton("Upload file", callback_data="etype_file"),
            ]
        ]
        await update.message.reply_text(
            "Nilai parameter ini nanti diminta dalam bentuk apa?",
            reply_markup=InlineKeyboardMarkup(keyboard),
        )
        return EDIT_GUIDE_ACTION_TYPE

    keyboard = [
        [InlineKeyboardButton("Satu-satu per parameter", callback_data="emode_separate")],
        [InlineKeyboardButton("Baris per baris (berpasangan)", callback_data="emode_paired")],
    ]
    await update.message.reply_text(
        "Cara input datanya nanti gimana?\n\n"
        "• Satu-satu per parameter: bot tanya tiap parameter terpisah "
        f"({', '.join(flags)}), boleh multi-nilai dipisah koma tiap parameter.\n"
        "• Baris per baris (berpasangan): kamu isi sekali jalan, satu baris = "
        f"satu kombinasi ({', '.join(f.lstrip('-') for f in flags)}), contoh:\n"
        "8000044321, SS20\n8000044322, SS21",
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    return EDIT_GUIDE_ACTION_MODE


async def editguide_action_mode(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    flags = context.user_data.get("_edit_action_flags", [])
    input_mode = "paired" if query.data == "emode_paired" else "separate"
    action = {
        "script_path": context.user_data.pop("_edit_action_script_path", ""),
        "params": [{"flag": f, "type": "text"} for f in flags],
    }
    if input_mode == "paired":
        action["input_mode"] = "paired"
    context.user_data["edit_action"] = action
    return await show_edit_menu(update, context)


async def editguide_action_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    param_type = "file" if query.data == "etype_file" else "text"
    flags = context.user_data.get("_edit_action_flags", [])
    context.user_data["edit_action"] = {
        "script_path": context.user_data.pop("_edit_action_script_path", ""),
        "params": [{"flag": flags[0], "type": param_type}] if flags else [],
    }
    return await show_edit_menu(update, context)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Terjadi error tak terduga:", exc_info=context.error)

    if isinstance(context.error, (TimedOut, NetworkError, BadRequest, Conflict)):
        return

    target_chat_id = None
    target_thread_id = None
    if isinstance(update, Update) and update.effective_chat:
        target_chat_id = update.effective_chat.id
        target_thread_id = _thread_id_from_update(update)
    elif config.TELEGRAM_USER_ID:
        target_chat_id = config.TELEGRAM_USER_ID

    if not target_chat_id:
        return

    try:
        await context.bot.send_message(
            chat_id=target_chat_id,
            message_thread_id=target_thread_id,
            text=(
                "⚠️ Terjadi error saat memproses perintah terakhir kamu:\n"
                f"{context.error}\n\n"
                "Coba ulangi lagi, atau ketik /cancel lalu mulai ulang."
            ),
        )
    except Exception:
        pass  # Jangan log lagi untuk menghindari loop error


SDP_NOTIFY_STATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "sdp_notify_state.json")


def _load_sdp_notify_state() -> dict:
    try:
        with open(SDP_NOTIFY_STATE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {"last_seen_id": 0}


def _save_sdp_notify_state(state: dict):
    try:
        with open(SDP_NOTIFY_STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f)
    except Exception as e:
        logger.error(f"Gagal menyimpan sdp notify state: {e}")


async def check_new_sdp_tickets(context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    state = _load_sdp_notify_state()
    last_seen_id = state.get("last_seen_id", 0)

    try:
        requests_list = await asyncio.to_thread(
            sdp.list_requests, 30, None, config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal cek tiket baru ServiceDesk Plus: {e}")
        return

    max_id_seen = last_seen_id
    new_items = []
    for req in requests_list:
        try:
            rid = int(req.get("id", 0))
        except (TypeError, ValueError):
            continue
        max_id_seen = max(max_id_seen, rid)
        if rid > last_seen_id:
            new_items.append((rid, req))

    if last_seen_id == 0:
        state["last_seen_id"] = max_id_seen
        _save_sdp_notify_state(state)
        return

    new_items.sort(key=lambda pair: pair[0])
    for _, req in new_items:
        group_name = (req.get("group") or {}).get("name", "")
        text = f"*New Ticket ({group_name}):\n\n{_format_sdp_request_line(req)}"
        await _broadcast_notify(context, text)

    if new_items:
        state["last_seen_id"] = max_id_seen
        _save_sdp_notify_state(state)


@restricted
async def sdreminder_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        await update.message.reply_text(
            "Fitur ini butuh SDP_NOTIFY_GROUPS diisi dulu di .env (nama-nama group yang mau dipantau)."
        )
        return

    parts = update.message.text.split(maxsplit=1)
    arg = parts[1].strip().lower() if len(parts) > 1 else ""
    state = _load_sdp_notify_state()

    if not arg or arg == "status":
        interval = state.get("open_reminder_interval_minutes", config.SDP_OPEN_REMINDER_DEFAULT_MINUTES)
        if interval and interval > 0:
            await update.message.reply_text(
                f"Reminder tiket Open sedang AKTIF, tiap {interval} menit.\n\n"
                "Ubah interval: /sdreminder <menit>\n"
                "Matikan: /sdreminder off"
            )
        else:
            await update.message.reply_text(
                "Reminder tiket Open sedang NONAKTIF.\n\n"
                "Aktifkan: /sdreminder <menit>, contoh: /sdreminder 30"
            )
        return

    if arg == "off":
        state["open_reminder_interval_minutes"] = 0
        _save_sdp_notify_state(state)
        await update.message.reply_text("Reminder tiket Open dimatikan.")
        return

    if not arg.isdigit() or int(arg) <= 0:
        await update.message.reply_text(
            "Ketik salah satu:\n"
            "/sdreminder <menit> - aktifkan/ubah interval, contoh: /sdreminder 30\n"
            "/sdreminder off - matikan\n"
            "/sdreminder status - cek status sekarang"
        )
        return

    minutes = int(arg)
    state["open_reminder_interval_minutes"] = minutes
    state["last_open_reminder_at"] = None
    _save_sdp_notify_state(state)
    await update.message.reply_text(
        f"Oke, reminder tiket Open sekarang aktif tiap {minutes} menit untuk group:\n"
        f"{', '.join(config.SDP_NOTIFY_GROUPS)}"
    )


# ==============================================================================
# SLA / OVERDUE MONITOR (deteksi tiket yang lewat DueBy)
# ==============================================================================
# Status yang dianggap "sudah selesai" — tidak perlu dicek SLA-nya
_SLA_DONE_STATUSES = {"closed", "resolved", "completed"}


def _epoch_from_sdp_time(time_val) -> float:
    """Ambil epoch (detik) dari field waktu SDP (bisa dict {value: ms} atau angka ms)."""
    if not time_val:
        return 0
    try:
        if isinstance(time_val, dict):
            raw = time_val.get("value", 0)
        else:
            raw = time_val
        raw = int(raw)
        # SDP return epoch dalam milidetik
        return raw / 1000 if raw > 1e12 else raw
    except (ValueError, TypeError):
        return 0


def _get_overdue_tickets() -> list:
    """Ambil semua tiket aktif yang sudah lewat DueBy (SLA) untuk group yang dikonfigurasi.

    Return list of dict: {id, subject, technician, status, due_by_display, delay_str, due_epoch}
    """
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return []

    active_statuses = [
        "Open",
        "In Progress Investigation",
        "Transfer L1",
        "Waiting User Confirmation",
        "Pending",
        "Onhold",
    ]

    now_epoch = dt.datetime.now(TZ).timestamp()
    seen_ids = set()
    overdue = []

    for status in active_statuses:
        try:
            tickets = sdp.list_requests(100, status, config.SDP_NOTIFY_GROUPS)
        except SDPError as e:
            logger.error(f"Gagal ambil tiket status {status} untuk cek SLA: {e}")
            continue

        for t in tickets:
            tid = t.get("id")
            if not tid or tid in seen_ids:
                continue

            due_val = t.get("due_by_time")
            due_epoch = _epoch_from_sdp_time(due_val)
            is_overdue_flag = t.get("is_overdue")

            # Tiket overdue kalau flag is_overdue True, ATAU due_by sudah lewat
            is_over = bool(is_overdue_flag) or (due_epoch > 0 and due_epoch < now_epoch)
            if not is_over:
                continue

            seen_ids.add(tid)

            due_display = due_val.get("display_value", "") if isinstance(due_val, dict) else ""
            # Hitung lama delay
            delay_str = ""
            if due_epoch > 0:
                delta_sec = now_epoch - due_epoch
                if delta_sec > 0:
                    days = int(delta_sec // 86400)
                    hours = int((delta_sec % 86400) // 3600)
                    mins = int((delta_sec % 3600) // 60)
                    parts = []
                    if days:
                        parts.append(f"{days} hari")
                    if hours:
                        parts.append(f"{hours} jam")
                    if mins and not days:
                        parts.append(f"{mins} menit")
                    delay_str = " ".join(parts) if parts else "baru saja"

            overdue.append({
                "id": tid,
                "subject": t.get("subject", ""),
                "technician": (t.get("technician") or {}).get("name", "") if t.get("technician") else "",
                "status": (t.get("status") or {}).get("name", ""),
                "due_by_display": due_display,
                "delay_str": delay_str,
                "due_epoch": due_epoch,
            })

    # Urutkan dari yang paling lama overdue (due_epoch terkecil dulu)
    overdue.sort(key=lambda x: x["due_epoch"] if x["due_epoch"] > 0 else float("inf"))
    return overdue


def _format_overdue_message(overdue: list, max_show: int = 20) -> str:
    """Susun pesan notifikasi tiket overdue."""
    lines = [
        "⚠️ <b>SLA OverDue Alert</b>",
        "━━━━━━━━━━━━━━━━━━━━━━",
        f"Ada <b>{len(overdue)}</b> tiket yang sudah lewat batas SLA:",
        "",
    ]

    for t in overdue[:max_show]:
        tech = t["technician"] or "Unassigned"
        line = (
            f"🎫 <b>#{t['id']}</b> — {html.escape(t['subject'])}\n"
            f"   👤 {html.escape(tech)} | 📌 {t['status']}\n"
            f"   ⏰ DueBy: {t['due_by_display']}"
        )
        if t["delay_str"]:
            line += f" (delay {t['delay_str']})"
        lines.append(line)
        lines.append("")

    if len(overdue) > max_show:
        lines.append(f"...dan {len(overdue) - max_show} tiket overdue lainnya.")

    lines.append(f"🕐 {dt.datetime.now(TZ).strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━")
    return "\n".join(lines).rstrip()


async def check_sla_overdue_tickets(context: ContextTypes.DEFAULT_TYPE):
    """Background job: cek tiket yang lewat SLA dan kirim notif (anti-spam per tiket)."""
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    try:
        overdue = await asyncio.to_thread(_get_overdue_tickets)
    except Exception as e:
        logger.error(f"SLA OverDue Monitor error: {e}")
        return

    if not overdue:
        return

    # Anti-spam: hanya notif tiket yang BELUM pernah dinotif (atau reset harian)
    state = _load_sdp_notify_state()
    today_str = dt.datetime.now(TZ).strftime("%Y-%m-%d")
    notified = state.get("overdue_notified", {})

    # Reset kalau ganti hari
    if notified.get("_date") != today_str:
        notified = {"_date": today_str}

    new_overdue = [t for t in overdue if str(t["id"]) not in notified]

    if not new_overdue:
        logger.info("SLA OverDue Monitor: tidak ada tiket overdue baru.")
        return

    await _broadcast_notify(context, _format_overdue_message(new_overdue))

    # Tandai sudah dinotif
    for t in new_overdue:
        notified[str(t["id"])] = True
    state["overdue_notified"] = notified
    _save_sdp_notify_state(state)
    logger.info(f"SLA OverDue Monitor: kirim notif {len(new_overdue)} tiket overdue baru.")


@restricted
async def overdue_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /overdue — cek manual tiket yang lewat SLA."""
    await update.message.reply_text("⏳ Mengecek tiket yang lewat SLA...")
    try:
        overdue = await asyncio.to_thread(_get_overdue_tickets)
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal cek tiket overdue:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    if not overdue:
        await update.message.reply_text(
            "✅ Tidak ada tiket yang lewat SLA saat ini.",
            parse_mode=ParseMode.HTML,
        )
        return

    await update.message.reply_text(
        _format_overdue_message(overdue),
        parse_mode=ParseMode.HTML,
    )


# ==============================================================================
# QUOTE OF THE DAY (/quote) — daftar lokal bahasa Indonesia
# ==============================================================================
_LOCAL_QUOTES = [
    {"text": "Kesuksesan adalah kemampuan untuk melangkah dari satu kegagalan ke kegagalan lain tanpa kehilangan semangat.", "author": "Winston Churchill"},
    {"text": "Satu-satunya cara melakukan pekerjaan hebat adalah dengan mencintai apa yang kamu kerjakan.", "author": "Steve Jobs"},
    {"text": "Jangan menilai setiap hari dari panen yang kamu tuai, tapi dari benih yang kamu tanam.", "author": "Robert Louis Stevenson"},
    {"text": "Masa depan milik mereka yang percaya pada keindahan mimpi mereka.", "author": "Eleanor Roosevelt"},
    {"text": "Cara terbaik memprediksi masa depan adalah dengan menciptakannya.", "author": "Peter Drucker"},
    {"text": "Kualitas bukanlah suatu tindakan, melainkan sebuah kebiasaan.", "author": "Aristoteles"},
    {"text": "Mulailah dari mana kamu berada. Gunakan apa yang kamu punya. Lakukan apa yang kamu bisa.", "author": "Arthur Ashe"},
    {"text": "Kerja keras mengalahkan bakat ketika bakat tidak bekerja keras.", "author": "Tim Notke"},
    {"text": "Jangan takut gagal. Takutlah untuk tidak mencoba.", "author": "Anonim"},
    {"text": "Hidup itu seperti mengendarai sepeda. Untuk menjaga keseimbangan, kamu harus terus bergerak.", "author": "Albert Einstein"},
    {"text": "Orang yang tidak pernah melakukan kesalahan adalah orang yang tidak pernah mencoba hal baru.", "author": "Albert Einstein"},
    {"text": "Semua impian kita bisa menjadi kenyataan jika kita punya keberanian untuk mengejarnya.", "author": "Walt Disney"},
    {"text": "Kebahagiaan bukanlah sesuatu yang sudah jadi. Ia datang dari tindakanmu sendiri.", "author": "Dalai Lama"},
    {"text": "Jangan menunggu waktu yang tepat, karena waktu tidak akan pernah benar-benar tepat.", "author": "Napoleon Hill"},
    {"text": "Yang membedakan orang sukses dan tidak sukses bukan kekuatan atau pengetahuan, tapi kemauan.", "author": "Vince Lombardi"},
    {"text": "Kegagalan hanyalah kesempatan untuk memulai lagi dengan lebih cerdas.", "author": "Henry Ford"},
    {"text": "Percaya pada dirimu sendiri dan semua yang kamu miliki. Kamu lebih kuat dari yang kamu kira.", "author": "Anonim"},
    {"text": "Sukses adalah hasil dari persiapan, kerja keras, dan belajar dari kegagalan.", "author": "Colin Powell"},
    {"text": "Jangan pernah menyerah pada mimpi hanya karena butuh waktu untuk mewujudkannya.", "author": "Anonim"},
    {"text": "Perubahan besar dimulai dari langkah kecil yang konsisten.", "author": "Anonim"},
    {"text": "Lakukan hari ini apa yang orang lain tidak mau, agar besok kamu bisa melakukan apa yang orang lain tidak bisa.", "author": "Jerry Rice"},
    {"text": "Kesabaran adalah kunci. Segala sesuatu yang baik butuh waktu untuk tumbuh.", "author": "Anonim"},
    {"text": "Fokus pada kemajuan, bukan kesempurnaan.", "author": "Anonim"},
    {"text": "Semakin keras kamu bekerja untuk sesuatu, semakin besar rasa bangga saat mencapainya.", "author": "Anonim"},
    {"text": "Jangan biarkan kemarin menghabiskan terlalu banyak waktu hari ini.", "author": "Will Rogers"},
    {"text": "Sikapmu, bukan bakatmu, yang menentukan seberapa tinggi kamu bisa mencapai.", "author": "Zig Ziglar"},
    {"text": "Setiap ahli pernah menjadi pemula. Teruslah belajar.", "author": "Anonim"},
    {"text": "Motivasi membuatmu memulai. Kebiasaan membuatmu terus melangkah.", "author": "Jim Ryun"},
    {"text": "Hidup dimulai di akhir zona nyamanmu.", "author": "Neale Donald Walsch"},
    {"text": "Jika kamu ingin mencapai hal-hal hebat, berhentilah meminta izin.", "author": "Anonim"},
    {"text": "Rezeki tidak akan tertukar. Yang penting terus berusaha dan berdoa.", "author": "Anonim"},
    {"text": "Bekerja dalam diam, biarkan kesuksesanmu yang bersuara.", "author": "Frank Ocean"},
    {"text": "Tidak ada yang mustahil bagi orang yang mau berusaha.", "author": "Alexander Agung"},
    {"text": "Kesulitan hari ini adalah kekuatan untuk masa depan.", "author": "Anonim"},
    {"text": "Bermimpilah besar dan beranilah gagal.", "author": "Norman Vaughan"},
    {"text": "Jadilah versi terbaik dari dirimu, bukan tiruan dari orang lain.", "author": "Anonim"},
    {"text": "Disiplin adalah jembatan antara tujuan dan pencapaian.", "author": "Jim Rohn"},
    {"text": "Waktu yang paling tepat untuk memulai adalah sekarang.", "author": "Anonim"},
    {"text": "Kekuatan tidak datang dari kemenangan. Perjuanganmulah yang menumbuhkan kekuatan.", "author": "Arnold Schwarzenegger"},
    {"text": "Jangan berhenti ketika lelah. Berhentilah ketika selesai.", "author": "Anonim"},
]


@restricted
async def quote_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /quote — quote motivasi harian. Utamakan Google Sheets (tab Quotes),
    fallback ke daftar lokal kalau Sheets kosong/gagal."""
    import random

    quote = None
    try:
        quote = await asyncio.to_thread(kb_store.random_quote)
    except Exception as e:
        logger.info(f"Gagal ambil quote dari Sheets, pakai lokal: {e}")

    if not quote:
        quote = random.choice(_LOCAL_QUOTES)

    await update.message.reply_text(
        f"💬 <b>Quote Hari Ini:</b>\n\n"
        f"\"{html.escape(quote['text'])}\"\n"
        f"— <i>{html.escape(quote['author'])}</i>",
        parse_mode=ParseMode.HTML,
    )


@restricted
async def addquote_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /addquote <teks> | <author> — tambah quote ke Google Sheets."""
    raw = update.message.text.partition(" ")[2].strip()
    if not raw:
        await update.message.reply_text(
            "<b>💬 Tambah Quote</b>\n\n"
            "Format:\n"
            "<code>/addquote teks quote | nama author</code>\n\n"
            "Contoh:\n"
            "<code>/addquote Kerja keras tidak mengkhianati hasil | Anonim</code>\n\n"
            "Author opsional — kalau tidak diisi otomatis 'Anonim'.",
            parse_mode=ParseMode.HTML,
        )
        return

    if "|" in raw:
        text_part, _, author_part = raw.partition("|")
        text = text_part.strip()
        author = author_part.strip() or "Anonim"
    else:
        text = raw
        author = "Anonim"

    if not text:
        await update.message.reply_text("⚠️ Teks quote tidak boleh kosong.", parse_mode=ParseMode.HTML)
        return

    try:
        entry = await asyncio.to_thread(kb_store.add_quote, text, author)
        await update.message.reply_text(
            f"✅ <b>Quote berhasil ditambahkan!</b>\n\n"
            f"\"{html.escape(entry['text'])}\"\n"
            f"— <i>{html.escape(entry['author'])}</i>",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal tambah quote:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )


@restricted
async def seedquotes_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /seedquotes — isi tab Quotes di Sheets dengan daftar quote lokal
    (hanya jika tab masih kosong). Berguna sekali di awal migrasi."""
    await update.message.reply_text("⏳ Menyiapkan quote di Google Sheets...")
    try:
        existing = await asyncio.to_thread(kb_store.list_quotes)
        if existing:
            await update.message.reply_text(
                f"ℹ️ Tab Quotes sudah berisi <b>{len(existing)}</b> quote. Seeding dilewati.",
                parse_mode=ParseMode.HTML,
            )
            return

        def _seed():
            for q in _LOCAL_QUOTES:
                kb_store.add_quote(q["text"], q["author"])
            return len(_LOCAL_QUOTES)

        count = await asyncio.to_thread(_seed)
        await update.message.reply_text(
            f"✅ Berhasil menambahkan <b>{count}</b> quote ke tab Quotes di Google Sheets.\n"
            "Sekarang kamu bisa kelola quote langsung dari spreadsheet.",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal seeding quotes:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )


# ==============================================================================
# COFFEE RECOMMENDATION (/coffee) — "endpoint internal" /api/coffee (mock lokal)
# ==============================================================================
# Database/config lokal rekomendasi kopi + kata penenang.
_COFFEE_MENU = [
    {"name": "Espresso", "desc": "Shot kopi pekat untuk dorongan energi cepat."},
    {"name": "Cappuccino", "desc": "Espresso dengan susu berbusa lembut, klasik dan seimbang."},
    {"name": "Caffè Latte", "desc": "Espresso dengan banyak susu hangat, halus dan creamy."},
    {"name": "Americano", "desc": "Espresso yang diencerkan dengan air panas, ringan dan bersih."},
    {"name": "Kopi Tubruk", "desc": "Kopi tradisional Indonesia, diseduh langsung dengan gula."},
    {"name": "Cold Brew", "desc": "Kopi seduh dingin 12 jam, rendah asam dan menyegarkan."},
    {"name": "Flat White", "desc": "Espresso ganda dengan microfoam tipis, kuat tapi lembut."},
    {"name": "Mocha", "desc": "Perpaduan espresso, cokelat, dan susu untuk yang suka manis."},
]

_CALMING_WORDS = [
    "Tarik napas dalam-dalam, semua akan baik-baik saja.",
    "Satu langkah kecil hari ini lebih baik daripada tidak sama sekali.",
    "Istirahat sejenak bukan berarti menyerah, tapi mengisi ulang tenaga.",
    "Kamu sudah melakukan yang terbaik, hargai dirimu.",
    "Pelan-pelan saja, tidak semua harus selesai hari ini.",
    "Nikmati kopimu, biarkan pikiran tenang sejenak.",
    "Tenang, badai pasti berlalu. Seduh kopi dulu.",
    "Fokus pada hal yang bisa kamu kendalikan, lepaskan sisanya.",
]


def _get_coffee_recommendation() -> dict:
    """Endpoint internal /api/coffee (mock): kembalikan 1 rekomendasi kopi +
    kata penenang acak dari config lokal.

    Return: {"coffee": {"name","desc"}, "calming_word": "..."}
    """
    import random
    return {
        "coffee": random.choice(_COFFEE_MENU),
        "calming_word": random.choice(_CALMING_WORDS),
    }


@restricted
async def coffee_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /coffee — rekomendasi kopi + kata penenang acak."""
    data = _get_coffee_recommendation()
    coffee = data["coffee"]
    await update.message.reply_text(
        f"☕ <b>Rekomendasi Kopi Hari Ini:</b>\n\n"
        f"<b>{html.escape(coffee['name'])}</b>\n"
        f"{html.escape(coffee['desc'])}\n\n"
        f"🧘 <i>{html.escape(data['calming_word'])}</i>",
        parse_mode=ParseMode.HTML,
    )


@restricted
async def testsearch_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /testsearch — diagnostik koneksi Web Search (Serper.dev)."""
    if not config.WEB_SEARCH_ENABLED:
        await update.message.reply_text(
            "ℹ️ Web Search sedang <b>dinonaktifkan</b> (WEB_SEARCH_ENABLED=false).",
            parse_mode=ParseMode.HTML,
        )
        return
    if not config.SERPER_API_KEY:
        await update.message.reply_text(
            "⚠️ <code>SERPER_API_KEY</code> belum diset di environment (Railway).",
            parse_mode=ParseMode.HTML,
        )
        return

    query = update.message.text.partition(" ")[2].strip() or "berita teknologi terbaru"
    await update.message.reply_text(f"🔎 Test search: <code>{html.escape(query)}</code>", parse_mode=ParseMode.HTML)
    try:
        results = await asyncio.to_thread(_web_search, query, 3)
        if not results:
            await update.message.reply_text("ℹ️ Koneksi OK tapi hasil kosong.")
            return
        await update.message.reply_text(
            _format_web_results(query, results),
            parse_mode=ParseMode.HTML,
            disable_web_page_preview=True,
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        await update.message.reply_text(
            f"⚠️ <b>Web Search gagal:</b>\n<pre>{html.escape(f'{type(e).__name__}: {e}')[:600]}</pre>\n"
            f"<pre>{html.escape(tb[-800:])}</pre>",
            parse_mode=ParseMode.HTML,
        )


# ==============================================================================
# GRABEXPRESS TRACKING (/lacakgrab, /trackgrab)
# ==============================================================================

def _format_grab_delivery(delivery_id: str, data: dict) -> str:
    """Susun pesan status pengiriman GrabExpress dari response API."""
    status = data.get("status", "-")

    courier = data.get("courier") or data.get("driver") or {}
    driver_name = courier.get("name", "") or "-"
    license_plate = courier.get("licensePlate") or courier.get("license_plate") or ""
    driver_phone = courier.get("phone") or courier.get("phoneNumber") or "-"

    tracking_url = data.get("trackingURL") or data.get("trackingUrl") or data.get("tracking_url") or ""

    driver_line = html.escape(driver_name)
    if license_plate:
        driver_line += f" ({html.escape(license_plate)})"

    lines = [
        "🛵 <b>Status Pengiriman GrabExpress</b>",
        "",
        f"<b>ID Pengiriman:</b> {html.escape(str(delivery_id))}",
        f"<b>Status:</b> {html.escape(str(status))}",
        f"<b>Kurir:</b> {driver_line}",
        f"<b>No. Telp Kurir:</b> {html.escape(str(driver_phone))}",
    ]
    if tracking_url:
        lines.append("")
        lines.append(f"📍 <i>Live Tracking Peta:</i> <a href=\"{html.escape(tracking_url)}\">Klik di sini untuk buka peta</a>")

    return "\n".join(lines)


@restricted
async def lacakgrab_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /lacakgrab & /trackgrab [deliveryID] — cek status pengiriman GrabExpress."""
    delivery_id = update.message.text.partition(" ")[2].strip()

    if not delivery_id:
        await update.message.reply_text(
            "<b>🛵 Lacak Pengiriman GrabExpress</b>\n\n"
            "Format:\n"
            "<code>/lacakgrab [deliveryID]</code>\n\n"
            "Contoh:\n"
            "<code>/lacakgrab DELIV-GRAB-982131</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    if not grab:
        await update.message.reply_text(
            "⚠️ Layanan GrabExpress belum dikonfigurasi (CLIENT_ID/SECRET belum diisi).",
            parse_mode=ParseMode.HTML,
        )
        return

    await update.message.reply_text(f"⏳ Melacak pengiriman <code>{html.escape(delivery_id)}</code>...", parse_mode=ParseMode.HTML)

    try:
        data = await asyncio.to_thread(grab.get_delivery, delivery_id)
    except GrabNotFound:
        await update.message.reply_text(
            f"Maaf, ID Pengiriman <code>{html.escape(delivery_id)}</code> tidak ditemukan di sistem GrabExpress.",
            parse_mode=ParseMode.HTML,
        )
        return
    except GrabError as e:
        logger.error(f"GrabExpress error saat lacak {delivery_id}: {e}")
        await update.message.reply_text(
            "Gagal terhubung ke layanan GrabExpress. Silakan coba beberapa saat lagi."
        )
        return
    except Exception as e:
        logger.exception(f"Error tak terduga saat lacak Grab {delivery_id}: {e}")
        await update.message.reply_text(
            "Gagal terhubung ke layanan GrabExpress. Silakan coba beberapa saat lagi."
        )
        return

    await update.message.reply_text(
        _format_grab_delivery(delivery_id, data),
        parse_mode=ParseMode.HTML,
        disable_web_page_preview=True,
    )


# ==============================================================================
# ERASPACE ORDER TRACKING (dumpdo) — /cekorder
# ==============================================================================

def _fetch_eraspace_order(pomp: str) -> dict:
    """POST ke endpoint dumpdo Eraspace. Return dict data order pertama.

    Raise ValueError kalau order tidak ditemukan, RuntimeError untuk kendala API.
    """
    params = {"user": config.ERASPACE_USER, "pomp": pomp}
    headers = {"Accept": "application/json"}
    if config.ERASPACE_COOKIE:
        headers["Cookie"] = config.ERASPACE_COOKIE

    try:
        resp = requests.post(config.ERASPACE_DUMPDO_URL, params=params, headers=headers, timeout=20)
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Gagal terhubung: {e}")

    if not resp.ok:
        raise RuntimeError(f"HTTP {resp.status_code}: {resp.text[:200]}")

    try:
        data = resp.json()
    except ValueError:
        raise RuntimeError("Respon bukan JSON valid.")

    # Cari objek order di dalam response (bisa dict langsung, atau di dalam
    # 'data'/'result'/list). Ambil yang mengandung 'orderNumber'.
    def _find_order(obj):
        if isinstance(obj, dict):
            if "orderNumber" in obj:
                return obj
            for v in obj.values():
                found = _find_order(v)
                if found:
                    return found
        elif isinstance(obj, list):
            for item in obj:
                found = _find_order(item)
                if found:
                    return found
        return None

    order = _find_order(data)
    if not order:
        raise ValueError("Order tidak ditemukan")
    return order


def _format_eraspace_order(order: dict) -> str:
    """Susun pesan status order Eraspace dari field yang diminta."""
    order_number = order.get("orderNumber", "-")
    order_status = order.get("orderstatus") or order.get("orderStatus") or "-"
    courier_code = order.get("courierCode", "") or "-"
    courier_desc = order.get("courierDescription", "") or "-"
    tracking_no = order.get("trackingNo", "") or "-"
    lat_long = order.get("latLong", "") or "-"
    customer_phone = order.get("customerPhone", "") or "-"

    lines = [
        "📦 <b>Status Order Eraspace</b>",
        "",
        f"<b>Order Number:</b> {html.escape(str(order_number))}",
        f"<b>Status:</b> {html.escape(str(order_status))}",
        f"<b>Kurir:</b> {html.escape(str(courier_code))} - {html.escape(str(courier_desc))}",
        f"<b>No. Resi:</b> {html.escape(str(tracking_no))}",
        f"<b>No. Telp Customer:</b> {html.escape(str(customer_phone))}",
    ]

    # Kalau latLong ada, buat link ke Google Maps
    if lat_long and lat_long != "-":
        maps_url = f"https://www.google.com/maps?q={html.escape(str(lat_long))}"
        lines.append(f"📍 <b>Lokasi:</b> <a href=\"{maps_url}\">{html.escape(str(lat_long))}</a>")
    else:
        lines.append(f"📍 <b>Lokasi:</b> -")

    return "\n".join(lines)


@restricted
async def cekorder_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /cekorder [pomp] — cek status order & tracking dari Eraspace dumpdo API."""
    pomp = update.message.text.partition(" ")[2].strip()

    if not pomp:
        await update.message.reply_text(
            "<b>📦 Cek Status Order Eraspace</b>\n\n"
            "Format:\n"
            "<code>/cekorder [nomor pomp]</code>\n\n"
            "Contoh:\n"
            "<code>/cekorder 4601361781</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    await update.message.reply_text(f"⏳ Mengecek order <code>{html.escape(pomp)}</code>...", parse_mode=ParseMode.HTML)

    try:
        order = await asyncio.to_thread(_fetch_eraspace_order, pomp)
    except ValueError:
        await update.message.reply_text(
            f"Maaf, order <code>{html.escape(pomp)}</code> tidak ditemukan di sistem Eraspace.",
            parse_mode=ParseMode.HTML,
        )
        return
    except Exception as e:
        logger.error(f"Eraspace dumpdo error untuk {pomp}: {e}")
        await update.message.reply_text(
            "Gagal terhubung ke layanan Eraspace. Silakan coba beberapa saat lagi."
        )
        return

    await update.message.reply_text(
        _format_eraspace_order(order),
        parse_mode=ParseMode.HTML,
        disable_web_page_preview=True,
    )


async def check_open_ticket_reminders(context: ContextTypes.DEFAULT_TYPE):
    if not sdp or not config.SDP_NOTIFY_GROUPS:
        return

    state = _load_sdp_notify_state()
    interval = state.get("open_reminder_interval_minutes", config.SDP_OPEN_REMINDER_DEFAULT_MINUTES)
    if not interval or interval <= 0:
        return

    now = dt.datetime.now(TZ)
    last_at_str = state.get("last_open_reminder_at")
    if last_at_str:
        try:
            last_at = dt.datetime.fromisoformat(last_at_str)
        except ValueError:
            last_at = None
    else:
        last_at = None

    if last_at and (now - last_at) < dt.timedelta(minutes=interval):
        return

    try:
        open_tickets = await asyncio.to_thread(
            sdp.list_requests, 50, "Open", config.SDP_NOTIFY_GROUPS
        )
    except SDPError as e:
        logger.error(f"Gagal cek tiket Open untuk reminder: {e}")
        return

    state["last_open_reminder_at"] = now.isoformat()
    _save_sdp_notify_state(state)

    if not open_tickets:
        return

    lines = [f"⏰ Reminder: masih ada {len(open_tickets)} tiket berstatus Open:", ""]
    for req in open_tickets[:15]:
        lines.append(_format_sdp_request_line(req))
        lines.append("")
    if len(open_tickets) > 15:
        lines.append(f"...dan {len(open_tickets) - 15} tiket Open lainnya.")
    await _broadcast_notify(context, _safe_telegram_text("\n".join(lines).rstrip()))


async def _broadcast_notify(context: ContextTypes.DEFAULT_TYPE, text: str):
    for chat_id, thread_id in config.notify_targets():
        try:
            await context.bot.send_message(
                chat_id=chat_id, 
                text=text, 
                message_thread_id=thread_id, 
                parse_mode=ParseMode.HTML
            )
        except Exception:
            logger.exception(f"Gagal kirim notifikasi ke chat_id={chat_id} thread_id={thread_id}")


async def daily_reminder(context: ContextTypes.DEFAULT_TYPE):
    await _broadcast_notify(context, "Reminder: Jangan lupa isi logwork")


async def interval_reminder(context: ContextTypes.DEFAULT_TYPE):
    now = dt.datetime.now(TZ)
    if not (config.REMINDER_START_HOUR <= now.hour < config.REMINDER_END_HOUR):
        return
    await _broadcast_notify(context, "Reminder: jangan lupa isi logwork")


# ==============================================================================
# SLA MONITOR (PLAYWRIGHT DASHBOARD SCRAPING)
# ==============================================================================
SLA_TEAM = [
    "Adelia Pebriani",
    "Bagus Dwi Susworo",
    "Bambang Purnomo Sidi",
    "Prizky Stefajar Darmaliz Gagah Utomo",
    "Rendi Surya Hadinata",
    "Teguh Wiguna",
    "Tri Sutrisno",
    "Unassigned",
]


def _run_sla_scrape() -> dict:
    """Ambil data tiket dari SDP API per status untuk group yang dikonfigurasi."""
    from collections import Counter

    if not sdp:
        raise RuntimeError("SDP Client tidak aktif (SDP_BASE_URL/SDP_API_KEY belum diisi).")

    groups = config.SDP_NOTIFY_GROUPS
    if not groups:
        raise RuntimeError("SDP_NOTIFY_GROUPS kosong.")

    statuses_to_check = [
        "Open",
        "In Progress Investigation",
        "Transfer L1",
        "Waiting User Confirmation",
        "Pending",
        "Onhold",
    ]

    status_counts = {}
    all_tickets = []

    # Ambil tiket per status
    for status in statuses_to_check:
        tickets = sdp.list_requests(100, status, groups)
        status_counts[status] = len(tickets)
        all_tickets.extend(tickets)

    # Hitung Unassigned = tiket aktif yang TIDAK punya technician
    unassigned_count = 0
    for t in all_tickets:
        tech = t.get("technician")
        if not tech or not tech.get("name"):
            unassigned_count += 1
    status_counts["Unassigned"] = unassigned_count

    # Ambil tiket Closed
    closed_tickets = sdp.list_requests(100, "Closed", groups)
    status_counts["Closed"] = len(closed_tickets)

    # Hitung Solved dalam 1 minggu (tiket Closed yang completed_time dalam 7 hari terakhir)
    now = dt.datetime.now(TZ)
    week_ago = now - dt.timedelta(days=7)
    solved_this_week = 0
    for t in closed_tickets:
        completed_time = t.get("completed_time") or t.get("resolved_time")
        if completed_time:
            # SDP API biasanya return epoch ms
            try:
                if isinstance(completed_time, dict):
                    epoch = int(completed_time.get("value", 0)) / 1000
                else:
                    epoch = int(completed_time) / 1000
                ticket_dt = dt.datetime.fromtimestamp(epoch, tz=TZ)
                if ticket_dt >= week_ago:
                    solved_this_week += 1
            except (ValueError, TypeError, OSError):
                pass
    # Jika tidak bisa cek waktu (API tidak return waktu), hitung semua closed sebagai solved minggu ini
    if solved_this_week == 0 and len(closed_tickets) > 0:
        solved_this_week = len(closed_tickets)

    # Hitung OverDue = tiket aktif yang sudah lewat DueBy (is_overdue / due_by_time)
    now_epoch = now.timestamp()
    overdue_count = 0
    for t in all_tickets:
        due_epoch = _epoch_from_sdp_time(t.get("due_by_time"))
        if bool(t.get("is_overdue")) or (due_epoch > 0 and due_epoch < now_epoch):
            overdue_count += 1
    status_counts["OverDue"] = overdue_count

    # Parse assignee dari semua tiket aktif (hanya yang punya technician)
    technicians = []
    for t in all_tickets:
        tech = t.get("technician")
        if tech and tech.get("name"):
            technicians.append(tech["name"])
    counter = Counter(technicians)

    assignee_info = []
    for member in SLA_TEAM:
        if member == "Unassigned":
            continue
        count = counter.get(member, 0)
        if count > 0:
            assignee_info.append(f"  • {member} : {count}")

    total_active = sum(v for k, v in status_counts.items() if k not in ("Closed", "Unassigned"))

    return {
        "status_counts": status_counts,
        "total_active": total_active,
        "assignee_info": assignee_info,
        "solved_this_week": solved_this_week,
        "overdue_count": overdue_count,
    }


async def sla_monitor_job(context: ContextTypes.DEFAULT_TYPE):
    """Background job: ambil status tiket dari SDP API dan kirim laporan ke Telegram."""
    try:
        result = await asyncio.to_thread(_run_sla_scrape)
    except Exception as e:
        logger.error(f"SLA Monitor error: {e}")
        await _broadcast_notify(context, f"❌ <b>SLA Monitor ERROR</b>\n<code>{html.escape(str(e))}</code>")
        return

    sc = result["status_counts"]
    now_str = dt.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

    # Format dengan monospace agar rata
    status_lines = [
        f"Open                      : {sc.get('Open', 0):>3}",
        f"In Progress Investigation : {sc.get('In Progress Investigation', 0):>3}",
        f"Transfer L1               : {sc.get('Transfer L1', 0):>3}",
        f"Waiting User Confirmation : {sc.get('Waiting User Confirmation', 0):>3}",
        f"Pending                   : {sc.get('Pending', 0):>3}",
        f"Unassigned                : {sc.get('Unassigned', 0):>3}",
        f"OverDue                   : {sc.get('OverDue', 0):>3}",
        f"Onhold                    : {sc.get('Onhold', 0):>3}",
        f"Closed                    : {sc.get('Closed', 0):>3}",
    ]

    summary_lines = [
        f"Total Active Ticket       : {result['total_active']:>3}",
        f"Solved (7 hari)           : {result['solved_this_week']:>3}",
        f"Total OverDue             : {result['overdue_count']:>3}",
    ]

    lines = [
        "🚨 <b>Summary Ticket Status</b>",
        "━━━━━━━━━━━━━━━━━━━━━━",
        "",
        "<pre>" + "\n".join(status_lines) + "</pre>",
        "",
        "<pre>" + "\n".join(summary_lines) + "</pre>",
    ]

    if result["assignee_info"]:
        lines.append("")
        lines.append("<b>📋 Assigned To:</b>")
        lines.extend(result["assignee_info"])

    lines.append("")
    lines.append(f"🕐 {now_str}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━")

    await _broadcast_notify(context, "\n".join(lines))
    logger.info(f"SLA Monitor: berhasil kirim laporan ({result['total_active']} tiket aktif)")


@restricted
async def ticketupdate_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /ticketupdate — trigger manual notifikasi Summary Ticket Status."""
    await update.message.reply_text("⏳ Mengambil data tiket dari SDP API...")
    try:
        result = await asyncio.to_thread(_run_sla_scrape)
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ <b>Gagal ambil data tiket:</b>\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    sc = result["status_counts"]
    now_str = dt.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")

    status_lines = [
        f"Open                      : {sc.get('Open', 0):>3}",
        f"In Progress Investigation : {sc.get('In Progress Investigation', 0):>3}",
        f"Transfer L1               : {sc.get('Transfer L1', 0):>3}",
        f"Waiting User Confirmation : {sc.get('Waiting User Confirmation', 0):>3}",
        f"Pending                   : {sc.get('Pending', 0):>3}",
        f"Unassigned                : {sc.get('Unassigned', 0):>3}",
        f"OverDue                   : {sc.get('OverDue', 0):>3}",
        f"Onhold                    : {sc.get('Onhold', 0):>3}",
        f"Closed                    : {sc.get('Closed', 0):>3}",
    ]

    summary_lines = [
        f"Total Active Ticket       : {result['total_active']:>3}",
        f"Solved (7 hari)           : {result['solved_this_week']:>3}",
        f"Total OverDue             : {result['overdue_count']:>3}",
    ]

    lines = [
        "🚨 <b>Summary Ticket Status</b>",
        "━━━━━━━━━━━━━━━━━━━━━━",
        "",
        "<pre>" + "\n".join(status_lines) + "</pre>",
        "",
        "<pre>" + "\n".join(summary_lines) + "</pre>",
    ]

    if result["assignee_info"]:
        lines.append("")
        lines.append("<b>📋 Assigned To:</b>")
        lines.extend(result["assignee_info"])

    lines.append("")
    lines.append(f"🕐 {now_str}")
    lines.append("━━━━━━━━━━━━━━━━━━━━━━")

    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


# ==============================================================================
# DELETE RESERVATION VIA TELEGRAM
# ==============================================================================
DELRES_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "deletereservation", "cancel_reservation.ps1")
DELRES_CSV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "deletereservation")


def _parse_delres_text(text: str) -> list:
    """
    Parse input teks user menjadi list item reservasi.
    Format yang didukung:
      bucode: E370
      ordernumber: 8302562258
      sku: 8100258103
      qty: 2

    Atau multi-item dipisah baris kosong.
    """
    items = []
    current = {}

    for line in text.splitlines():
        line = line.strip()
        if not line:
            if current:
                items.append(current)
                current = {}
            continue

        # Parse key: value
        match = re.match(r"^(bucode|ordernumber|sku|qty)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            key = match.group(1).lower()
            val = match.group(2).strip()
            if key == "bucode":
                current["businessUnitCode"] = val
            elif key == "ordernumber":
                current["transactionNumber"] = val
            elif key == "sku":
                if "itemCodes" not in current:
                    current["itemCodes"] = []
                current["itemCodes"].append(val)
            elif key == "qty":
                current["qty"] = int(val) if val.isdigit() else 1

    if current:
        items.append(current)

    # Validasi minimal
    valid_items = []
    for item in items:
        if item.get("businessUnitCode") and item.get("transactionNumber") and item.get("itemCodes"):
            item.setdefault("qty", 1)
            valid_items.append(item)

    return valid_items


def _build_delres_csv(items: list) -> str:
    """Buat file CSV sementara dari list item reservasi, return path file."""
    timestamp = dt.datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(DELRES_CSV_DIR, f"delres_temp_{timestamp}.csv")

    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("businessUnitCode,transactionNumber,itemCode,qty,orderDate\n")
        for item in items:
            bu = item["businessUnitCode"]
            trx = item["transactionNumber"]
            qty = item.get("qty", 1)
            for sku in item["itemCodes"]:
                # orderDate pakai hari ini sebagai default
                order_date = dt.datetime.now(TZ).strftime("%Y-%m-%d 00:00:00")
                f.write(f"{bu},{trx},{sku},{qty},{order_date}\n")

    return csv_path


@restricted
async def delreservation_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /delreservation — mulai proses delete reservation."""
    context.user_data.clear()
    await update.message.reply_text(
        "<b>🗑 Delete Reservation</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Upload file CSV</b>\n"
        "Format kolom: businessUnitCode, transactionNumber, itemCode, qty, orderDate\n\n"
        "<b>2. Ketik langsung:</b>\n"
        "<code>bucode: E370\n"
        "ordernumber: 8302562258\n"
        "sku: 8100258103\n"
        "qty: 1</code>\n\n"
        "Untuk multiple item, pisahkan dengan baris kosong.\n"
        "Kirim sekarang (file atau teks), atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return DELRES_INPUT


async def delreservation_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user: bisa file CSV atau teks."""
    if update.message.document:
        # User upload file CSV
        doc = update.message.document
        if not doc.file_name.lower().endswith(".csv"):
            await update.message.reply_text("File harus berformat .csv. Coba upload ulang atau ketik manual.")
            return DELRES_INPUT

        file = await doc.get_file()
        csv_path = os.path.join(DELRES_CSV_DIR, f"delres_upload_{dt.datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.csv")
        await file.download_to_drive(csv_path)

        context.user_data["delres_csv"] = csv_path
        context.user_data["delres_source"] = "file"

        # Baca isi untuk preview
        with open(csv_path, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        data_count = len(lines) - 1  # minus header

        await update.message.reply_text(
            f"📄 File diterima: <b>{doc.file_name}</b>\n"
            f"Jumlah data: <b>{data_count} baris</b>\n\n"
            "Eksekusi delete reservation sekarang?",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Eksekusi", callback_data="delres_yes"),
                    InlineKeyboardButton("❌ Batal", callback_data="delres_no"),
                ]
            ]),
        )
        return DELRES_CONFIRM

    elif update.message.text:
        # User ketik teks manual
        text = update.message.text.strip()
        items = _parse_delres_text(text)

        if not items:
            await update.message.reply_text(
                "Format tidak dikenali. Pastikan format seperti ini:\n\n"
                "<code>bucode: E370\n"
                "ordernumber: 8302562258\n"
                "sku: 8100258103\n"
                "qty: 1</code>\n\n"
                "Coba lagi atau /cancel untuk batal.",
                parse_mode=ParseMode.HTML,
            )
            return DELRES_INPUT

        csv_path = _build_delres_csv(items)
        context.user_data["delres_csv"] = csv_path
        context.user_data["delres_source"] = "text"
        context.user_data["delres_items"] = items

        # Preview
        preview_lines = []
        for i, item in enumerate(items, 1):
            skus = ", ".join(item["itemCodes"])
            preview_lines.append(
                f"{i}. {item['transactionNumber']} | BU: {item['businessUnitCode']} | "
                f"SKU: {skus} | Qty: {item.get('qty', 1)}"
            )

        await update.message.reply_text(
            f"<b>🗑 Konfirmasi Delete Reservation</b>\n\n"
            + "\n".join(preview_lines)
            + "\n\nEksekusi sekarang?",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Eksekusi", callback_data="delres_yes"),
                    InlineKeyboardButton("❌ Batal", callback_data="delres_no"),
                ]
            ]),
        )
        return DELRES_CONFIRM

    await update.message.reply_text("Kirim file CSV atau ketik data reservasi. /cancel untuk batal.")
    return DELRES_INPUT


async def delreservation_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Konfirmasi dan eksekusi delete reservation via Python requests."""
    query = update.callback_query
    await query.answer()

    if query.data == "delres_no":
        csv_path = context.user_data.get("delres_csv")
        if csv_path and os.path.exists(csv_path):
            try:
                os.remove(csv_path)
            except OSError:
                pass
        await query.edit_message_text("❌ Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    csv_path = context.user_data.get("delres_csv")
    if not csv_path or not os.path.exists(csv_path):
        await query.edit_message_text("⚠️ File CSV tidak ditemukan. Coba ulang /delreservation.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text("⏳ Menjalankan delete reservation, mohon tunggu...")

    try:
        output = await asyncio.to_thread(_execute_delreservation, csv_path)

        if len(output) > 3500:
            output = output[:3500] + "\n... (output dipotong)"

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"✅ <b>Delete Reservation Selesai</b>\n\n<pre>{html.escape(output)}</pre>",
            parse_mode=ParseMode.HTML,
            message_thread_id=_thread_id_from_update(update),
        )

    except Exception as e:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"⚠️ Gagal menjalankan delete reservation:\n{html.escape(str(e))}",
            message_thread_id=_thread_id_from_update(update),
        )

    # Cleanup temp file
    if context.user_data.get("delres_source") == "text" and csv_path and os.path.exists(csv_path):
        try:
            os.remove(csv_path)
        except OSError:
            pass

    context.user_data.clear()
    return ConversationHandler.END


def _execute_delreservation(csv_path: str) -> str:
    """Eksekusi cancel reservation dari CSV via Python requests."""
    import requests as req_lib
    import csv
    import time as _time

    API_URL = "https://erpi.eraspace.com/api/v1/pos/hub/main"
    CLIENT_ID = "ERAFONEDOTCOM"
    CLIENT_SIGNATURE = "6f19a056e10b367dc354cbec47de2fd93fad1c25f505517bf8473529e2959073"

    # Baca CSV
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return "File CSV kosong."

        # Detect kolom qty
        header_lower = [h.strip().lower() for h in header]
        has_qty = "qty" in header_lower

        rows = []
        for row in reader:
            if not row or len(row) < 4:
                continue
            fields = [c.strip().strip('"') for c in row]
            last_idx = len(fields) - 1

            if has_qty:
                qty_idx = last_idx - 1
                qty_raw = fields[qty_idx]
                qty = int(qty_raw) if qty_raw.isdigit() else 1
                item_codes = [c for c in fields[2:qty_idx] if c]
            else:
                qty = 1
                item_codes = [c for c in fields[2:last_idx] if c]

            rows.append({
                "businessUnitCode": fields[0],
                "transactionNumber": fields[1],
                "itemCodes": item_codes,
                "qty": qty,
                "orderDate": fields[last_idx],
            })

    if not rows:
        return "Tidak ada data valid di CSV."

    # Epoch helper
    from datetime import datetime as _dt
    epoch_origin = _dt(1970, 1, 1)

    def to_epoch(date_str):
        try:
            parsed = _dt.strptime(date_str.strip(), "%Y-%m-%d %H:%M:%S")
        except ValueError:
            try:
                parsed = _dt.strptime(date_str.strip(), "%Y-%m-%d")
            except ValueError:
                return "0"
        return str(int((parsed - epoch_origin).total_seconds()))

    total = 0
    success_count = 0
    failed_count = 0
    output_lines = []

    for row in rows:
        trx = row["transactionNumber"]
        if not trx:
            continue

        total += 1
        bu_code = row["businessUnitCode"]
        order_epoch = to_epoch(row["orderDate"])

        child_procedures = []
        for item_code in row["itemCodes"]:
            child_procedures.append({
                "procedureCode": "O2O - CANCEL ORDER ITEM",
                "parametersIn": {
                    "transactionNumber": trx,
                    "itemCode": item_code,
                    "quantity": row["qty"],
                    "price": "0"
                }
            })

        if not child_procedures:
            continue

        now_utc = _dt.utcnow()
        timestamps = str(int((now_utc - epoch_origin).total_seconds()))
        cancel_date = now_utc.strftime("%Y%m%d%H%M%S")

        body = {
            "procedureCode": "O2O - CANCEL ORDER",
            "parametersIn": {
                "clientId": CLIENT_ID,
                "timestamps": timestamps,
                "businessUnitCode": bu_code,
                "clientSignature": CLIENT_SIGNATURE,
                "transactionNumber": trx,
                "orderDate": order_epoch,
                "cancelDate": cancel_date,
                "orderStatus": "cancel"
            },
            "childProcedure": child_procedures
        }

        try:
            resp = req_lib.post(API_URL, json=body, headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}, timeout=30)
            resp_text = resp.text

            if resp.status_code == 200:
                # Cek logical error — response yang sukses biasanya mengandung "Successfully"
                resp_lower = resp_text.lower()
                if "successfully" in resp_lower or '"outerror":0' in resp_lower.replace(" ", ""):
                    success_count += 1
                    output_lines.append(f"[{total}] {trx} -> Sukses")
                elif "not exists" in resp_lower or "not found" in resp_lower or "invalid" in resp_lower or "gagal" in resp_lower:
                    failed_count += 1
                    output_lines.append(f"[{total}] {trx} -> GAGAL (logika): {resp_text[:100]}")
                else:
                    success_count += 1
                    output_lines.append(f"[{total}] {trx} -> Sukses (HTTP 200)")
            else:
                failed_count += 1
                output_lines.append(f"[{total}] {trx} -> GAGAL (HTTP {resp.status_code}): {resp_text[:100]}")
        except Exception as e:
            failed_count += 1
            output_lines.append(f"[{total}] {trx} -> ERROR: {str(e)[:80]}")

        _time.sleep(0.2)

    output_lines.append("")
    output_lines.append(f"===== SELESAI =====")
    output_lines.append(f"Total  : {total}")
    output_lines.append(f"Sukses : {success_count}")
    output_lines.append(f"Gagal  : {failed_count}")

    # Simpan log ke file
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "deletereservation", "logs")
    os.makedirs(log_dir, exist_ok=True)
    from datetime import datetime as _dt2
    log_file = os.path.join(log_dir, f"cancel_log_{_dt2.now().strftime('%Y%m%d_%H%M%S')}.txt")
    with open(log_file, "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))

    return "\n".join(output_lines)


# ==============================================================================
# RELEASE VOUCHER VIA TELEGRAM
# ==============================================================================
RELVOUCHER_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "promovoucher", "release_voucher.ps1")
RELVOUCHER_CSV_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "promovoucher")


def _parse_relvoucher_text(text: str) -> list:
    """
    Parse input teks user menjadi list promo ID.
    Format yang didukung:
      - Satu ID per baris: 462016
      - Dengan label: id: 462016
      - Dipisah koma: 462016, 462017, 462018
    """
    ids = []

    # Coba split koma dulu
    if "," in text:
        for part in text.split(","):
            part = part.strip()
            # Hapus label kalau ada
            match = re.match(r"^(?:id\s*[:=]\s*)?(\d+)", part, re.IGNORECASE)
            if match:
                ids.append(match.group(1))
    else:
        for line in text.splitlines():
            line = line.strip()
            if not line:
                continue
            match = re.match(r"^(?:id\s*[:=]\s*)?(\d+)", line, re.IGNORECASE)
            if match:
                ids.append(match.group(1))

    return ids


def _build_relvoucher_csv(ids: list) -> str:
    """Buat file CSV sementara dari list promo ID, return path file."""
    timestamp = dt.datetime.now(TZ).strftime("%Y%m%d_%H%M%S")
    csv_path = os.path.join(RELVOUCHER_CSV_DIR, f"relvoucher_temp_{timestamp}.csv")

    with open(csv_path, "w", encoding="utf-8") as f:
        f.write("id\n")
        for promo_id in ids:
            f.write(f"{promo_id}\n")

    return csv_path


@restricted
async def releasevoucher_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /releasevoucher — mulai proses release voucher."""
    context.user_data.clear()
    await update.message.reply_text(
        "<b>🎟 Release Voucher / Promo</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Upload file CSV</b>\n"
        "Format kolom: id\n\n"
        "<b>2. Ketik langsung (satu atau lebih ID):</b>\n"
        "<code>462016\n"
        "462017\n"
        "462018</code>\n\n"
        "Atau pisahkan dengan koma: <code>462016, 462017, 462018</code>\n\n"
        "Kirim sekarang (file atau teks), atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return RELVOUCHER_INPUT


async def releasevoucher_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user: bisa file CSV atau teks."""
    if update.message.document:
        # User upload file CSV
        doc = update.message.document
        if not doc.file_name.lower().endswith(".csv"):
            await update.message.reply_text("File harus berformat .csv. Coba upload ulang atau ketik manual.")
            return RELVOUCHER_INPUT

        file = await doc.get_file()
        csv_path = os.path.join(RELVOUCHER_CSV_DIR, f"relvoucher_upload_{dt.datetime.now(TZ).strftime('%Y%m%d_%H%M%S')}.csv")
        await file.download_to_drive(csv_path)

        context.user_data["relvoucher_csv"] = csv_path
        context.user_data["relvoucher_source"] = "file"

        # Baca isi untuk preview
        with open(csv_path, "r", encoding="utf-8") as f:
            lines = [l.strip() for l in f.readlines() if l.strip()]
        data_count = len(lines) - 1  # minus header

        await update.message.reply_text(
            f"📄 File diterima: <b>{doc.file_name}</b>\n"
            f"Jumlah promo ID: <b>{data_count}</b>\n\n"
            "Eksekusi release voucher sekarang?",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Eksekusi", callback_data="relvoucher_yes"),
                    InlineKeyboardButton("❌ Batal", callback_data="relvoucher_no"),
                ]
            ]),
        )
        return RELVOUCHER_CONFIRM

    elif update.message.text:
        # User ketik teks manual
        text = update.message.text.strip()
        ids = _parse_relvoucher_text(text)

        if not ids:
            await update.message.reply_text(
                "Tidak ada ID yang dikenali. Pastikan format seperti ini:\n\n"
                "<code>462016\n462017\n462018</code>\n\n"
                "Atau: <code>462016, 462017, 462018</code>\n\n"
                "Coba lagi atau /cancel untuk batal.",
                parse_mode=ParseMode.HTML,
            )
            return RELVOUCHER_INPUT

        csv_path = _build_relvoucher_csv(ids)
        context.user_data["relvoucher_csv"] = csv_path
        context.user_data["relvoucher_source"] = "text"
        context.user_data["relvoucher_ids"] = ids

        # Preview
        preview = ", ".join(ids[:20])
        if len(ids) > 20:
            preview += f" ... (+{len(ids) - 20} lainnya)"

        await update.message.reply_text(
            f"<b>🎟 Konfirmasi Release Voucher</b>\n\n"
            f"Jumlah ID: <b>{len(ids)}</b>\n"
            f"ID: <code>{preview}</code>\n\n"
            "Eksekusi sekarang?",
            parse_mode=ParseMode.HTML,
            reply_markup=InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("✅ Eksekusi", callback_data="relvoucher_yes"),
                    InlineKeyboardButton("❌ Batal", callback_data="relvoucher_no"),
                ]
            ]),
        )
        return RELVOUCHER_CONFIRM

    await update.message.reply_text("Kirim file CSV atau ketik promo ID. /cancel untuk batal.")
    return RELVOUCHER_INPUT


async def releasevoucher_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Konfirmasi dan eksekusi release voucher via Python requests."""
    query = update.callback_query
    await query.answer()

    if query.data == "relvoucher_no":
        csv_path = context.user_data.get("relvoucher_csv")
        if csv_path and os.path.exists(csv_path):
            try:
                os.remove(csv_path)
            except OSError:
                pass
        await query.edit_message_text("❌ Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    csv_path = context.user_data.get("relvoucher_csv")
    if not csv_path or not os.path.exists(csv_path):
        await query.edit_message_text("⚠️ File CSV tidak ditemukan. Coba ulang /releasevoucher.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text("⏳ Menjalankan release voucher, mohon tunggu...")

    try:
        output = await asyncio.to_thread(_execute_releasevoucher, csv_path)

        if len(output) > 3500:
            output = output[:3500] + "\n... (output dipotong)"

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"✅ <b>Release Voucher Selesai</b>\n\n<pre>{html.escape(output)}</pre>",
            parse_mode=ParseMode.HTML,
            message_thread_id=_thread_id_from_update(update),
        )

    except Exception as e:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"⚠️ Gagal menjalankan release voucher:\n{html.escape(str(e))}",
            message_thread_id=_thread_id_from_update(update),
        )

    # Cleanup temp file
    if context.user_data.get("relvoucher_source") == "text" and csv_path and os.path.exists(csv_path):
        try:
            os.remove(csv_path)
        except OSError:
            pass

    context.user_data.clear()
    return ConversationHandler.END


def _execute_releasevoucher(csv_path: str) -> str:
    """Eksekusi release voucher dari CSV via Python requests."""
    import requests as req_lib
    import time as _time

    API_BASE = "https://sculptor.eraspace.com/promos/v1/promo/release-by-system"
    AUTH_HEADER = "e9738b21b981a6f33d096f51830fac27"

    # Baca CSV — kolom pertama berisi ID
    with open(csv_path, "r", encoding="utf-8") as f:
        lines = [l.strip() for l in f.readlines() if l.strip()]

    if len(lines) < 2:
        return "File CSV kosong (hanya header atau kosong)."

    # Skip header, ambil kolom pertama
    ids = []
    for line in lines[1:]:
        parts = line.split(",")
        val = parts[0].strip().strip('"')
        if val and val.isdigit():
            ids.append(val)

    if not ids:
        return "Tidak ada ID valid di CSV."

    total = 0
    success_count = 0
    failed_count = 0
    output_lines = []

    for promo_id in ids:
        total += 1
        url = f"{API_BASE}/{promo_id}"

        try:
            resp = req_lib.post(
                url,
                headers={"X-Auth-Signature": AUTH_HEADER, "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"},
                timeout=20,
            )

            if resp.status_code == 200:
                success_count += 1
                output_lines.append(f"[{total}] ID {promo_id} -> Sukses (HTTP 200)")
            else:
                failed_count += 1
                body = resp.text[:100]
                output_lines.append(f"[{total}] ID {promo_id} -> GAGAL (HTTP {resp.status_code}): {body}")
        except Exception as e:
            failed_count += 1
            output_lines.append(f"[{total}] ID {promo_id} -> ERROR: {str(e)[:80]}")

        _time.sleep(0.2)

    output_lines.append("")
    output_lines.append(f"===== SELESAI =====")
    output_lines.append(f"Total  : {total}")
    output_lines.append(f"Sukses : {success_count}")
    output_lines.append(f"Gagal  : {failed_count}")

    # Simpan log ke file
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "promovoucher", "logs")
    os.makedirs(log_dir, exist_ok=True)
    from datetime import datetime as _dt2
    log_file = os.path.join(log_dir, f"release_log_{_dt2.now().strftime('%Y%m%d_%H%M%S')}.txt")
    with open(log_file, "w", encoding="utf-8") as f:
        f.write("\n".join(output_lines))

    return "\n".join(output_lines)


# ==============================================================================
# CHECK STOCK VIA TELEGRAM
# ==============================================================================
CHECKSTOCK_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "checkstock", "oaa_stock_query.ps1")


def _parse_checkstock_text(text: str) -> dict:
    """
    Parse input teks user menjadi parameter check stock.
    Format yang didukung:
      sku: 8000044321
      source: SS20

    Atau multiple:
      sku: 8000044321, 8000044322
      source: SS20, SS21
    """
    result = {"article_ids": [], "source_ids": [], "units": []}

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        match = re.match(r"^(sku|article|articleid|article_id)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            values = [v.strip() for v in match.group(2).split(",") if v.strip()]
            result["article_ids"].extend(values)
            continue

        match = re.match(r"^(source|sourceid|source_id|toko|store)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            values = [v.strip() for v in match.group(2).split(",") if v.strip()]
            result["source_ids"].extend(values)
            continue

        match = re.match(r"^(unit)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            values = [v.strip() for v in match.group(2).split(",") if v.strip()]
            result["units"].extend(values)
            continue

    return result if result["article_ids"] and result["source_ids"] else None


@restricted
async def checkstock_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /stock — mulai proses cek stock."""
    context.user_data.clear()

    # Cek apakah user langsung kirim parameter setelah command
    parts = update.message.text.split(maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        # Coba parse langsung dari command, misal: /stock 8000044321 SS20
        args = parts[1].strip().split()
        if len(args) >= 2:
            context.user_data["checkstock_data"] = {
                "article_ids": [args[0]],
                "source_ids": [args[1]],
                "units": [args[2]] if len(args) > 2 else ["EA"],
            }
            return await _checkstock_show_confirm(update, context)

    await update.message.reply_text(
        "<b>📦 Check Stock (OAA)</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Ketik langsung:</b>\n"
        "<code>sku: 8000044321\n"
        "source: SS20</code>\n\n"
        "<b>2. Multiple SKU/Source:</b>\n"
        "<code>sku: 8000044321, 8000044322\n"
        "source: SS20, SS21</code>\n\n"
        "<b>3. Shortcut:</b>\n"
        "<code>/stock 8000044321 SS20</code>\n\n"
        "Opsional tambahkan unit (default: EA):\n"
        "<code>unit: PC</code>\n\n"
        "Kirim sekarang, atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return CHECKSTOCK_INPUT


async def checkstock_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user untuk check stock."""
    text = update.message.text.strip()
    data = _parse_checkstock_text(text)

    if not data:
        await update.message.reply_text(
            "Format tidak dikenali. Pastikan minimal ada <b>sku</b> dan <b>source</b>:\n\n"
            "<code>sku: 8000044321\n"
            "source: SS20</code>\n\n"
            "Coba lagi atau /cancel untuk batal.",
            parse_mode=ParseMode.HTML,
        )
        return CHECKSTOCK_INPUT

    context.user_data["checkstock_data"] = data
    return await _checkstock_show_confirm(update, context)


async def _checkstock_show_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Tampilkan preview dan minta konfirmasi."""
    data = context.user_data["checkstock_data"]
    skus = ", ".join(data["article_ids"])
    sources = ", ".join(data["source_ids"])
    units = ", ".join(data["units"]) if data["units"] else "EA"

    await update.message.reply_text(
        f"<b>📦 Konfirmasi Check Stock</b>\n\n"
        f"<b>SKU:</b> <code>{skus}</code>\n"
        f"<b>Source:</b> <code>{sources}</code>\n"
        f"<b>Unit:</b> <code>{units}</code>\n\n"
        "Jalankan query sekarang?",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Cek Stock", callback_data="checkstock_yes"),
                InlineKeyboardButton("❌ Batal", callback_data="checkstock_no"),
            ]
        ]),
    )
    return CHECKSTOCK_CONFIRM


async def checkstock_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Konfirmasi dan eksekusi check stock via Python requests."""
    query = update.callback_query
    await query.answer()

    if query.data == "checkstock_no":
        await query.edit_message_text("❌ Dibatalkan.")
        context.user_data.clear()
        return ConversationHandler.END

    data = context.user_data.get("checkstock_data")
    if not data:
        await query.edit_message_text("⚠️ Data tidak ditemukan. Coba ulang /stock.")
        context.user_data.clear()
        return ConversationHandler.END

    await query.edit_message_text("⏳ Mengecek stock, mohon tunggu...")

    try:
        import uuid
        articles = []
        for i, article_id in enumerate(data["article_ids"]):
            unit_val = data["units"][i] if i < len(data["units"]) else (data["units"][0] if data["units"] else "EA")
            articles.append({"ARTICLE_ID": article_id, "UNIT": unit_val})

        sources = [{"SOURCE_ID": s} for s in data["source_ids"]]

        payload = {
            "DATA": {
                "IDENTIFICATION": {
                    "INTERFACE_NAME": "stockquery",
                    "PARTNERS": "300",
                    "MSGID_EXTERNAL": str(uuid.uuid4()),
                    "ORDER_REF": "",
                    "SYSTEM_ORIGIN": ""
                },
                "ARTICLES": articles,
                "SOURCES": sources
            }
        }

        def _do_stock_request():
            import requests as req_lib
            resp = req_lib.post(
                "https://ping.erajaya.com/msoaa/api/v1/oaa-stock-query",
                headers={
                    "dbCode": "oaa-prod",
                    "Authorization": "Basic QXplYzpTM3J2aXMxbnQzcm40bA==",
                    "Content-Type": "application/json",
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                },
                json=payload,
                timeout=30,
            )
            return resp.text

        raw_output = await asyncio.to_thread(_do_stock_request)
        formatted = _format_stock_response(raw_output)

        if len(formatted) > 3500:
            formatted = formatted[:3500] + "\n... (output dipotong)"

        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"✅ <b>Hasil Check Stock</b>\n\n<pre>{html.escape(formatted)}</pre>",
            parse_mode=ParseMode.HTML,
            message_thread_id=_thread_id_from_update(update),
        )

    except Exception as e:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"⚠️ Gagal menjalankan check stock:\n{html.escape(str(e))}",
            message_thread_id=_thread_id_from_update(update),
        )

    context.user_data.clear()
    return ConversationHandler.END


def _format_stock_response(raw_output: str) -> str:
    """Coba parse JSON response dan format jadi tabel yang readable."""
    clean = re.sub(r"^Berhasil:\s*", "", raw_output, flags=re.IGNORECASE).strip()

    try:
        data = json.loads(clean)
        items = None
        if isinstance(data, dict):
            d = data.get("DATA") or data
            items = (
                d.get("ARTICLES_STOCKS")
                or d.get("STOCK_DATA")
                or d.get("ARTICLES")
            )

        if items and isinstance(items, list):
            lines = []
            for item in items:
                article = item.get("ARTICLE_ID", "")
                source = item.get("SOURCE_ID", "")
                qty = item.get("QUANTITY", item.get("ATP_QTY", item.get("QTY", "")))
                unit = item.get("UNIT", "")
                atp_date = item.get("ATP_DATE", "")

                lines.append(f"SKU      : {article}")
                lines.append(f"Source   : {source}")
                lines.append(f"Stock    : {qty} {unit}")
                if atp_date:
                    # Format date dari 20260825 ke 2026-08-25
                    if len(atp_date) == 8:
                        atp_date = f"{atp_date[:4]}-{atp_date[4:6]}-{atp_date[6:]}"
                    lines.append(f"ATP Date : {atp_date}")
                lines.append("──────────────────────────────")

            return "\n".join(lines).rstrip("──────────────────────────────\n")
    except (json.JSONDecodeError, KeyError, TypeError):
        pass

    return raw_output


# ==============================================================================
# CEK PROMO VIA TELEGRAM
# ==============================================================================
CEKPROMO_API_URL = "https://erpi.eraspace.com/api/v1/promo-srp-price/item"
CEKPROMO_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "cekpromo", "cek_promo_srp.ps1")


def _parse_cekpromo_text(text: str) -> dict:
    """
    Parse input teks user menjadi parameter cek promo.
    Format:
      sku: 8100102377
      bucode: AZ02
      channelid: 50       (opsional, default 50)
      membergroup: 00     (opsional, default 00)
    """
    result = {}

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        match = re.match(r"^(sku|artikle?|article)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["sku"] = match.group(2).strip()
            continue

        match = re.match(r"^(bucode|bu|store|toko)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["bucode"] = match.group(2).strip()
            continue

        match = re.match(r"^(channelid|channel)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["channelid"] = match.group(2).strip()
            continue

        match = re.match(r"^(membergroup|member)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["membergroup"] = match.group(2).strip()
            continue

    return result if result.get("sku") and result.get("bucode") else None


def _call_cekpromo_api(sku: str, bucode: str, channelid: str = "50", membergroup: str = "00") -> str:
    """Panggil API promo-srp-price via Python requests."""
    import requests as req_lib

    url = (
        f"{CEKPROMO_API_URL}?bucode={bucode}&channelid={channelid}"
        f"&membergroup={membergroup}&promoservice=0&qty=1&sku={sku}"
    )

    resp = req_lib.get(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        },
        timeout=30,
    )
    return resp.text


def _format_promo_response(data, sku: str, bucode: str) -> str:
    """Format response API jadi teks yang mudah dibaca."""
    if not data:
        return "Tidak ada data promo."

    # Struktur response: { promoPrice: [...], srpPrice: [...], bundling: [...] }
    promo_prices = []
    srp_prices = []

    if isinstance(data, dict):
        promo_prices = data.get("promoPrice") or []
        srp_prices = data.get("srpPrice") or []

    if not promo_prices and not srp_prices:
        return json.dumps(data, indent=2, ensure_ascii=False)

    # Ambil nama produk dari srpPrice
    product_name = ""
    srp_price_val = ""
    if srp_prices and isinstance(srp_prices, list):
        srp_item = srp_prices[0]
        product_name = srp_item.get("name", "")
        srp_price_val = srp_item.get("price", "")

    if not promo_prices:
        # Tidak ada promo, tampilkan info dasar dari srpPrice
        lines = [
            f"SKU           : {sku}",
            f"Name          : {product_name}",
            f"Original Price: {srp_price_val}",
            f"Price         : {srp_price_val}",
            f"BU Code       : {bucode}",
            "",
            "(Tidak ada promo aktif untuk SKU ini)",
        ]
        return "\n".join(lines)

    lines = []
    for promo_item in promo_prices:
        if not isinstance(promo_item, dict):
            continue

        item_sku = promo_item.get("sku") or sku
        original_price = promo_item.get("originalPrice") or ""
        price = promo_item.get("price") or ""
        bu_obj = promo_item.get("businessUnit") or {}
        item_bucode = bu_obj.get("buCode") or bucode

        # Ambil discount details
        discounts = promo_item.get("discount") or []

        if discounts:
            for disc in discounts:
                if not isinstance(disc, dict):
                    continue
                bonus_buy_id = disc.get("bonusBuyID") or ""
                amount = disc.get("amount") or ""
                condition_type = disc.get("conditionType") or ""
                optional = disc.get("optional") or ""
                is_pwp = disc.get("isPwp") if disc.get("isPwp") is not None else ""

                if lines:
                    lines.append("")
                lines.append(f"SKU           : {item_sku}")
                lines.append(f"Name          : {product_name}")
                lines.append(f"Original Price: {original_price}")
                lines.append(f"Price         : {price}")
                lines.append(f"BU Code       : {item_bucode}")
                lines.append(f"Bonus Buy ID  : {bonus_buy_id}")
                lines.append(f"Amount        : {amount}")
                lines.append(f"Condition Type: {condition_type}")
                lines.append(f"Optional      : {optional}")
                lines.append(f"Is PWP        : {is_pwp}")
        else:
            if lines:
                lines.append("")
            lines.append(f"SKU           : {item_sku}")
            lines.append(f"Name          : {product_name}")
            lines.append(f"Original Price: {original_price}")
            lines.append(f"Price         : {price}")
            lines.append(f"BU Code       : {item_bucode}")
            lines.append(f"Bonus Buy ID  : -")
            lines.append(f"Amount        : -")
            lines.append(f"Condition Type: -")
            lines.append(f"Optional      : -")
            lines.append(f"Is PWP        : -")

    return "\n".join(lines).strip()


@restricted
async def cekpromo_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /promo — mulai proses cek promo."""
    context.user_data.clear()

    # Cek apakah user langsung kirim parameter setelah command
    # Format shortcut: /promo 8100102377 AZ02
    parts = update.message.text.split(maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        args = parts[1].strip().split()
        if len(args) >= 2:
            data = {
                "sku": args[0],
                "bucode": args[1],
            }
            if len(args) > 2:
                data["channelid"] = args[2]
            if len(args) > 3:
                data["membergroup"] = args[3]
            context.user_data["cekpromo_data"] = data
            return await _cekpromo_execute(update, context)

    await update.message.reply_text(
        "<b>🏷 Cek Promo SRP Price</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Ketik langsung:</b>\n"
        "<code>sku: 8100102377\n"
        "bucode: AZ02</code>\n\n"
        "<b>2. Dengan opsi tambahan:</b>\n"
        "<code>sku: 8100102377\n"
        "bucode: AZ02\n"
        "channelid: 50\n"
        "membergroup: 00</code>\n\n"
        "<b>3. Shortcut:</b>\n"
        "<code>/promo 8100102377 AZ02</code>\n\n"
        "Default: channelid=50, membergroup=00\n\n"
        "Kirim sekarang, atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return CEKPROMO_INPUT


async def cekpromo_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user untuk cek promo."""
    text = update.message.text.strip()
    data = _parse_cekpromo_text(text)

    if not data:
        await update.message.reply_text(
            "Format tidak dikenali. Minimal harus ada <b>sku</b> dan <b>bucode</b>:\n\n"
            "<code>sku: 8100102377\n"
            "bucode: AZ02</code>\n\n"
            "Coba lagi atau /cancel untuk batal.",
            parse_mode=ParseMode.HTML,
        )
        return CEKPROMO_INPUT

    context.user_data["cekpromo_data"] = data
    return await _cekpromo_execute(update, context)


async def _cekpromo_execute(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Langsung eksekusi (karena cuma read/query, tanpa konfirmasi)."""
    data = context.user_data["cekpromo_data"]
    sku = data["sku"]
    bucode = data["bucode"]
    channelid = data.get("channelid", "50")
    membergroup = data.get("membergroup", "00")

    await update.message.reply_text(
        f"⏳ Mengecek promo SKU <code>{sku}</code> di BU <code>{bucode}</code>...",
        parse_mode=ParseMode.HTML,
    )

    try:
        raw_response = await asyncio.to_thread(
            _call_cekpromo_api, sku, bucode, channelid, membergroup
        )

        # Coba parse JSON
        try:
            data = json.loads(raw_response)
            formatted = _format_promo_response(data, sku, bucode)
        except json.JSONDecodeError:
            formatted = raw_response

        if len(formatted) > 3500:
            formatted = formatted[:3500] + "\n... (output dipotong)"

        await update.message.reply_text(
            f"🏷 <b>Hasil Cek Promo</b>\n\n<pre>{html.escape(formatted)}</pre>",
            parse_mode=ParseMode.HTML,
        )

    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal cek promo:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )

    context.user_data.clear()
    return ConversationHandler.END


# ==============================================================================
# CEK AWB / TRACKING ORDER VIA TELEGRAM
# ==============================================================================
CEKAWB_API_URL = "https://erpi.eraspace.com/api/v1/oms/tracking/order"
CEKAWB_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "cekawb", "tracking_order.ps1")


def _parse_cekawb_text(text: str) -> dict:
    """
    Parse input teks user menjadi parameter cek AWB.
    Format:
      ordernumber: 3301352973
      source: IBOX
    """
    result = {}

    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue

        match = re.match(r"^(ordernumber|order|no)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["orderNumber"] = match.group(2).strip()
            continue

        match = re.match(r"^(source|src|channel)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            result["source"] = match.group(2).strip().upper()
            continue

    return result if result.get("orderNumber") and result.get("source") else None


def _call_cekawb_api(order_number: str, source: str) -> str:
    """Panggil API OMS tracking via Python requests."""
    import requests as req_lib

    payload = {"orderNumber": order_number, "source": source}
    resp = req_lib.post(
        CEKAWB_API_URL,
        headers={
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        },
        json=payload,
        timeout=30,
    )
    return resp.text


def _format_awb_response(data, order_number: str, source: str) -> str:
    """Format response API tracking jadi teks yang mudah dibaca."""
    if not data or not isinstance(data, dict):
        return "Tidak ada data tracking."

    status_code = data.get("statusCode", "")
    message = data.get("message", "")

    # Kalau error
    if status_code and str(status_code) != "200" and str(status_code) != "0":
        return f"Order: {order_number} | Source: {source}\n\n⚠️ {message}"

    courier_code = data.get("courierCode") or "-"
    tracking_no = data.get("trackingNo") or "-"
    enter_date = data.get("enterDate") or "-"
    process_date = data.get("processDate") or "-"
    shipped_date = data.get("shippedDate") or "-"
    delivered_date = data.get("deliveredDate") or "-"
    sales_no = data.get("salesNo") or "-"
    current_status = data.get("currentStatus") or "-"
    delivery_no = data.get("deliveryNo") or "-"
    source_system = data.get("sourceSystem") or "-"
    billing_no = data.get("billingNo") or "-"

    lines = [
        f"Order Number  : {order_number}",
        f"Source        : {source}",
        f"Sales No      : {sales_no}",
        f"Courier       : {courier_code}",
        f"AWB / Resi    : {tracking_no}",
        f"Status        : {current_status}",
        f"Delivery No   : {delivery_no}",
        f"Billing No    : {billing_no}",
        f"Source System : {source_system}",
        "",
        "📅 Timeline:",
        f"  Enter     : {enter_date}",
        f"  Process   : {process_date}",
        f"  Shipped   : {shipped_date}",
        f"  Delivered : {delivered_date}",
    ]

    return "\n".join(lines)


@restricted
async def cekawb_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /awb — mulai proses cek AWB/tracking."""
    context.user_data.clear()

    # Cek apakah user langsung kirim parameter: /awb 3301352973 IBOX
    parts = update.message.text.split(maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        args = parts[1].strip().split()
        if len(args) >= 2:
            context.user_data["cekawb_data"] = {
                "orderNumber": args[0],
                "source": args[1].upper(),
            }
            return await _cekawb_execute(update, context)

    await update.message.reply_text(
        "<b>📦 Cek AWB / Tracking Order</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Ketik langsung:</b>\n"
        "<code>ordernumber: 3301352973\n"
        "source: IBOX</code>\n\n"
        "<b>2. Shortcut:</b>\n"
        "<code>/awb 3301352973 IBOX</code>\n\n"
        "Kirim sekarang, atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return CEKAWB_INPUT


async def cekawb_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user untuk cek AWB."""
    text = update.message.text.strip()
    data = _parse_cekawb_text(text)

    if not data:
        await update.message.reply_text(
            "Format tidak dikenali. Minimal harus ada <b>ordernumber</b> dan <b>source</b>:\n\n"
            "<code>ordernumber: 3301352973\n"
            "source: IBOX</code>\n\n"
            "Coba lagi atau /cancel untuk batal.",
            parse_mode=ParseMode.HTML,
        )
        return CEKAWB_INPUT

    context.user_data["cekawb_data"] = data
    return await _cekawb_execute(update, context)


async def _cekawb_execute(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Langsung eksekusi (query, tanpa konfirmasi)."""
    data = context.user_data["cekawb_data"]
    order_number = data["orderNumber"]
    source = data["source"]

    await update.message.reply_text(
        f"⏳ Mengecek tracking order <code>{order_number}</code> ({source})...",
        parse_mode=ParseMode.HTML,
    )

    try:
        raw_response = await asyncio.to_thread(
            _call_cekawb_api, order_number, source
        )

        try:
            resp_data = json.loads(raw_response)
            formatted = _format_awb_response(resp_data, order_number, source)
        except json.JSONDecodeError:
            formatted = raw_response

        if len(formatted) > 3500:
            formatted = formatted[:3500] + "\n... (output dipotong)"

        await update.message.reply_text(
            f"📦 <b>Hasil Cek AWB</b>\n\n<pre>{html.escape(formatted)}</pre>",
            parse_mode=ParseMode.HTML,
        )

    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal cek AWB:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )

    context.user_data.clear()
    return ConversationHandler.END


# ==============================================================================
# QUERY DATABASE (ADMINER VIA PLAYWRIGHT) - INTEGRATED FROM bot_core
# ==============================================================================
BOTCORE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "coreitops", "bot_core")


def _run_botcore_query(specific_files: list = None) -> list:
    """Jalankan bot_notifier query via Playwright (sync, dipanggil dari thread).
    specific_files: jika diisi, hanya jalankan file query tertentu.
    Return list of (query_file, excel_path, row_count, headers, data_list) untuk setiap query yang berhasil."""
    import sys
    if BOTCORE_DIR not in sys.path:
        sys.path.insert(0, BOTCORE_DIR)

    from bot_notifier import load_sql_query, create_excel_report
    from playwright.sync_api import sync_playwright
    from bs4 import BeautifulSoup
    import time as _time

    # Load config
    config_path = os.path.join(BOTCORE_DIR, "config.json")
    with open(config_path, "r", encoding="utf-8") as f:
        config_data = json.load(f)

    # Support format lama (single "adminer") dan baru (array "databases")
    databases = config_data.get("databases")
    if not databases:
        # Format lama: konversi ke format baru
        databases = [{
            "name": config_data.get("adminer", {}).get("db", "default"),
            "adminer": config_data["adminer"],
            "query_files": config_data.get("query_files", config_data.get("query_file", ["query.sql"])),
        }]

    results = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)

        try:
            for db_entry in databases:
                adm = db_entry["adminer"]
                db_name = db_entry.get("name", adm.get("db", "unknown"))

                # Tentukan query files untuk database ini
                if specific_files:
                    # Cek apakah file ada di query_files database ini
                    raw_files = db_entry.get("query_files", [])
                    if isinstance(raw_files, str):
                        raw_files = [raw_files]
                    flat_files = []
                    for item in raw_files:
                        if isinstance(item, list):
                            flat_files.extend(item)
                        else:
                            flat_files.append(item)
                    # Jalankan jika file terdaftar di DB ini, ATAU jika ini DB pertama (fallback)
                    query_files = [f for f in specific_files if f in flat_files]
                    if not query_files:
                        # Fallback: jika file tidak terdaftar di DB manapun, coba di DB pertama
                        if db_entry == databases[0]:
                            query_files = [f for f in specific_files if os.path.exists(os.path.join(BOTCORE_DIR, f))]
                        if not query_files:
                            continue
                else:
                    raw_files = db_entry.get("query_files", [])
                    if isinstance(raw_files, str):
                        query_files = [raw_files]
                    elif isinstance(raw_files, list):
                        query_files = []
                        for item in raw_files:
                            if isinstance(item, list):
                                query_files.extend(item)
                            else:
                                query_files.append(item)
                    else:
                        query_files = []

                if not query_files:
                    continue

                # Buat context browser baru per database (credentials bisa beda)
                http_credentials = None
                if adm.get("basic_auth_user") and adm.get("basic_auth_pass"):
                    http_credentials = {
                        "username": adm["basic_auth_user"],
                        "password": adm["basic_auth_pass"]
                    }

                ctx = browser.new_context(
                    http_credentials=http_credentials,
                    user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
                )
                page = ctx.new_page()

                try:
                    page.goto(adm["url"], wait_until="networkidle")

                    driver_select = page.locator("select[name='auth[driver]']")
                    if driver_select.count() > 0 and adm.get("driver"):
                        try:
                            driver_select.select_option(value=adm["driver"], timeout=2000)
                        except Exception:
                            pass

                    page.fill("input[name='auth[server]']", adm["server"])
                    page.fill("input[name='auth[username]']", adm["username"])
                    page.fill("input[name='auth[password]']", adm["password"])

                    db_input = page.locator("input[name='auth[db]']")
                    if db_input.count() > 0:
                        db_input.fill(adm["db"])

                    page.click("input[type='submit']")
                    page.wait_for_load_state("networkidle")

                    if page.locator("div.error").count() > 0:
                        err_txt = page.locator("div.error").first.text_content()
                        logger.error(f"Gagal Login Adminer [{db_name}]: {err_txt.strip()}")
                        ctx.close()
                        continue

                    for q_file in query_files:
                        try:
                            sql_query = load_sql_query(os.path.join(BOTCORE_DIR, q_file))
                        except Exception as file_err:
                            logger.error(f"Gagal membaca file {q_file}: {file_err}")
                            continue

                        sql_link = page.locator("a:has-text('SQL command'), a:has-text('SQL dotaz'), a[href*='sql=']").first
                        if sql_link.count() > 0:
                            sql_link.click()
                            page.wait_for_load_state("networkidle")
                        else:
                            query_url = f"{adm['url']}?username={adm['username']}&db={adm['db']}&sql="
                            page.goto(query_url, wait_until="networkidle")

                        page.wait_for_selector("textarea[name='query']", state="attached", timeout=10000)

                        with page.expect_navigation(wait_until="networkidle"):
                            page.evaluate("""
                                (queryText) => {
                                    const textarea = document.querySelector("textarea[name='query']");
                                    if (textarea) {
                                        textarea.value = queryText;
                                        const form = textarea.closest("form");
                                        if (form) { form.submit(); }
                                    }
                                }
                            """, sql_query)

                        html_content = page.content()
                        soup = BeautifulSoup(html_content, "html.parser")

                        error_div = soup.find("div", class_="error")
                        if error_div:
                            logger.error(f"Adminer Query Error [{db_name}] ({q_file}): {error_div.text.strip()}")
                            continue

                        table = soup.find("table", class_="printable") or soup.find("table")
                        rows = table.find_all("tr") if table else []

                        data_list = []
                        headers = []
                        if rows and len(rows) >= 2:
                            headers = [th.text.strip() for th in rows[0].find_all(["th", "td"])]
                            for row in rows[1:]:
                                cols = [td.text.strip() for td in row.find_all("td")]
                                if cols and any(cols):
                                    data_list.append(cols)

                        base_name = os.path.splitext(q_file)[0]
                        excel_filename = f"{base_name}.xlsx"
                        excel_filepath = os.path.join(BOTCORE_DIR, excel_filename)

                        if not data_list:
                            create_excel_report(["Status"], [["No Rows"]], excel_filepath)
                        else:
                            create_excel_report(headers, data_list, excel_filepath)

                        results.append((q_file, excel_filepath, len(data_list), headers, data_list))
                        _time.sleep(1)

                finally:
                    ctx.close()

        finally:
            browser.close()

    return results


async def query_scheduler_job(context: ContextTypes.DEFAULT_TYPE):
    """Background job: jalankan query database dan kirim hasilnya ke Telegram."""
    try:
        results = await asyncio.to_thread(_run_botcore_query)
    except Exception as e:
        logger.error(f"Query scheduler error: {e}")
        return

    for q_file, excel_path, row_count, _, _ in results:
        if not os.path.exists(excel_path):
            continue
        try:
            caption = f"📊 Hasil Query ({q_file})\nTotal SKU: {row_count}" if row_count > 0 else f"hasil export ({q_file})\nStatus: No Rows"
            for chat_id, thread_id in config.notify_targets():
                with open(excel_path, "rb") as f:
                    await context.bot.send_document(
                        chat_id=chat_id,
                        document=f,
                        filename=os.path.basename(excel_path),
                        caption=caption,
                        message_thread_id=thread_id,
                        read_timeout=120,
                        write_timeout=120,
                        connect_timeout=30,
                    )
        except Exception as e:
            logger.error(f"Gagal kirim hasil query {q_file}: {e}")
        finally:
            try:
                os.remove(excel_path)
            except OSError:
                pass

    if results:
        logger.info(f"Query scheduler: {len(results)} query berhasil dieksekusi dan dikirim.")


@restricted
async def query_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /query — trigger manual eksekusi query database.
    Tanpa argumen: tampilkan daftar query yang tersedia.
    Dengan argumen: /query query1 → jalankan query1.sql langsung.
    """
    parts = update.message.text.split(maxsplit=1)

    if len(parts) > 1 and parts[1].strip():
        arg = parts[1].strip()
        if not arg.endswith(".sql"):
            arg += ".sql"
        file_path = os.path.join(BOTCORE_DIR, arg)
        if not os.path.exists(file_path):
            available = [f for f in os.listdir(BOTCORE_DIR) if f.endswith(".sql")]
            await update.message.reply_text(
                f"⚠️ File <code>{arg}</code> tidak ditemukan.\n\n"
                f"File query yang tersedia:\n"
                + "\n".join(f"• <code>{f}</code>" for f in available),
                parse_mode=ParseMode.HTML,
            )
            return
        # Langsung eksekusi
        await _query_execute(update, context, [arg])
        return

    # Tanpa argumen: tampilkan daftar query sebagai tombol
    available = sorted([f for f in os.listdir(BOTCORE_DIR) if f.endswith(".sql")])
    if not available:
        await update.message.reply_text("Tidak ada file query (.sql) yang tersedia.")
        return

    buttons = []
    for f in available:
        name = os.path.splitext(f)[0]
        buttons.append([InlineKeyboardButton(f"📋 {name}", callback_data=f"qrun_{name}")])

    await update.message.reply_text(
        "<b>📂 Pilih query yang ingin dijalankan:</b>",
        parse_mode=ParseMode.HTML,
        reply_markup=InlineKeyboardMarkup(buttons),
    )


async def query_pick_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback ketika user pilih query dari daftar."""
    cb = update.callback_query
    await cb.answer()

    query_name = cb.data.replace("qrun_", "")
    sql_file = f"{query_name}.sql"
    file_path = os.path.join(BOTCORE_DIR, sql_file)

    if not os.path.exists(file_path):
        await cb.edit_message_text(f"⚠️ File <code>{sql_file}</code> tidak ditemukan.", parse_mode=ParseMode.HTML)
        return

    await cb.edit_message_text(f"⏳ Menjalankan <code>{sql_file}</code>, mohon tunggu...", parse_mode=ParseMode.HTML)
    await _query_execute(update, context, [sql_file])


async def _query_execute(update: Update, context: ContextTypes.DEFAULT_TYPE, specific_files: list):
    """Eksekusi query dan tampilkan opsi output."""
    # Tentukan chat untuk reply
    msg = update.effective_message

    try:
        results = await asyncio.to_thread(_run_botcore_query, specific_files)
    except Exception as e:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"⚠️ Gagal menjalankan query:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
            message_thread_id=_thread_id_from_update(update),
        )
        return

    if not results:
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text="Tidak ada hasil query yang berhasil dieksekusi.",
            message_thread_id=_thread_id_from_update(update),
        )
        return

    # Simpan hasil di user_data untuk dipakai callback
    context.user_data["query_results"] = results

    # Tampilkan summary + opsi output
    summary_lines = []
    for i, (q_file, _, row_count, _, _) in enumerate(results):
        summary_lines.append(f"• <code>{q_file}</code> — {row_count} baris")

    await context.bot.send_message(
        chat_id=update.effective_chat.id,
        text="<b>✅ Query berhasil!</b>\n\n"
        + "\n".join(summary_lines)
        + "\n\nPilih format output:",
        parse_mode=ParseMode.HTML,
        message_thread_id=_thread_id_from_update(update),
        reply_markup=InlineKeyboardMarkup([
            [
                InlineKeyboardButton("📄 Tampilkan Text", callback_data="qout_text"),
                InlineKeyboardButton("📊 Export Excel", callback_data="qout_excel"),
            ]
        ]),
    )


async def query_output_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback untuk pilihan output query: text atau excel."""
    query = update.callback_query
    await query.answer()

    results = context.user_data.get("query_results")
    if not results:
        await query.edit_message_text("⚠️ Hasil query tidak ditemukan. Coba jalankan /query lagi.")
        return

    choice = query.data  # "qout_text" atau "qout_excel"

    if choice == "qout_text":
        await query.edit_message_text("📄 Menampilkan hasil sebagai teks...")

        for q_file, excel_path, row_count, headers, data_list in results:
            if not data_list:
                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=f"📄 <b>{q_file}</b>\n\nNo Rows",
                    parse_mode=ParseMode.HTML,
                    message_thread_id=_thread_id_from_update(update),
                )
            else:
                # Format sebagai tabel text dengan label per field
                text_lines = [
                    f"List SKU ini harap dicek dan disesuaikan agar tidak terjadi jebol stock",
                    f"Total SKU: {row_count}",
                    "",
                ]

                # Data (batasi 30 baris agar tidak terlalu panjang)
                max_rows = 30
                for idx, row in enumerate(data_list[:max_rows], 1):
                    text_lines.append(f"Baris #{idx}")
                    for col_idx, cell in enumerate(row):
                        col_name = headers[col_idx] if col_idx < len(headers) else f"col{col_idx}"
                        text_lines.append(f"• {col_name}: {cell}")
                    if idx < min(len(data_list), max_rows):
                        text_lines.append("──────────────────────────────")

                if len(data_list) > max_rows:
                    text_lines.append(f"\n... dan {len(data_list) - max_rows} baris lainnya (gunakan Excel untuk data lengkap)")

                text_output = "\n".join(text_lines)

                # Telegram limit 4096 chars
                if len(text_output) > 3800:
                    text_output = text_output[:3800] + "\n\n... (dipotong, gunakan Excel untuk data lengkap)"

                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=text_output,
                    parse_mode=ParseMode.HTML,
                    message_thread_id=_thread_id_from_update(update),
                )

            # Cleanup excel file
            if excel_path and os.path.exists(excel_path):
                try:
                    os.remove(excel_path)
                except OSError:
                    pass

    elif choice == "qout_excel":
        await query.edit_message_text("📊 Mengirim file Excel...")

        for q_file, excel_path, row_count, _, _ in results:
            if not os.path.exists(excel_path):
                continue
            try:
                caption = f"📊 Hasil Query ({q_file})\nTotal SKU: {row_count}" if row_count > 0 else f"hasil export ({q_file})\nStatus: No Rows"
                with open(excel_path, "rb") as f:
                    await context.bot.send_document(
                        chat_id=update.effective_chat.id,
                        document=f,
                        filename=os.path.basename(excel_path),
                        caption=caption,
                        message_thread_id=_thread_id_from_update(update),
                        read_timeout=120,
                        write_timeout=120,
                        connect_timeout=30,
                    )
            except Exception as e:
                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=f"⚠️ Gagal kirim file {q_file}: {e}",
                    message_thread_id=_thread_id_from_update(update),
                )
            finally:
                try:
                    os.remove(excel_path)
                except OSError:
                    pass

    context.user_data.pop("query_results", None)


# ==============================================================================
# AWB JNE TRACKING VIA TELEGRAM
# ==============================================================================
AWBJNE_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "tools", "awbjne", "tracking_jne.ps1")


def _parse_awbjne_text(text: str) -> list:
    """
    Parse input teks user menjadi list parameter AWB JNE.
    Support multiple pairs dipisah baris kosong.
    Format:
      ordernumber: 8402663858
      awb: 0157352600237230

      ordernumber: 8402679862
      awb: 0157352600240929
    Return: list of dict {"order_number": ..., "awb": ...}
    """
    items = []
    current = {}

    for line in text.splitlines():
        line = line.strip()
        if not line:
            if current.get("order_number") and current.get("awb"):
                items.append(current)
                current = {}
            continue

        match = re.match(r"^(ordernumber|order|no)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            # Jika sudah ada order_number di current, simpan dulu (pair baru)
            if current.get("order_number") and current.get("awb"):
                items.append(current)
                current = {}
            current["order_number"] = match.group(2).strip()
            continue

        match = re.match(r"^(awb|resi|tracking)\s*[:=]\s*(.+)", line, re.IGNORECASE)
        if match:
            current["awb"] = match.group(2).strip()
            continue

    # Sisa terakhir
    if current.get("order_number") and current.get("awb"):
        items.append(current)

    return items if items else None


def _call_awbjne_api(order_number: str, awb: str) -> str:
    """Panggil API tracking JNE via Python requests."""
    import requests as req_lib

    resp = req_lib.post(
        "https://jeanne.eraspace.com/shippings/v2/tracking/order/oms/jne",
        headers={
            "authorization": "Basic c2hpcHBpbmdiYXNpYzo3NmNkNDJlZTQzZTUxNTIzZTAzNTVjZDE3NTMxY2ZjZjQxYjE2MWNmZDJjNTgwNDJkZjkxZTVmODU1MDQwYTQx",
            "x-source": "eraspace",
            "x-platform": "omsservice",
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        },
        json={"order_number": order_number, "awb": awb},
        timeout=30,
    )
    return resp.text


def _format_awbjne_response(data, order_number: str, awb: str) -> str:
    """Format response API tracking JNE."""
    if not data or not isinstance(data, dict):
        return "Tidak ada data tracking."

    # Cek error
    error_msg = data.get("error_message")
    if error_msg:
        return f"Order: {order_number} | AWB: {awb}\n\n⚠️ {error_msg}"

    d = data.get("data")
    if not d:
        return f"Order: {order_number} | AWB: {awb}\n\nTidak ada data."

    status = d.get("status") or "-"
    service = d.get("service") or "-"
    shipment_date = d.get("date_of_shipment") or "-"
    origin = d.get("origin") or "-"
    destination = d.get("destination") or "-"
    shipper = d.get("shipper") or "-"
    consignee = d.get("consignee") or "-"

    lines = [
        f"Order Number : {order_number}",
        f"AWB / Resi   : {awb}",
        f"Status       : {status}",
        f"Service      : {service}",
        f"Tanggal Kirim: {shipment_date}",
        f"Origin       : {origin}",
        f"Destination  : {destination}",
        f"Shipper      : {shipper}",
        f"Consignee    : {consignee}",
    ]

    # History tracking
    history = d.get("history_shipping") or []
    if history:
        lines.append("")
        lines.append("📦 Tracking History:")
        for h in history:
            date = h.get("date", "")
            status_text = h.get("status", "")
            lines.append(f"  [{date}] {status_text}")

    return "\n".join(lines)


@restricted
async def awbjne_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /awbjne — cek tracking AWB JNE (support multiple)."""
    context.user_data.clear()

    # Shortcut: /awbjne 8402663858 0157352600237230 8402679862 0157352600240929
    parts = update.message.text.split(maxsplit=1)
    if len(parts) > 1 and parts[1].strip():
        arg_text = parts[1].strip()

        # Coba parse sebagai label format dulu
        items = _parse_awbjne_text(arg_text)

        # Jika tidak match label format, coba parse sebagai pairs (ordernumber awb ordernumber awb ...)
        if not items:
            args = arg_text.split()
            if len(args) >= 2 and len(args) % 2 == 0:
                items = []
                for i in range(0, len(args), 2):
                    items.append({"order_number": args[i], "awb": args[i + 1]})
            elif len(args) == 2:
                items = [{"order_number": args[0], "awb": args[1]}]

        if items:
            context.user_data["awbjne_items"] = items
            return await _awbjne_execute_multi(update, context)

    await update.message.reply_text(
        "<b>📦 Cek AWB JNE</b>\n\n"
        "Pilih salah satu cara input:\n\n"
        "<b>1. Single:</b>\n"
        "<code>/awbjne 8402663858 0157352600237230</code>\n\n"
        "<b>2. Multiple (shortcut):</b>\n"
        "<code>/awbjne 8402663858 0157352600237230 8402679862 0157352600240929</code>\n\n"
        "<b>3. Multiple (label):</b>\n"
        "<code>ordernumber: 8402663858\n"
        "awb: 0157352600237230\n\n"
        "ordernumber: 8402679862\n"
        "awb: 0157352600240929</code>\n\n"
        "Kirim sekarang, atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return AWBJNE_INPUT


async def awbjne_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima input dari user untuk cek AWB JNE (support multiple)."""
    text = update.message.text.strip()
    items = _parse_awbjne_text(text)

    if not items:
        await update.message.reply_text(
            "Format tidak dikenali. Minimal harus ada <b>ordernumber</b> dan <b>awb</b>:\n\n"
            "<code>ordernumber: 8402663858\n"
            "awb: 0157352600237230</code>\n\n"
            "Untuk multiple, pisahkan dengan baris kosong.\n"
            "Coba lagi atau /cancel untuk batal.",
            parse_mode=ParseMode.HTML,
        )
        return AWBJNE_INPUT

    context.user_data["awbjne_items"] = items
    return await _awbjne_execute_multi(update, context)


async def _awbjne_execute_multi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Eksekusi tracking JNE untuk satu atau banyak AWB."""
    items = context.user_data.get("awbjne_items", [])

    await update.message.reply_text(
        f"⏳ Mengecek {len(items)} AWB JNE, mohon tunggu...",
        parse_mode=ParseMode.HTML,
    )

    for item in items:
        order_number = item["order_number"]
        awb = item["awb"]

        try:
            raw_response = await asyncio.to_thread(_call_awbjne_api, order_number, awb)

            try:
                resp_data = json.loads(raw_response)
                formatted = _format_awbjne_response(resp_data, order_number, awb)
            except json.JSONDecodeError:
                if "Cloudflare" in raw_response or "<!DOCTYPE" in raw_response:
                    formatted = "⚠️ API diblok Cloudflare. Cookie mungkin expired.\nCoba lagi nanti atau update cookie di kode."
                else:
                    formatted = raw_response[:500]

            if len(formatted) > 3500:
                formatted = formatted[:3500] + "\n... (output dipotong)"

            await update.message.reply_text(
                f"📦 <b>Tracking JNE</b>\n\n<pre>{html.escape(formatted)}</pre>",
                parse_mode=ParseMode.HTML,
            )

        except Exception as e:
            await update.message.reply_text(
                f"⚠️ Gagal cek AWB JNE ({awb}):\n<code>{html.escape(str(e))}</code>",
                parse_mode=ParseMode.HTML,
            )

    context.user_data.clear()
    return ConversationHandler.END


# ==============================================================================
# UPDATE JADWAL SHIFT VIA TELEGRAM
# ==============================================================================
SHIFT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "jadwal_shift.xlsx")


@restricted
async def updateshift_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /updateshift — upload file atau update via teks.

    Format teks:
        /updateshift 2026-08-26 07:00 15:00 SHIFT_1 Bagus, Tri
    Jika tanpa argumen, minta upload file xlsx.
    """
    # Cek apakah ada argumen teks inline
    if context.args and len(context.args) >= 5:
        # Parse inline: tanggal jam_mulai jam_selesai shift teknisi...
        tanggal_str = context.args[0]
        jam_mulai_str = context.args[1]
        jam_selesai_str = context.args[2]
        shift_name = context.args[3]
        teknisi_str = " ".join(context.args[4:])

        # Validasi format tanggal
        try:
            tanggal = dt.datetime.strptime(tanggal_str, "%Y-%m-%d").date()
        except ValueError:
            await update.message.reply_text(
                "⚠️ Format tanggal salah. Gunakan <code>YYYY-MM-DD</code>\n"
                "Contoh: <code>/updateshift 2026-08-26 07:00 15:00 SHIFT_1 Bagus, Tri</code>",
                parse_mode=ParseMode.HTML,
            )
            return ConversationHandler.END

        # Validasi format jam
        try:
            dt.datetime.strptime(jam_mulai_str, "%H:%M")
            dt.datetime.strptime(jam_selesai_str, "%H:%M")
        except ValueError:
            await update.message.reply_text(
                "⚠️ Format jam salah. Gunakan <code>HH:MM</code>\n"
                "Contoh: <code>/updateshift 2026-08-26 07:00 15:00 SHIFT_1 Bagus, Tri</code>",
                parse_mode=ParseMode.HTML,
            )
            return ConversationHandler.END

        # Tulis/update ke Google Sheets
        try:
            _sheets_upsert_shift(tanggal, jam_mulai_str, jam_selesai_str, shift_name, teknisi_str)
            await update.message.reply_text(
                f"✅ <b>Jadwal shift berhasil diupdate!</b>\n\n"
                f"📅 Tanggal: <code>{tanggal_str}</code>\n"
                f"⏰ Jam: <code>{jam_mulai_str} - {jam_selesai_str}</code>\n"
                f"🏷 Shift: <code>{shift_name}</code>\n"
                f"👤 Teknisi: <code>{teknisi_str}</code>",
                parse_mode=ParseMode.HTML,
            )
        except Exception as e:
            await update.message.reply_text(
                f"⚠️ Gagal update jadwal:\n<code>{html.escape(str(e))}</code>",
                parse_mode=ParseMode.HTML,
            )
        return ConversationHandler.END

    # Tanpa argumen → minta upload file
    context.user_data.clear()
    await update.message.reply_text(
        "<b>📅 Update Jadwal Shift</b>\n\n"
        "Upload file <code>jadwal_shift.xlsx</code> baru.\n\n"
        "<b>Atau gunakan format teks:</b>\n"
        "<code>/updateshift YYYY-MM-DD HH:MM HH:MM SHIFT_X Nama1, Nama2</code>\n\n"
        "Format kolom file:\n"
        "A: Tanggal (2026-08-25)\n"
        "B: Jam Mulai (07:00)\n"
        "C: Jam Selesai (15:00)\n"
        "D: Nama Shift (SHIFT_1)\n"
        "E: Teknisi (nama, pisah koma jika &gt;1)\n\n"
        "Kirim file sekarang, atau /cancel untuk batal.",
        parse_mode=ParseMode.HTML,
    )
    return UPDATESHIFT_UPLOAD


async def updateshift_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Terima file xlsx dan merge (upsert) ke Google Sheets."""
    if not update.message.document:
        await update.message.reply_text("Kirim file .xlsx, bukan teks. Coba lagi atau /cancel.")
        return UPDATESHIFT_UPLOAD

    doc = update.message.document
    if not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("File harus berformat .xlsx. Coba upload ulang.")
        return UPDATESHIFT_UPLOAD

    await update.message.reply_text("⏳ Mengunduh dan memproses file...")

    try:
        file = await doc.get_file()
        temp_path = SHIFT_FILE + ".tmp"
        await file.download_to_drive(temp_path)

        # Baca dan validasi file Excel
        wb = openpyxl.load_workbook(temp_path, data_only=True)
        sheet = wb.active
        new_rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 5 or not row[0]:
                continue
            cell_tgl = row[0]
            if isinstance(cell_tgl, (dt.datetime, dt.date)):
                tgl_str = cell_tgl.strftime("%Y-%m-%d")
            else:
                tgl_str = str(cell_tgl).strip()
            # Validasi format tanggal
            try:
                dt.datetime.strptime(tgl_str, "%Y-%m-%d")
            except ValueError:
                continue

            jam_mulai = row[1].strftime("%H:%M") if isinstance(row[1], (dt.time, dt.datetime)) else str(row[1]).strip()
            jam_selesai = row[2].strftime("%H:%M") if isinstance(row[2], (dt.time, dt.datetime)) else str(row[2]).strip()
            shift_name = str(row[3]).strip()
            teknisi = str(row[4]).strip()
            new_rows.append((tgl_str, jam_mulai, jam_selesai, shift_name, teknisi))

        wb.close()
        os.remove(temp_path)

        if not new_rows:
            await update.message.reply_text(
                "⚠️ File tidak valid — tidak ada data jadwal ditemukan (baris kosong atau format salah).\n"
                "Pastikan format kolom benar. Coba upload ulang atau /cancel."
            )
            return UPDATESHIFT_UPLOAD

        # Merge/upsert ke Google Sheets (data lama yang tidak ada di file TIDAK dihapus)
        added = 0
        updated = 0
        for tgl_str, jam_mulai, jam_selesai, shift_name, teknisi in new_rows:
            tanggal = dt.datetime.strptime(tgl_str, "%Y-%m-%d").date()
            # _sheets_upsert_shift: update jika tanggal+shift sama, tambah jika belum ada
            _sheets_upsert_shift(tanggal, jam_mulai, jam_selesai, shift_name, teknisi)

        await update.message.reply_text(
            f"✅ <b>Jadwal shift berhasil di-merge ke Google Sheets!</b>\n\n"
            f"File: <code>{doc.file_name}</code>\n"
            f"Total baris diproses: <b>{len(new_rows)}</b>\n\n"
            "• Baris dengan tanggal+shift yang sudah ada → <b>diupdate</b>\n"
            "• Baris baru → <b>ditambahkan</b>\n"
            "• Data lama di sheet yang tidak ada di file → <b>tetap aman</b>",
            parse_mode=ParseMode.HTML,
        )

    except Exception as e:
        temp_path = SHIFT_FILE + ".tmp"
        if os.path.exists(temp_path):
            os.remove(temp_path)
        await update.message.reply_text(
            f"⚠️ Gagal memproses file:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )

    context.user_data.clear()
    return ConversationHandler.END


# ==============================================================================
# HELPER: UPSERT / TAMBAH / HAPUS BARIS DI JADWAL SHIFT EXCEL
# ==============================================================================

def _ensure_shift_workbook():
    """Buka atau buat file jadwal_shift.xlsx dengan header."""
    if os.path.exists(SHIFT_FILE):
        wb = openpyxl.load_workbook(SHIFT_FILE)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Tanggal", "Jam Mulai", "Jam Selesai", "Shift", "Teknisi"])
        wb.save(SHIFT_FILE)
    return wb


def _upsert_shift_row(tanggal: dt.date, jam_mulai: str, jam_selesai: str, shift_name: str, teknisi: str):
    """Update baris jika tanggal+shift sudah ada, atau tambah baris baru."""
    wb = _ensure_shift_workbook()
    ws = wb.active

    target_tgl_str = tanggal.strftime("%Y-%m-%d")

    # Cari baris dengan tanggal + shift yang sama untuk di-update
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_tgl = row[0].value
        cell_shift = row[3].value if len(row) > 3 else None

        if isinstance(cell_tgl, (dt.date, dt.datetime)):
            existing_tgl = cell_tgl.strftime("%Y-%m-%d")
        else:
            existing_tgl = str(cell_tgl).strip() if cell_tgl else ""

        existing_shift = str(cell_shift).strip() if cell_shift else ""

        if existing_tgl == target_tgl_str and existing_shift.upper() == shift_name.upper():
            # Update baris existing
            ws.cell(row=row_idx, column=1, value=tanggal)
            ws.cell(row=row_idx, column=2, value=jam_mulai)
            ws.cell(row=row_idx, column=3, value=jam_selesai)
            ws.cell(row=row_idx, column=4, value=shift_name)
            ws.cell(row=row_idx, column=5, value=teknisi)
            wb.save(SHIFT_FILE)
            wb.close()
            return

    # Tidak ditemukan → tambah baris baru
    ws.append([tanggal, jam_mulai, jam_selesai, shift_name, teknisi])
    wb.save(SHIFT_FILE)
    wb.close()


def _add_shift_row(tanggal: dt.date, jam_mulai: str, jam_selesai: str, shift_name: str, teknisi: str):
    """Selalu tambah baris baru (tidak cek duplikat)."""
    wb = _ensure_shift_workbook()
    ws = wb.active
    ws.append([tanggal, jam_mulai, jam_selesai, shift_name, teknisi])
    wb.save(SHIFT_FILE)
    wb.close()


def _delete_shift_rows(tanggal: dt.date = None, shift_name: str = None) -> int:
    """Hapus baris berdasarkan tanggal dan/atau shift. Return jumlah baris terhapus."""
    if not os.path.exists(SHIFT_FILE):
        return 0

    wb = openpyxl.load_workbook(SHIFT_FILE)
    ws = wb.active

    target_tgl_str = tanggal.strftime("%Y-%m-%d") if tanggal else None
    rows_to_delete = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_tgl = row[0].value
        cell_shift = row[3].value if len(row) > 3 else None

        if isinstance(cell_tgl, (dt.date, dt.datetime)):
            existing_tgl = cell_tgl.strftime("%Y-%m-%d")
        else:
            existing_tgl = str(cell_tgl).strip() if cell_tgl else ""

        existing_shift = str(cell_shift).strip().upper() if cell_shift else ""

        match = True
        if target_tgl_str and existing_tgl != target_tgl_str:
            match = False
        if shift_name and existing_shift != shift_name.upper():
            match = False
        if match:
            rows_to_delete.append(row_idx)

    # Hapus dari bawah ke atas agar index tidak bergeser
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    wb.save(SHIFT_FILE)
    wb.close()
    return len(rows_to_delete)


def _read_shift_rows(tanggal_start: dt.date = None, tanggal_end: dt.date = None) -> list:
    """Baca baris jadwal shift dalam rentang tanggal. Return list of dict."""
    if not os.path.exists(SHIFT_FILE):
        return []

    wb = openpyxl.load_workbook(SHIFT_FILE, data_only=True)
    ws = wb.active
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5 or not row[0]:
            continue

        cell_tgl = row[0]
        if isinstance(cell_tgl, dt.datetime):
            tgl = cell_tgl.date()
        elif isinstance(cell_tgl, dt.date):
            tgl = cell_tgl
        else:
            try:
                tgl = dt.datetime.strptime(str(cell_tgl).strip(), "%Y-%m-%d").date()
            except ValueError:
                continue

        if tanggal_start and tgl < tanggal_start:
            continue
        if tanggal_end and tgl > tanggal_end:
            continue

        results.append({
            "tanggal": tgl,
            "jam_mulai": row[1].strftime("%H:%M") if isinstance(row[1], (dt.time, dt.datetime)) else str(row[1]).strip() if row[1] else "",
            "jam_selesai": row[2].strftime("%H:%M") if isinstance(row[2], (dt.time, dt.datetime)) else str(row[2]).strip() if row[2] else "",
            "shift": str(row[3]).strip() if row[3] else "",
            "teknisi": str(row[4]).strip() if row[4] else "",
        })

    wb.close()
    # Sort by tanggal lalu jam_mulai
    results.sort(key=lambda x: (x["tanggal"], x["jam_mulai"]))
    return results


# ==============================================================================
# COMMAND: /lihatshift — Melihat jadwal shift hari ini / minggu ini
# ==============================================================================

@restricted
async def lihatshift_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /lihatshift [hari/minggu/YYYY-MM-DD]

    - /lihatshift           → jadwal hari ini
    - /lihatshift minggu    → jadwal minggu ini (Senin–Minggu)
    - /lihatshift 2026-08-26  → jadwal tanggal tertentu
    """
    now = dt.datetime.now(TZ)
    today = now.date()

    arg = context.args[0].lower().strip() if context.args else "hari"

    if arg in ("hari", "today"):
        start_date = today
        end_date = today
        label = f"Hari Ini ({today.strftime('%A, %d %b %Y')})"
    elif arg in ("minggu", "week", "weekly"):
        # Senin minggu ini
        start_date = today - dt.timedelta(days=today.weekday())
        end_date = start_date + dt.timedelta(days=6)
        label = f"Minggu Ini ({start_date.strftime('%d %b')} – {end_date.strftime('%d %b %Y')})"
    else:
        # Coba parse sebagai tanggal
        try:
            start_date = dt.datetime.strptime(arg, "%Y-%m-%d").date()
            end_date = start_date
            label = f"Tanggal {start_date.strftime('%A, %d %b %Y')}"
        except ValueError:
            await update.message.reply_text(
                "⚠️ Format salah. Gunakan:\n"
                "<code>/lihatshift</code> — hari ini\n"
                "<code>/lihatshift minggu</code> — minggu ini\n"
                "<code>/lihatshift 2026-08-26</code> — tanggal tertentu",
                parse_mode=ParseMode.HTML,
            )
            return

    rows = _sheets_read_shifts(start_date, end_date)

    if not rows:
        await update.message.reply_text(
            f"📅 <b>Jadwal Shift — {label}</b>\n\n"
            "Tidak ada jadwal ditemukan.",
            parse_mode=ParseMode.HTML,
        )
        return

    lines = [f"📅 <b>Jadwal Shift — {label}</b>\n"]
    current_date = None
    for r in rows:
        if r["tanggal"] != current_date:
            current_date = r["tanggal"]
            lines.append(f"\n<b>{current_date.strftime('%A, %d %b %Y')}</b>")
        lines.append(
            f"  • {r['jam_mulai']}–{r['jam_selesai']} | "
            f"<code>{r['shift']}</code> | {r['teknisi']}"
        )

    await update.message.reply_text("\n".join(lines), parse_mode=ParseMode.HTML)


# ==============================================================================
# COMMAND: /tambahshift — Tambah baris ke jadwal tanpa replace file
# ==============================================================================

@restricted
async def tambahshift_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /tambahshift YYYY-MM-DD HH:MM HH:MM SHIFT_X Nama1, Nama2

    Menambah baris baru ke jadwal (tidak menimpa jika sudah ada shift sama di tanggal itu).
    """
    if not context.args or len(context.args) < 5:
        await update.message.reply_text(
            "<b>📝 Tambah Jadwal Shift</b>\n\n"
            "Format:\n"
            "<code>/tambahshift YYYY-MM-DD HH:MM HH:MM SHIFT_X Nama1, Nama2</code>\n\n"
            "Contoh:\n"
            "<code>/tambahshift 2026-08-26 07:00 15:00 SHIFT_1 Bagus, Tri</code>\n"
            "<code>/tambahshift 2026-08-26 15:00 23:00 SHIFT_2 Adelia</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    tanggal_str = context.args[0]
    jam_mulai_str = context.args[1]
    jam_selesai_str = context.args[2]
    shift_name = context.args[3]
    teknisi_str = " ".join(context.args[4:])

    # Validasi
    try:
        tanggal = dt.datetime.strptime(tanggal_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text(
            "⚠️ Format tanggal salah. Gunakan <code>YYYY-MM-DD</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        dt.datetime.strptime(jam_mulai_str, "%H:%M")
        dt.datetime.strptime(jam_selesai_str, "%H:%M")
    except ValueError:
        await update.message.reply_text(
            "⚠️ Format jam salah. Gunakan <code>HH:MM</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    try:
        _sheets_add_shift(tanggal, jam_mulai_str, jam_selesai_str, shift_name, teknisi_str)
        await update.message.reply_text(
            f"✅ <b>Baris jadwal berhasil ditambahkan!</b>\n\n"
            f"📅 Tanggal: <code>{tanggal_str}</code>\n"
            f"⏰ Jam: <code>{jam_mulai_str} - {jam_selesai_str}</code>\n"
            f"🏷 Shift: <code>{shift_name}</code>\n"
            f"👤 Teknisi: <code>{teknisi_str}</code>",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal tambah jadwal:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )


# ==============================================================================
# COMMAND: /hapusshift — Hapus jadwal shift berdasarkan tanggal/shift
# ==============================================================================

@restricted
async def hapusshift_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /hapusshift YYYY-MM-DD [SHIFT_X]

    - /hapusshift 2026-08-26          → hapus SEMUA shift di tanggal tersebut
    - /hapusshift 2026-08-26 SHIFT_1  → hapus hanya SHIFT_1 di tanggal tersebut
    """
    if not context.args:
        await update.message.reply_text(
            "<b>🗑 Hapus Jadwal Shift</b>\n\n"
            "Format:\n"
            "<code>/hapusshift YYYY-MM-DD</code> — hapus semua shift di tanggal tsb\n"
            "<code>/hapusshift YYYY-MM-DD SHIFT_1</code> — hapus shift tertentu saja\n\n"
            "Contoh:\n"
            "<code>/hapusshift 2026-08-26</code>\n"
            "<code>/hapusshift 2026-08-26 SHIFT_2</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    tanggal_str = context.args[0]
    shift_filter = context.args[1] if len(context.args) > 1 else None

    try:
        tanggal = dt.datetime.strptime(tanggal_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text(
            "⚠️ Format tanggal salah. Gunakan <code>YYYY-MM-DD</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    deleted = _sheets_delete_shifts(tanggal, shift_filter)

    if deleted == 0:
        filter_text = f" shift <code>{shift_filter}</code>" if shift_filter else ""
        await update.message.reply_text(
            f"ℹ️ Tidak ada jadwal ditemukan untuk tanggal <code>{tanggal_str}</code>{filter_text}.",
            parse_mode=ParseMode.HTML,
        )
    else:
        filter_text = f" ({shift_filter})" if shift_filter else " (semua shift)"
        await update.message.reply_text(
            f"✅ <b>Berhasil menghapus {deleted} baris jadwal</b>\n\n"
            f"📅 Tanggal: <code>{tanggal_str}</code>{filter_text}",
            parse_mode=ParseMode.HTML,
        )





# ==============================================================================
# COMMAND: /updatetimeshift — Update jam mulai/selesai shift secara massal
# ==============================================================================

@restricted
async def updatetimeshift_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler /updatetimeshift YYYY-MM-DD SHIFT_X HH:MM HH:MM

    Mengubah jam mulai dan jam selesai untuk shift tertentu di tanggal tertentu.
    Contoh: /updatetimeshift 2026-08-27 SHIFT_1 09:00 13:30
    """
    if not context.args or len(context.args) < 4:
        await update.message.reply_text(
            "<b>⏰ Update Waktu Shift</b>\n\n"
            "Format:\n"
            "<code>/updatetimeshift YYYY-MM-DD SHIFT_X HH:MM HH:MM</code>\n\n"
            "Contoh:\n"
            "<code>/updatetimeshift 2026-08-27 SHIFT_1 09:00 13:30</code>\n"
            "<code>/updatetimeshift 2026-08-27 SHIFT_2 14:00 21:00</code>\n\n"
            "Mengubah jam shift di tanggal yang ditentukan.",
            parse_mode=ParseMode.HTML,
        )
        return

    tanggal_str = context.args[0]
    shift_name = context.args[1].upper()
    jam_mulai_str = context.args[2]
    jam_selesai_str = context.args[3]

    # Validasi format tanggal
    try:
        tanggal = dt.datetime.strptime(tanggal_str, "%Y-%m-%d").date()
    except ValueError:
        await update.message.reply_text(
            "⚠️ Format tanggal salah. Gunakan <code>YYYY-MM-DD</code>\n"
            "Contoh: <code>/updatetimeshift 2026-08-27 SHIFT_1 09:00 13:30</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Validasi format jam
    try:
        dt.datetime.strptime(jam_mulai_str, "%H:%M")
        dt.datetime.strptime(jam_selesai_str, "%H:%M")
    except ValueError:
        await update.message.reply_text(
            "⚠️ Format jam salah. Gunakan <code>HH:MM</code>\n"
            "Contoh: <code>/updatetimeshift 2026-08-27 SHIFT_1 09:00 13:30</code>",
            parse_mode=ParseMode.HTML,
        )
        return

    # Update baris dengan tanggal + shift tersebut
    try:
        updated = _sheets_update_time(shift_name, jam_mulai_str, jam_selesai_str, tanggal)
        if updated == 0:
            await update.message.reply_text(
                f"ℹ️ Tidak ada baris dengan shift <code>{shift_name}</code> "
                f"di tanggal <code>{tanggal_str}</code> ditemukan di jadwal.",
                parse_mode=ParseMode.HTML,
            )
        else:
            await update.message.reply_text(
                f"✅ <b>Waktu shift berhasil diupdate!</b>\n\n"
                f"📅 Tanggal: <code>{tanggal_str}</code>\n"
                f"🏷 Shift: <code>{shift_name}</code>\n"
                f"⏰ Jam baru: <code>{jam_mulai_str} - {jam_selesai_str}</code>\n"
                f"📝 Total baris diupdate: <b>{updated}</b>",
                parse_mode=ParseMode.HTML,
            )
    except Exception as e:
        await update.message.reply_text(
            f"⚠️ Gagal update waktu shift:\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )


def _update_time_for_shift(shift_name: str, jam_mulai: str, jam_selesai: str, tanggal: dt.date = None) -> int:
    """Update jam mulai & selesai untuk baris dengan shift (dan tanggal) tertentu. Return jumlah baris terupdate."""
    if not os.path.exists(SHIFT_FILE):
        return 0

    wb = openpyxl.load_workbook(SHIFT_FILE)
    ws = wb.active
    updated = 0
    target_tgl_str = tanggal.strftime("%Y-%m-%d") if tanggal else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell_shift = row[3].value if len(row) > 3 else None
        existing_shift = str(cell_shift).strip().upper() if cell_shift else ""

        if existing_shift != shift_name.upper():
            continue

        # Cek tanggal jika diberikan
        if target_tgl_str:
            cell_tgl = row[0].value
            if isinstance(cell_tgl, (dt.date, dt.datetime)):
                existing_tgl = cell_tgl.strftime("%Y-%m-%d")
            else:
                existing_tgl = str(cell_tgl).strip() if cell_tgl else ""
            if existing_tgl != target_tgl_str:
                continue

        ws.cell(row=row_idx, column=2, value=jam_mulai)
        ws.cell(row=row_idx, column=3, value=jam_selesai)
        updated += 1

    if updated > 0:
        wb.save(SHIFT_FILE)
    wb.close()
    return updated


# ==============================================================================
# COMMAND: /testsheets — Test koneksi ke Google Sheets
# ==============================================================================

@restricted
async def testsheets_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Test koneksi Google Sheets dan append baris percobaan."""
    await update.message.reply_text("⏳ Testing koneksi Google Sheets...")

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds_json = config.GOOGLE_SHEETS_CREDENTIALS
        sheet_id = config.GOOGLE_SHEETS_SPREADSHEET_ID

        if not creds_json:
            await update.message.reply_text("⚠️ <code>GOOGLE_SHEETS_CREDENTIALS</code> env kosong!", parse_mode=ParseMode.HTML)
            return
        if not sheet_id:
            await update.message.reply_text("⚠️ <code>GOOGLE_SHEETS_SPREADSHEET_ID</code> env kosong!", parse_mode=ParseMode.HTML)
            return

        creds_dict = json.loads(creds_json)
        credentials = Credentials.from_service_account_info(
            creds_dict,
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )

        gc = gspread.authorize(credentials)
        spreadsheet = gc.open_by_key(sheet_id)
        worksheet = spreadsheet.sheet1

        # Append test row
        timestamp = dt.datetime.now(TZ).strftime("%Y-%m-%d %H:%M:%S")
        test_row = [timestamp, "TEST", "Test Connection", "test@test.com", "Test", "TEST_SHIFT", "SUCCESS", ""]
        worksheet.append_row(test_row, value_input_option="USER_ENTERED")

        await update.message.reply_text(
            f"✅ <b>Koneksi Google Sheets berhasil!</b>\n\n"
            f"Spreadsheet: <code>{spreadsheet.title}</code>\n"
            f"Sheet: <code>{worksheet.title}</code>\n"
            f"Test row berhasil ditulis pada {timestamp}",
            parse_mode=ParseMode.HTML,
        )

    except json.JSONDecodeError as e:
        await update.message.reply_text(
            f"⚠️ <b>GOOGLE_SHEETS_CREDENTIALS bukan JSON valid:</b>\n<code>{html.escape(str(e))}</code>",
            parse_mode=ParseMode.HTML,
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        error_msg = f"{type(e).__name__}: {str(e)}\n\n{tb[-1500:]}"
        await update.message.reply_text(
            f"⚠️ <b>Gagal koneksi Google Sheets:</b>\n<pre>{html.escape(error_msg)}</pre>",
            parse_mode=ParseMode.HTML,
        )


async def push_error_daily_job(context: ContextTypes.DEFAULT_TYPE):
    """Daily job: jalankan push_error.sql jam 17:00 dan kirim hasilnya."""
    try:
        results = await asyncio.to_thread(_run_botcore_query, ["push_error.sql"])
    except Exception as e:
        logger.error(f"Push Error daily job error: {e}")
        return

    for q_file, excel_path, row_count, headers, data_list in results:
        if row_count == 0:
            await _broadcast_notify(context, "✅ <b>Push Error Report (17:00)</b>\n\nTidak ada error hari ini.")
        else:
            # Kirim sebagai text jika sedikit, excel jika banyak
            if row_count <= 30:
                text_lines = [
                    f"⚠️ <b>Push Error Report (17:00)</b>",
                    f"Total Error: {row_count}",
                    "",
                ]
                for idx, row in enumerate(data_list[:30], 1):
                    text_lines.append(f"<b>#{idx}</b>")
                    for col_idx, cell in enumerate(row):
                        col_name = headers[col_idx] if col_idx < len(headers) else f"col{col_idx}"
                        text_lines.append(f"• {col_name}: {cell}")
                    text_lines.append("──────────────────────────────")

                text_output = "\n".join(text_lines)
                if len(text_output) > 3800:
                    text_output = text_output[:3800] + "\n\n... (dipotong)"
                await _broadcast_notify(context, text_output)
            else:
                # Kirim sebagai Excel
                if os.path.exists(excel_path):
                    caption = f"⚠️ Push Error Report (17:00)\nTotal Error: {row_count}"
                    for chat_id, thread_id in config.notify_targets():
                        try:
                            with open(excel_path, "rb") as f:
                                await context.bot.send_document(
                                    chat_id=chat_id,
                                    document=f,
                                    filename=os.path.basename(excel_path),
                                    caption=caption,
                                    message_thread_id=thread_id,
                                    read_timeout=120,
                                    write_timeout=120,
                                )
                        except Exception as e:
                            logger.error(f"Gagal kirim push_error excel: {e}")

        # Cleanup
        if excel_path and os.path.exists(excel_path):
            try:
                os.remove(excel_path)
            except OSError:
                pass

    if results:
        logger.info(f"Push Error daily job: {results[0][2]} error ditemukan.")


# ---------------- main ----------------

def main():
    config.validate()

    try:
        asyncio.get_event_loop()
    except RuntimeError:
        asyncio.set_event_loop(asyncio.new_event_loop())

    app = (
        Application.builder()
        .token(config.TELEGRAM_BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .pool_timeout(30)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("chatid", chatid_command))

    myjira_conv = ConversationHandler(
        entry_points=[CommandHandler("myjira", myjira_command)],
        states={
            MYJIRA_EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, myjira_email)],
            MYJIRA_TOKEN: [MessageHandler(filters.TEXT & ~filters.COMMAND, myjira_token)],
            MYJIRA_CONFIRM: [CallbackQueryHandler(myjira_confirm, pattern="^myjira_save_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(myjira_conv)
    app.add_handler(CommandHandler("today", today_summary))
    app.add_handler(CommandHandler("week", week_summary))

    log_conv = ConversationHandler(
        entry_points=[CommandHandler("log", log_start)],
        states={
            LOG_ISSUE: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_issue)],
            LOG_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_time)],
            LOG_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, log_desc)],
            LOG_DATE: [
                CallbackQueryHandler(log_date_choice, pattern="^date_"),
                MessageHandler(filters.TEXT & ~filters.COMMAND, log_date_manual),
            ],
            LOG_CONFIRM: [CallbackQueryHandler(log_confirm, pattern="^confirm_")],
            LOG_BULK_CONFIRM: [CallbackQueryHandler(bulk_log_confirm_callback, pattern="^bulk_confirm_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(log_conv)

    edit_conv = ConversationHandler(
        entry_points=[CommandHandler("edit", edit_start)],
        states={
            PICK_ISSUE_FOR_EDIT: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_pick_issue)],
            PICK_WORKLOG_EDIT: [CallbackQueryHandler(edit_pick_worklog, pattern="^wl_")],
            EDIT_TIME: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_time)],
            EDIT_DESC: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_desc)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(edit_conv)

    delete_conv = ConversationHandler(
        entry_points=[CommandHandler("delete", delete_start)],
        states={
            PICK_ISSUE_FOR_DELETE: [MessageHandler(filters.TEXT & ~filters.COMMAND, delete_pick_issue)],
            PICK_WORKLOG_DELETE: [CallbackQueryHandler(delete_pick_worklog, pattern="^wl_")],
            CONFIRM_DELETE: [CallbackQueryHandler(delete_confirm, pattern="^del_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(delete_conv)
    app.add_handler(CommandHandler("tasks", list_tasks))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("tools", tools_command))
    app.add_handler(CallbackQueryHandler(tools_callback, pattern="^tools_"))
    app.add_handler(CommandHandler("sdtickets", sdtickets_command))
    app.add_handler(CommandHandler("sdticket", sdticket_command))
    app.add_handler(CommandHandler("logactivity", log_activity_command))
    app.add_handler(CommandHandler("sdreminder", sdreminder_command))
    app.add_handler(
        MessageHandler(
            filters.TEXT & ~filters.COMMAND & filters.Regex(r"^([A-Za-z]{2,10}-|[A-Z]{2,10})$"),
            list_tasks,
        )
    )

    app.add_handler(CommandHandler("guide", guide_command))
    app.add_handler(CommandHandler("listguide", list_guide))
    app.add_handler(CallbackQueryHandler(guide_pick_callback, pattern="^gid_"))

    addguide_conv = ConversationHandler(
        entry_points=[CommandHandler("addguide", addguide_start)],
        states={
            ADD_GUIDE_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_title)],
            ADD_GUIDE_KEYWORDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_keywords)],
            ADD_GUIDE_CONTENT: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, addguide_content
                )
            ],
            ADD_GUIDE_ACTION_ASK: [CallbackQueryHandler(addguide_action_ask, pattern="^ga_")],
            ADD_GUIDE_ACTION_SCRIPT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_action_script)
            ],
            ADD_GUIDE_ACTION_FLAG: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, addguide_action_flag)
            ],
            ADD_GUIDE_ACTION_TYPE: [CallbackQueryHandler(addguide_action_type, pattern="^atype_")],
            ADD_GUIDE_ACTION_MODE: [CallbackQueryHandler(addguide_action_mode, pattern="^amode_")],
            ADD_GUIDE_CONFIRM: [CallbackQueryHandler(addguide_confirm, pattern="^guide_save_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(addguide_conv)

    delguide_conv = ConversationHandler(
        entry_points=[CommandHandler("delguide", delguide_start)],
        states={
            DEL_GUIDE_PICK: [CallbackQueryHandler(delguide_pick, pattern="^delg_")],
            DEL_GUIDE_CONFIRM: [CallbackQueryHandler(delguide_confirm, pattern="^delg_(yes|no)$")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(delguide_conv)

    editguide_conv = ConversationHandler(
        entry_points=[CommandHandler("editguide", editguide_start)],
        states={
            EDIT_GUIDE_PICK: [CallbackQueryHandler(editguide_pick, pattern="^editg_")],
            EDIT_GUIDE_MENU: [CallbackQueryHandler(editguide_menu_choice, pattern="^ef_")],
            EDIT_GUIDE_TITLE: [MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_title)],
            EDIT_GUIDE_KEYWORDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_keywords)],
            EDIT_GUIDE_CONTENT: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, editguide_content
                )
            ],
            EDIT_GUIDE_ACTION_SCRIPT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_action_script)
            ],
            EDIT_GUIDE_ACTION_FLAG: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, editguide_action_flag)
            ],
            EDIT_GUIDE_ACTION_TYPE: [
                CallbackQueryHandler(editguide_action_type, pattern="^etype_")
            ],
            EDIT_GUIDE_ACTION_MODE: [
                CallbackQueryHandler(editguide_action_mode, pattern="^emode_")
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(editguide_conv)

    run_action_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(run_action_start, pattern="^runaction_"),
            CommandHandler("run", run_command),
        ],
        states={
            RUN_ACTION_PARAM: [
                MessageHandler(
                    (filters.TEXT & ~filters.COMMAND) | filters.Document.ALL, run_action_param
                )
            ],
            RUN_ACTION_CONFIRM: [
                CallbackQueryHandler(run_action_confirm, pattern="^runconfirm_")
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(run_action_conv)

    # ConversationHandler delete reservation
    delres_conv = ConversationHandler(
        entry_points=[CommandHandler("delreservation", delreservation_start)],
        states={
            DELRES_INPUT: [
                MessageHandler(filters.Document.ALL, delreservation_input),
                MessageHandler(filters.TEXT & ~filters.COMMAND, delreservation_input),
            ],
            DELRES_CONFIRM: [CallbackQueryHandler(delreservation_confirm, pattern="^delres_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(delres_conv)

    # ConversationHandler release voucher
    relvoucher_conv = ConversationHandler(
        entry_points=[CommandHandler("releasevoucher", releasevoucher_start)],
        states={
            RELVOUCHER_INPUT: [
                MessageHandler(filters.Document.ALL, releasevoucher_input),
                MessageHandler(filters.TEXT & ~filters.COMMAND, releasevoucher_input),
            ],
            RELVOUCHER_CONFIRM: [CallbackQueryHandler(releasevoucher_confirm, pattern="^relvoucher_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(relvoucher_conv)

    # ConversationHandler check stock
    checkstock_conv = ConversationHandler(
        entry_points=[CommandHandler("stock", checkstock_start)],
        states={
            CHECKSTOCK_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, checkstock_input),
            ],
            CHECKSTOCK_CONFIRM: [CallbackQueryHandler(checkstock_confirm, pattern="^checkstock_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(checkstock_conv)
    app.add_handler(CallbackQueryHandler(checkstock_confirm, pattern="^checkstock_"))
    app.add_handler(CallbackQueryHandler(delreservation_confirm, pattern="^delres_"))

    # ConversationHandler cek promo
    cekpromo_conv = ConversationHandler(
        entry_points=[CommandHandler("promo", cekpromo_start)],
        states={
            CEKPROMO_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, cekpromo_input),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(cekpromo_conv)

    # ConversationHandler cek AWB
    cekawb_conv = ConversationHandler(
        entry_points=[CommandHandler("awb", cekawb_start)],
        states={
            CEKAWB_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, cekawb_input),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(cekawb_conv)

    # ConversationHandler AWB JNE
    awbjne_conv = ConversationHandler(
        entry_points=[CommandHandler("awbjne", awbjne_start)],
        states={
            AWBJNE_INPUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, awbjne_input),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(awbjne_conv)

    # ConversationHandler update shift
    updateshift_conv = ConversationHandler(
        entry_points=[CommandHandler("updateshift", updateshift_start)],
        states={
            UPDATESHIFT_UPLOAD: [
                MessageHandler(filters.Document.ALL, updateshift_upload),
                MessageHandler(filters.TEXT & ~filters.COMMAND, updateshift_upload),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        per_message=False,
    )
    app.add_handler(updateshift_conv)

    # Command shift management (tanpa ConversationHandler, langsung execute)
    app.add_handler(CommandHandler("lihatshift", lihatshift_command))
    app.add_handler(CommandHandler("tambahshift", tambahshift_command))
    app.add_handler(CommandHandler("hapusshift", hapusshift_command))
    app.add_handler(CommandHandler("updatetimeshift", updatetimeshift_command))
    app.add_handler(CommandHandler("testsheets", testsheets_command))
    app.add_handler(CommandHandler("ticketupdate", ticketupdate_command))
    app.add_handler(CommandHandler("overdue", overdue_command))
    app.add_handler(CommandHandler("quote", quote_command))
    app.add_handler(CommandHandler("addquote", addquote_command))
    app.add_handler(CommandHandler("seedquotes", seedquotes_command))
    app.add_handler(CommandHandler("coffee", coffee_command))
    app.add_handler(CommandHandler("testsearch", testsearch_command))

    # GrabExpress tracking
    app.add_handler(CommandHandler("lacakgrab", lacakgrab_command))
    app.add_handler(CommandHandler("trackgrab", lacakgrab_command))

    # Eraspace order tracking (dumpdo)
    app.add_handler(CommandHandler("cekorder", cekorder_command))

    # Knowledge Base / Bank Data commands
    app.add_handler(CommandHandler("tanyabot", tanya_command))
    app.add_handler(CommandHandler("addfaq", addfaq_command))
    app.add_handler(CommandHandler("listfaq", listfaq_command))
    app.add_handler(CommandHandler("delfaq", delfaq_command))
    app.add_handler(CommandHandler("pending", pending_command))
    app.add_handler(CommandHandler("answerfaq", answerfaq_command))

    # Command /query (manual trigger)
    app.add_handler(CommandHandler("query", query_command))
    app.add_handler(CallbackQueryHandler(query_pick_callback, pattern="^qrun_"))
    app.add_handler(CallbackQueryHandler(query_output_callback, pattern="^qout_"))

    # Handler natural language "cek tiket #12356" -- harus SEBELUM guide_free_text
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND & filters.Regex(CEK_TIKET_RE),
        cek_tiket_natural,
    ))

    # Handler natural text untuk tools (deteksi format input tanpa command)
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND & filters.Regex(NATURAL_TOOLS_RE),
        natural_tools_handler,
    ))

    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, guide_free_text))
    app.add_error_handler(error_handler)

    # --------------------------------------------------------------------------
    # BACKGROUND JOBS SERVICEDESK PLUS & REMINDERS
    # --------------------------------------------------------------------------
    if sdp and config.SDP_NOTIFY_GROUPS:
        # Job Otomasi Assign Tiket ServiceDesk Plus (Berjalan tiap 5 menit)
        app.job_queue.run_repeating(
            auto_assign_sdp_tickets,
            interval=dt.timedelta(minutes=5),
            first=10,
        )
        logger.info(
            f"Otomasi Auto-Assign Tiket SDP aktif (tiap 5 menit) untuk grup: {', '.join(config.SDP_NOTIFY_GROUPS)}"
        )

        # Job Cek Tiket Baru
        app.job_queue.run_repeating(
            check_new_sdp_tickets,
            interval=dt.timedelta(minutes=config.SDP_NOTIFY_INTERVAL_MINUTES),
            first=15,
        )
        logger.info(
            f"Notifikasi tiket baru SDP aktif tiap {config.SDP_NOTIFY_INTERVAL_MINUTES} menit "
            f"untuk group: {', '.join(config.SDP_NOTIFY_GROUPS)}"
        )

        # Job Reminder Tiket Open
        app.job_queue.run_repeating(
            check_open_ticket_reminders,
            interval=dt.timedelta(minutes=1),
            first=30,
        )

        # Job Monitor SLA / OverDue
        if config.SDP_SLA_MONITOR_ENABLED:
            app.job_queue.run_repeating(
                check_sla_overdue_tickets,
                interval=dt.timedelta(minutes=config.SDP_SLA_CHECK_INTERVAL_MINUTES),
                first=40,
            )
            logger.info(
                f"SLA OverDue Monitor aktif tiap {config.SDP_SLA_CHECK_INTERVAL_MINUTES} menit "
                f"untuk group: {', '.join(config.SDP_NOTIFY_GROUPS)}"
            )

    if config.REMINDER_INTERVAL_MINUTES > 0:
        app.job_queue.run_repeating(
            interval_reminder,
            interval=dt.timedelta(minutes=config.REMINDER_INTERVAL_MINUTES),
            first=10,
        )
        logger.info(
            f"Reminder aktif tiap {config.REMINDER_INTERVAL_MINUTES} menit, "
            f"jam {config.REMINDER_START_HOUR}:00-{config.REMINDER_END_HOUR}:00"
        )
    else:
        app.job_queue.run_daily(
            daily_reminder,
            time=dt.time(hour=config.REMINDER_HOUR, minute=config.REMINDER_MINUTE, tzinfo=TZ),
        )
        logger.info(f"Reminder aktif tiap hari jam {config.REMINDER_HOUR}:{config.REMINDER_MINUTE:02d}")

    # Job SLA Monitor (via SDP API)
    if config.SLA_MONITOR_ENABLED:
        app.job_queue.run_repeating(
            sla_monitor_job,
            interval=dt.timedelta(minutes=config.SLA_MONITOR_INTERVAL_MINUTES),
            first=20,
        )
        logger.info(
            f"SLA Monitor aktif tiap {config.SLA_MONITOR_INTERVAL_MINUTES} menit "
            f"(Groups: {', '.join(config.SDP_NOTIFY_GROUPS)})"
        )

    # Job Query Database (dari bot_core scheduler)
    query_interval = int(os.getenv("QUERY_SCHEDULER_HOURS", "0"))
    if query_interval > 0:
        app.job_queue.run_repeating(
            query_scheduler_job,
            interval=dt.timedelta(hours=query_interval),
            first=60,
        )
        logger.info(f"Query scheduler aktif tiap {query_interval} jam")

    # Job Daily push_error.sql jam 17:00
    app.job_queue.run_daily(
        push_error_daily_job,
        time=dt.time(hour=17, minute=0, tzinfo=TZ),
    )
    logger.info("Push Error query aktif setiap hari jam 17:00")

    logger.info("Bot mulai berjalan...")
    app.run_polling()


if __name__ == "__main__":
    main()